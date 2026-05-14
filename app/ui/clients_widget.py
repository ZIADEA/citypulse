"""
clients_widget.py — Gestion complète des clients CityPulse Logistics
=====================================================================
• Table paginée (100/page) avec SearchBar + CollapsibleSection filtres
• Dialogue 5 onglets : Général, Adresse, Livraison, Contact, Historique
• Import CSV/Excel avec preview 5 lignes + mapping colonnes + géocodage lot
• Export CSV / Excel / JSON
• Détection anomalies (z-score) sur données clients
• Vue Carte Leaflet (QWebEngineView si disponible)
• Opérations en lot : archiver, exporter sélection
• Double-clic → édition, clic droit → menu contextuel
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import csv
import json as _json
import logging
import os
import time

# ── PyQt6 ────────────────────────────────────────────────────────────────────
from PyQt6.QtWidgets import (
  QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
  QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit,
  QComboBox, QDialog, QFormLayout, QDoubleSpinBox, QSpinBox,
  QTextEdit, QMessageBox, QFileDialog, QFrame, QStackedWidget,
  QTabWidget, QSlider, QCheckBox, QMenu, QScrollArea,
  QSizePolicy, QAbstractItemView, QProgressBar,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QPoint
from PyQt6.QtGui import QFont, QColor, QCursor, QAction

# ── Local ─────────────────────────────────────────────────────────────────────
from ..database.db_manager import get_connection, log_action
from ..services.django_sync_service import get_django_service
from .help_dialog import show_help
from .toast import show_toast
from .loading_overlay import LoadingOverlay
from .import_dialog import _read_headers_and_preview
from .components import (
  SectionHeader, SearchBar, PaginationBar,
  CollapsibleSection, ConfirmDialog, EmptyState,
)
from .components.confirm_dialog import _dialog_qss
from .lucide_icons import apply_action_button

from .webengine_support import HAS_WEB, QWebEngineView, WEBENGINE_FALLBACK_SHORT

try:
  import requests as _requests
  HAS_REQUESTS = True
except ImportError:
  HAS_REQUESTS = False

try:
  import openpyxl as _openpyxl
  HAS_OPENPYXL = True
except ImportError:
  HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
C = {
  "bg":   "#0D1B2A", "panel":  "#112240", "input":  "#1A2E4A",
  "accent": "#00D4FF", "success": "#00FF88", "warning": "#FFB800",
  "danger": "#FF4757", "text":  "#E8F4FD", "text2":  "#8899AA",
  "border": "#1E3A5F", "hover":  "#1A3A5C",
}
_TYPE_COLORS = {
  "supermarche": "#00D4FF", "restaurant": "#FF6B6B",
  "bureau":   "#96CEB4", "pharmacie": "#C3A6FF",
  "particulier": "#FFD93D", "standard":  "#8899AA",
  "prioritaire": "#FFB800", "occasionnel":"#6C757D",
  "demo":    "#4A5568",
}
_CLIENT_TYPES = [
  "standard", "supermarche", "restaurant", "bureau",
  "pharmacie", "particulier", "prioritaire", "occasionnel", "demo",
]

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _priority_stars(p) -> str:
  try:
    p = max(1, min(5, int(p)))
  except (TypeError, ValueError):
    p = 3
  stars = 6 - p
  return "*" * stars + "-" * (5 - stars)

def _import_parse_float(val, default: float = 0.0) -> tuple[float, bool]:
  """
  Parse une valeur CSV/Excel en float pour l'import clients.
  Retourne (valeur, True) si la cellule était non vide mais non numérique.
  """
  if val in (None, "", "None"):
    return float(default), False
  if isinstance(val, (int, float)):
    return float(val), False
  s = str(val).strip().replace("\u00a0", "").replace(" ", "")
  if not s:
    return float(default), False
  try:
    return float(s.replace(",", ".")), False
  except (ValueError, TypeError):
    return float(default), True


def _import_coerce_int(val, default: int, lo: int, hi: int) -> tuple[int, bool]:
  """Entier borné pour ready_time, due_time, service_time, etc. (min depuis minuit)."""
  f, bad = _import_parse_float(val, float(default))
  try:
    iv = int(round(f))
  except (ValueError, OverflowError):
    return max(lo, min(hi, int(default))), True
  return max(lo, min(hi, iv)), bad


_IMPORT_NUMERIC_FIELDS = frozenset({
  "latitude", "longitude", "demand_kg", "demand_m3",
  "ready_time", "due_time", "service_time", "priority",
})


def _min_to_hhmm(minutes) -> str:
  try:
    h, m = divmod(int(minutes or 0), 60)
    return f"{h:02d}:{m:02d}"
  except Exception:
    return "--:--"

def _hhmm_to_min(hhmm: str) -> int:
  try:
    parts = str(hhmm).strip().split(":")
    return int(parts[0]) * 60 + int(parts[1])
  except Exception:
    return 0

def _type_color(ctype: str) -> str:
  return _TYPE_COLORS.get((ctype or "standard").lower(), C["text2"])

# ═══════════════════════════════════════════════════════════════════════════════
# THREADS
# ═══════════════════════════════════════════════════════════════════════════════

class _GeocoderThread(QThread):
  """Nominatim geocoder (single address)."""
  result = pyqtSignal(float, float, str)
  error = pyqtSignal(str)

  def __init__(self, address: str, parent=None):
    super().__init__(parent)
    self.address = address

  def run(self):
    if not HAS_REQUESTS:
      self.error.emit("Module 'requests' non disponible.")
      return
    try:
      resp = _requests.get(
        "https://nominatim.openstreetmap.org/search",
        params={"q": self.address, "format": "json", "limit": 1},
        headers={"User-Agent": "CityPulse/5.3"},
        timeout=10,
      )
      data = resp.json()
      if data:
        self.result.emit(float(data[0]["lat"]), float(data[0]["lon"]),
                 data[0].get("display_name", self.address))
      else:
        self.error.emit(f"Adresse introuvable : {self.address[:60]}")
    except Exception as e:
      self.error.emit(str(e))


class _ImportThread(QThread):
  """Import CSV/Excel → DB avec géocodage lot optionnel."""
  progress = pyqtSignal(str, int, int)
  finished = pyqtSignal(dict)

  def __init__(self, filepath: str, col_map: dict,
         geocode: bool = False, parent=None):
    super().__init__(parent)
    self.filepath = filepath
    self.col_map = col_map
    self.geocode = geocode

  def run(self):
    created = 0
    updated = 0
    error_list: list[str] = []
    rows: list[dict] = []

    # ── Lecture fichier ────────────────────────────────────────────
    try:
      ext = os.path.splitext(self.filepath)[1].lower()
      if ext in (".xls", ".xlsx"):
        if not HAS_OPENPYXL:
          self.finished.emit({"created": 0, "updated": 0,
                    "errors": 1, "error_list": ["openpyxl manquant"]})
          return
        wb = _openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else f"col{i}"
              for i, h in enumerate(next(it, []))]
        for vals in it:
          rows.append(dict(zip(headers, vals)))
        wb.close()
      else:
        with open(self.filepath, "r", encoding="utf-8-sig") as f:
          rows = list(csv.DictReader(f))
    except Exception as e:
      self.finished.emit({"created": 0, "updated": 0,
                "errors": 1, "error_list": [str(e)]})
      return

    total = len(rows)
    conn = get_connection()

    def _v(row, field, default=None):
      col = self.col_map.get(field)
      if col and col in row:
        v = row[col]
        if v not in (None, "", "None"):
          return v
      return default

    for i, row in enumerate(rows):
      self.progress.emit(f"Ligne {i + 1}/{total}…", i + 1, total)
      try:
        name = str(_v(row, "name", f"Client {i + 1}")).strip()
        if not name:
          error_list.append(f"Ligne {i + 1}: nom vide — ignorée")
          continue

        lat, bad_lat = _import_parse_float(_v(row, "latitude", None), 0.0)
        lon, bad_lon = _import_parse_float(_v(row, "longitude", None), 0.0)
        if bad_lat:
          error_list.append(
            f"Ligne {i + 1}: latitude non numérique — ignorée (0)")
          lat = 0.0
        if bad_lon:
          error_list.append(
            f"Ligne {i + 1}: longitude non numérique — ignorée (0)")
          lon = 0.0

        # Géocodage si coords manquantes
        if self.geocode and lat == 0 and lon == 0:
          addr = str(_v(row, "address", "") or "")
          if addr and HAS_REQUESTS:
            try:
              r = _requests.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": addr, "format": "json", "limit": 1},
                headers={"User-Agent": "CityPulse/5.3"}, timeout=8,
              )
              geo = r.json()
              if geo:
                lat = float(geo[0]["lat"])
                lon = float(geo[0]["lon"])
              time.sleep(1.1)
            except Exception:
              pass

        demand, bad_d = _import_parse_float(_v(row, "demand_kg", None), 0.0)
        if bad_d:
          error_list.append(
            f"Ligne {i + 1}: demande (kg) non numérique — 0 kg")
        demand_m3, bad_m3 = _import_parse_float(
          _v(row, "demand_m3", None), 0.0)
        if bad_m3:
          error_list.append(
            f"Ligne {i + 1}: volume (m³) non numérique — 0")
        ready, bad_r = _import_coerce_int(_v(row, "ready_time", None), 0, 0, 2880)
        if bad_r:
          error_list.append(
            f"Ligne {i + 1}: créneau début (min) non numérique — 0")
        due, bad_u = _import_coerce_int(
          _v(row, "due_time", None), 1440, 0, 2880)
        if bad_u:
          error_list.append(
            f"Ligne {i + 1}: créneau fin (min) non numérique — défaut")
        service, bad_s = _import_coerce_int(
          _v(row, "service_time", None), 10, 0, 1440)
        if bad_s:
          error_list.append(
            f"Ligne {i + 1}: durée visite non numérique — 10 min")
        prio_f, bad_p = _import_parse_float(_v(row, "priority", None), 3.0)
        prio = max(1, min(5, int(round(prio_f))))
        if bad_p:
          error_list.append(
            f"Ligne {i + 1}: priorité non numérique — 3")
        ctype  = str(_v(row, "client_type", "standard") or "standard")
        address = str(_v(row, "address",   "") or "")
        company = str(_v(row, "company_name", "") or "")
        phone  = str(_v(row, "phone",    "") or "")
        email  = str(_v(row, "email",    "") or "")
        tags  = str(_v(row, "tags",     "") or "")

        if ready >= due:
          due = ready + 240

        existing = conn.execute(
          "SELECT id FROM clients WHERE name= ? AND archived=0 LIMIT 1", (name,)
        ).fetchone()

        if existing:
          cid = existing[0]
          conn.execute("""
            UPDATE clients SET address= ?,latitude= ?,longitude= ?,
            demand_kg= ?,demand_m3= ?,ready_time= ?,due_time= ?,service_time= ?,
            priority= ?,client_type= ?,phone= ?,email= ?,
            updated_at=datetime('now')
            WHERE id=?
          """, (address, lat, lon, demand, demand_m3, ready, due, service,
             prio, ctype, phone, email, cid))
          for col, val in [("company_name", company), ("tags", tags)]:
            try:
              conn.execute(f"UPDATE clients SET {col}= ? WHERE id= ?", (val, cid))
            except Exception:
              pass
          updated += 1
        else:
          cur = conn.execute("""
            INSERT INTO clients
            (name,address,latitude,longitude,demand_kg,demand_m3,
             ready_time,due_time,service_time,priority,client_type,
             phone,email,archived,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,0,datetime('now'))
          """, (name, address, lat, lon, demand, demand_m3, ready, due,
             service, prio, ctype, phone, email))
          cid = cur.lastrowid
          for col, val in [("company_name", company), ("tags", tags)]:
            try:
              conn.execute(f"UPDATE clients SET {col}= ? WHERE id= ?", (val, cid))
            except Exception:
              pass
          created += 1

      except Exception as e:
        error_list.append(f"Ligne {i + 1}: {str(e)[:80]}")

    conn.commit()
    conn.close()
    log_action("CLIENT_IMPORT",
          f"{created} créés, {updated} MàJ, {len(error_list)} erreurs")
    self.finished.emit({
      "created": created, "updated": updated,
      "errors": len(error_list), "error_list": error_list,
    })


class _AnomalyThread(QThread):
  """Détection anomalies dans les données clients (z-score)."""
  finished = pyqtSignal(list)

  def run(self):
    results = []
    try:
      import numpy as _np
      conn = get_connection()
      rows = conn.execute("""
        SELECT id, name, latitude, longitude, demand_kg,
            service_time, ready_time, due_time
        FROM clients WHERE archived=0
      """).fetchall()
      conn.close()

      if len(rows) < 3:
        self.finished.emit([])
        return

      demands = _np.array([float(r["demand_kg"]  or 0) for r in rows])
      services = _np.array([float(r["service_time"] or 0) for r in rows])

      d_m, d_s = demands.mean(), max(demands.std(), 0.001)
      s_m, s_s = services.mean(), max(services.std(), 0.001)

      for r in rows:
        issues = []
        lat = float(r["latitude"] or 0)
        lon = float(r["longitude"] or 0)

        if lat == 0.0 and lon == 0.0:
          issues.append(("high", "Coordonnées (0,0) — client non localisé"))
        if (r["demand_kg"] or 0) < 0:
          issues.append(("high",
                  f"Demande négative : {r['demand_kg']} kg"))
        if (r["demand_kg"] or 0) == 0:
          issues.append(("low", "Demande nulle — livraison sans poids"))

        dz = abs((float(r["demand_kg"] or 0) - d_m) / d_s)
        if dz > 2.5:
          issues.append(("medium",
                  f"Demande anormale : {r['demand_kg']:.1f} kg (z={dz:.1f})"))

        sz = abs((float(r["service_time"] or 0) - s_m) / s_s)
        if sz > 2.5:
          issues.append(("medium",
                  f"Durée service anormale : {r['service_time']} min (z={sz:.1f})"))

        rt = r["ready_time"] or 0
        dt = r["due_time"] or 1440
        if rt >= dt:
          issues.append(("high",
                  f"Créneau inversé : début {rt} >= fin {dt}"))

        if issues:
          results.append({
            "id": r["id"], "name": r["name"], "issues": issues,
          })
    except ImportError:
      # numpy non disponible : analyse basique
      conn = get_connection()
      rows = conn.execute(
        "SELECT id, name, latitude, longitude, demand_kg,"
        " service_time, ready_time, due_time"
        " FROM clients WHERE archived=0"
      ).fetchall()
      conn.close()
      for r in rows:
        issues = []
        if float(r["latitude"] or 0) == 0 and float(r["longitude"] or 0) == 0:
          issues.append(("high", "Coordonnées (0,0)"))
        if (r["ready_time"] or 0) >= (r["due_time"] or 1440):
          issues.append(("high", "Créneau inversé"))
        if issues:
          results.append({"id": r["id"], "name": r["name"], "issues": issues})
    except Exception:
      logger.exception("_AnomalyThread error")

    self.finished.emit(results)


# ═══════════════════════════════════════════════════════════════════════════════
# WIDGETS AUXILIAIRES
# ═══════════════════════════════════════════════════════════════════════════════

class _TagsInput(QWidget):
  """Saisie de tags sous forme de chips + champ de texte."""

  def __init__(self, initial: str = "", parent=None):
    super().__init__(parent)
    self._tags: list[str] = [t.strip() for t in (initial or "").split(",") if t.strip()]
    self._setup()

  def _setup(self):
    lo = QVBoxLayout(self)
    lo.setContentsMargins(0, 0, 0, 0)
    lo.setSpacing(4)

    self._chips_frame = QFrame()
    self._chips_frame.setMinimumHeight(32)
    self._chips_frame.setStyleSheet(
      f"QFrame{{background:{C['input']};border:1px solid {C['border']};border-radius:6px;padding:4px;}}"
    )
    self._chips_lo = QHBoxLayout(self._chips_frame)
    self._chips_lo.setContentsMargins(4, 2, 4, 2)
    self._chips_lo.setSpacing(4)
    self._chips_lo.addStretch()
    lo.addWidget(self._chips_frame)

    row = QHBoxLayout()
    self._input = QLineEdit()
    self._input.setPlaceholderText("Ajouter tag… (Entrée)")
    self._input.setStyleSheet(
      f"QLineEdit{{background:{C['input']};color:{C['text']};border:1px solid {C['border']};"
      "border-radius:5px;padding:4px 8px;font-size:12px;}}"
    )
    self._input.returnPressed.connect(self._add)
    row.addWidget(self._input)
    lo.addLayout(row)
    self._refresh()

  def _refresh(self):
    for i in reversed(range(self._chips_lo.count() - 1)):
      w = self._chips_lo.itemAt(i).widget()
      if w:
        w.deleteLater()
    for tag in self._tags:
      chip = QFrame()
      chip.setStyleSheet(
        f"QFrame{{background:{C['accent']};border-radius:10px;padding:0;}}"
      )
      cl = QHBoxLayout(chip)
      cl.setContentsMargins(8, 2, 4, 2)
      cl.setSpacing(4)
      lbl = QLabel(tag)
      lbl.setStyleSheet(f"color:{C['bg']};font-size:11px;font-weight:600;background:transparent;")
      rm = QPushButton("×")
      rm.setToolTip("Retirer ce tag")
      rm.setFixedSize(14, 14)
      rm.setStyleSheet(
        f"QPushButton{{background:transparent;border:none;color:{C['bg']};font-size:13px;}}"
        f"QPushButton:hover{{color:{C['danger']};}}"
      )
      rm.clicked.connect(lambda _, t=tag: self._remove(t))
      cl.addWidget(lbl)
      cl.addWidget(rm)
      self._chips_lo.insertWidget(self._chips_lo.count() - 1, chip)

  def _add(self):
    t = self._input.text().strip()
    if t and t not in self._tags:
      self._tags.append(t)
      self._refresh()
    self._input.clear()

  def _remove(self, tag: str):
    self._tags = [t for t in self._tags if t != tag]
    self._refresh()

  def get_tags(self) -> str:
    return ",".join(self._tags)


# ── Minimap HTML ───────────────────────────────────────────────────────────────
def _minimap_html(lat: float, lon: float, name: str = "") -> str:
  safe = (name or "").replace("'", "\\'")[:50]
  return (
    f"<!DOCTYPE html><html><head><meta charset='utf-8'/>"
    "<link rel='stylesheet' href='https://unpkg.com/leaflet@1.9.4/dist/leaflet.css'/>"
    "<script src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js'></script>"
    "<style>html,body,#map{width:100%;height:100%;margin:0;padding:0;}</style>"
    f"</head><body><div id='map'></div><script>"
    f"var m=L.map('map',{{zoomControl:true}}).setView([{lat},{lon}],15);"
    "L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png').addTo(m);"
    f"L.marker([{lat},{lon}]).addTo(m).bindPopup('{safe}').openPopup();"
    "</script></body></html>"
  )


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT DIALOG (5 onglets)
# ═══════════════════════════════════════════════════════════════════════════════

class _ClientDialog(QDialog):

  def __init__(self, parent=None, client: dict = None):
    super().__init__(parent)
    self.client = client or {}
    self.setWindowTitle("Modifier client" if client else "Nouveau client")
    self.setMinimumSize(660, 540)
    self.resize(720, 570)
    self.setModal(True)
    self._geocoder: _GeocoderThread | None = None
    self.setStyleSheet(
      _dialog_qss()
      + f"QDialog{{background:{C['bg']};color:{C['text']};}}"
      f"QTabWidget::pane{{background:{C['panel']};border:1px solid {C['border']};border-radius:6px;}}"
      f"QTabBar::tab{{background:{C['input']};color:{C['text2']};padding:8px 14px;"
      "border-top-left-radius:4px;border-top-right-radius:4px;margin-right:2px;font-size:12px;}"
      f"QTabBar::tab:selected{{background:{C['accent']};color:{C['bg']};font-weight:700;}}"
      f"QTabBar::tab:hover{{background:{C['hover']};}}"
      f"QLineEdit,QDoubleSpinBox,QSpinBox,QComboBox,QTextEdit{{background:{C['input']};"
      f"color:{C['text']};border:1px solid {C['border']};border-radius:5px;padding:5px 8px;}}"
      f"QLabel{{background:transparent;color:{C['text']};}}"
      f"QSlider::groove:horizontal{{background:{C['border']};height:4px;border-radius:2px;}}"
      f"QSlider::handle:horizontal{{background:{C['accent']};width:14px;height:14px;"
      f"margin:-5px 0;border-radius:7px;}}"
    )
    self._setup_ui()

  # ── helpers ───────────────────────────────────────────────────────

  def _lbl(self, t: str) -> QLabel:
    l = QLabel(t)
    l.setStyleSheet(f"color:{C['text2']};font-size:11px;")
    return l

  def _le(self, val="", ph="") -> QLineEdit:
    w = QLineEdit(str(val) if val else "")
    w.setPlaceholderText(ph)
    return w

  def _spin(self, val=0, mn=0, mx=99999, step=1, dec=0):
    if dec:
      w = QDoubleSpinBox()
      w.setDecimals(dec)
      fmn, fmx = float(mn), float(mx)
      w.setRange(fmn, fmx)
      w.setSingleStep(float(step))
      fv = float(val or 0)
      w.setValue(min(fmx, max(fmn, fv)))
    else:
      w = QSpinBox()
      imn, imx = int(mn), int(mx)
      w.setRange(imn, imx)
      istep = int(round(step))
      w.setSingleStep(istep if istep > 0 else 1)
      iv = int(round(float(val or 0)))
      w.setValue(min(imx, max(imn, iv)))
    return w

  def _combo(self, items, cur="") -> QComboBox:
    w = QComboBox()
    w.addItems(items)
    idx = w.findText(str(cur or ""))
    if idx >= 0:
      w.setCurrentIndex(idx)
    return w

  # ── UI setup ──────────────────────────────────────────────────────

  def _setup_ui(self):
    lo = QVBoxLayout(self)
    lo.setContentsMargins(16, 16, 16, 12)
    lo.setSpacing(12)

    self._tabs = QTabWidget()
    self._tabs.addTab(self._tab_general(), " Général ")
    self._tabs.addTab(self._tab_address(), " Adresse ")
    self._tabs.addTab(self._tab_delivery(), " Livraison ")
    self._tabs.addTab(self._tab_contact(), " Contact ")
    if self.client.get("id"):
      self._tabs.addTab(self._tab_history(), " Historique ")
    lo.addWidget(self._tabs, 1)

    btn_row = QHBoxLayout()
    btn_row.addStretch()
    cancel = QPushButton("Annuler")
    cancel.setObjectName("secondaryBtn")
    cancel.setFixedHeight(34)
    cancel.clicked.connect(self.reject)
    save = QPushButton("Sauvegarder")
    save.setObjectName("primaryBtn")
    save.setFixedHeight(34)
    save.setMinimumWidth(120)
    save.clicked.connect(self._on_save)
    btn_row.addWidget(cancel)
    btn_row.addWidget(save)
    lo.addLayout(btn_row)

  # ── Tab 0 : Général ───────────────────────────────────────────────

  def _tab_general(self) -> QWidget:
    w = QWidget()
    fl = QFormLayout(w)
    fl.setSpacing(10)
    fl.setContentsMargins(16, 16, 16, 8)
    c = self.client

    self._name  = self._le(c.get("name", ""), "Nom du client *")
    self._company = self._le(c.get("company_name", ""), "Raison sociale")
    self._ctype  = self._combo(_CLIENT_TYPES, c.get("client_type", "standard"))
    self._status = self._combo(["actif", "archivé"],
                  "archivé" if c.get("archived") else "actif")
    self._tags  = _TagsInput(c.get("tags", "") or "")

    fl.addRow(self._lbl("Nom *"),   self._name)
    fl.addRow(self._lbl("Entreprise"), self._company)
    fl.addRow(self._lbl("Type"),    self._ctype)
    fl.addRow(self._lbl("Statut"),   self._status)
    fl.addRow(self._lbl("Tags"),    self._tags)
    return w

  # ── Tab 1 : Adresse ───────────────────────────────────────────────

  def _tab_address(self) -> QWidget:
    w = QWidget()
    lo = QVBoxLayout(w)
    lo.setSpacing(8)
    lo.setContentsMargins(16, 16, 16, 8)
    c = self.client

    fl = QFormLayout()
    fl.setSpacing(8)

    self._address = self._le(c.get("address", ""), "Adresse complète")

    coord_row = QHBoxLayout()
    self._lat = self._spin(c.get("latitude", 33.5731), -90, 90, 0.000001, 6)
    self._lon = self._spin(c.get("longitude", -7.5898), -180, 180, 0.000001, 6)
    for lbl, spin in [("Lat", self._lat), ("Lon", self._lon)]:
      coord_row.addWidget(QLabel(lbl))
      coord_row.addWidget(spin)

    self._geo_btn = QPushButton("Géocoder")
    self._geo_btn.setObjectName("primaryBtn")
    self._geo_btn.setFixedHeight(28)
    self._geo_btn.clicked.connect(self._do_geocode)
    coord_row.addWidget(self._geo_btn)

    self._geo_status = QLabel("")
    self._geo_status.setStyleSheet(f"color:{C['text2']};font-size:11px;")
    self._access_code = self._le(c.get("access_code", ""), "Code d'accès / digicode")
    self._instructions = QTextEdit()
    self._instructions.setMaximumHeight(65)
    self._instructions.setPlaceholderText("Instructions pour le chauffeur…")
    if c.get("instructions"):
      self._instructions.setPlainText(c["instructions"])

    fl.addRow(self._lbl("Adresse"),    self._address)
    fl.addRow(self._lbl("Coordonnées"),  coord_row)
    fl.addRow("",             self._geo_status)
    fl.addRow(self._lbl("Code accès"),  self._access_code)
    fl.addRow(self._lbl("Instructions"), self._instructions)
    lo.addLayout(fl)

    # Minimap
    if HAS_WEB:
      lo.addWidget(QLabel("Aperçu carte :"))
      self._map_view = QWebEngineView()
      self._map_view.setFixedHeight(200)
      self._map_view.setHtml(_minimap_html(
        float(c.get("latitude") or 33.5731),
        float(c.get("longitude") or -7.5898),
        c.get("name", ""),
      ))
      self._lat.valueChanged.connect(self._update_map)
      self._lon.valueChanged.connect(self._update_map)
      lo.addWidget(self._map_view)
    else:
      lo.addWidget(QLabel(
        f"(Minimap non disponible — {WEBENGINE_FALLBACK_SHORT})"
      ))
    return w

  def _do_geocode(self):
    addr = self._address.text().strip()
    if not addr:
      self._geo_status.setText("Entrez une adresse.")
      return
    self._geo_btn.setEnabled(False)
    self._geo_status.setText("Géocodage en cours…")
    self._geocoder = _GeocoderThread(addr, self)
    self._geocoder.result.connect(self._on_geo_ok)
    self._geocoder.error.connect(lambda e: (
      self._geo_status.setText(f"Erreur : {e[:50]}"),
      self._geo_btn.setEnabled(True),
    ))
    self._geocoder.start()

  def _on_geo_ok(self, lat, lon, name):
    self._lat.setValue(lat)
    self._lon.setValue(lon)
    self._geo_status.setText(f"OK : {name[:60]}")
    self._geo_btn.setEnabled(True)
    self._update_map()

  def _update_map(self):
    if HAS_WEB and hasattr(self, "_map_view"):
      self._map_view.setHtml(
        _minimap_html(self._lat.value(), self._lon.value(), self._name.text())
      )

  # ── Tab 2 : Livraison ─────────────────────────────────────────────

  def _tab_delivery(self) -> QWidget:
    w = QWidget()
    fl = QFormLayout(w)
    fl.setSpacing(10)
    fl.setContentsMargins(16, 16, 16, 8)
    c = self.client

    self._demand_kg = self._spin(c.get("demand_kg",  0), 0, 99999, 1,  1)
    self._demand_m3 = self._spin(c.get("demand_m3",  0), 0, 9999, 0.1, 2)
    self._svc_time  = self._spin(c.get("service_time",10), 0,  300)

    # Créneau 1
    tw1 = QHBoxLayout()
    self._ready1 = self._le(_min_to_hhmm(c.get("ready_time", 0)), "HH:MM")
    self._ready1.setFixedWidth(65)
    self._due1  = self._le(_min_to_hhmm(c.get("due_time", 1440)), "HH:MM")
    self._due1.setFixedWidth(65)
    for lbl, w2 in [("De", self._ready1), ("À", self._due1)]:
      tw1.addWidget(QLabel(lbl))
      tw1.addWidget(w2)
    tw1.addStretch()

    # Créneau 2
    tw2 = QHBoxLayout()
    self._ready2 = self._le(c.get("time_window2_start") or "", "HH:MM")
    self._ready2.setFixedWidth(65)
    self._due2  = self._le(c.get("time_window2_end")  or "", "HH:MM")
    self._due2.setFixedWidth(65)
    for lbl, w2 in [("De", self._ready2), ("À", self._due2)]:
      tw2.addWidget(QLabel(lbl))
      tw2.addWidget(w2)
    tw2.addStretch()

    adr_opts = ["", "ADR 1", "ADR 2", "ADR 3", "ADR 4",
          "ADR 5", "ADR 6", "ADR 7", "ADR 8", "ADR 9"]
    self._adr = self._combo(adr_opts, c.get("adr_class") or "")

    veh_opts = ["Aucun", "Frigo", "Poids lourd", "Camionnette", "Vélo cargo"]
    self._veh_req = self._combo(veh_opts, c.get("vehicle_requirement") or "Aucun")

    # Ponctualité slider 0→5 (0.0×→2.5×)
    punct_row = QHBoxLayout()
    self._punct = QSlider(Qt.Orientation.Horizontal)
    self._punct.setRange(0, 5)
    self._punct.setTickInterval(1)
    self._punct.setTickPosition(QSlider.TickPosition.TicksBelow)
    pf = float(c.get("punctuality_factor") or 1.0)
    self._punct.setValue(round(pf * 2))
    self._punct_lbl = QLabel(f"{pf:.1f}×")
    self._punct.valueChanged.connect(
      lambda v: self._punct_lbl.setText(f"{v / 2:.1f}×"))
    punct_row.addWidget(self._punct)
    punct_row.addWidget(self._punct_lbl)

    self._penalty = self._spin(c.get("delay_penalty_per_hour", 0), 0, 9999, 1, 1)

    fl.addRow(self._lbl("Demande (kg)"),    self._demand_kg)
    fl.addRow(self._lbl("Volume (m³)"),     self._demand_m3)
    fl.addRow(self._lbl("Durée visite (min)"),  self._svc_time)
    fl.addRow(self._lbl("Créneau 1 (HH:MM)"),  tw1)
    fl.addRow(self._lbl("Créneau 2 (optionnel)"),tw2)
    fl.addRow(self._lbl("Classe ADR"),      self._adr)
    fl.addRow(self._lbl("Véhicule requis"),   self._veh_req)
    fl.addRow(self._lbl("Ponctualité"),     punct_row)
    fl.addRow(self._lbl("Pénalité retard (€/h)"),self._penalty)
    return w

  # ── Tab 3 : Contact ───────────────────────────────────────────────

  def _tab_contact(self) -> QWidget:
    w = QWidget()
    fl = QFormLayout(w)
    fl.setSpacing(10)
    fl.setContentsMargins(16, 16, 16, 8)
    c = self.client

    self._contact = self._le(c.get("contact", ""), "Nom du contact")
    self._phone  = self._le(
      c.get("phone") or c.get("contact_phone") or "", "Téléphone"
    )
    self._email  = self._le(
      c.get("email") or c.get("contact_email") or "", "Email"
    )
    self._notes  = QTextEdit()
    self._notes.setMaximumHeight(75)
    self._notes.setPlaceholderText("Notes internes…")
    if c.get("notes"):
      self._notes.setPlainText(c["notes"])

    self._pref_driver = QComboBox()
    self._pref_driver.addItem("Aucune préférence", None)
    try:
      conn = get_connection()
      drivers = conn.execute(
        "SELECT id, first_name||' '||last_name AS name"
        " FROM drivers WHERE archived=0 ORDER BY last_name"
      ).fetchall()
      conn.close()
      for d in drivers:
        self._pref_driver.addItem(d["name"], d["id"])
      cur_id = c.get("preferred_driver_id")
      for i in range(self._pref_driver.count()):
        if self._pref_driver.itemData(i) == cur_id:
          self._pref_driver.setCurrentIndex(i)
          break
    except Exception:
      pass

    fl.addRow(self._lbl("Contact"),     self._contact)
    fl.addRow(self._lbl("Téléphone"),    self._phone)
    fl.addRow(self._lbl("Email"),      self._email)
    fl.addRow(self._lbl("Notes"),      self._notes)
    fl.addRow(self._lbl("Chauffeur préféré"),self._pref_driver)
    return w

  # ── Tab 4 : Historique ────────────────────────────────────────────

  def _tab_history(self) -> QWidget:
    w = QWidget()
    lo = QVBoxLayout(w)
    lo.setContentsMargins(16, 16, 16, 8)

    orders = []
    try:
      conn = get_connection()
      orders = conn.execute("""
        SELECT reference, status, quantity_kg, scheduled_date, actual_arrival
        FROM orders WHERE client_id= ? AND archived=0
        ORDER BY scheduled_date DESC LIMIT 10
      """, (self.client["id"],)).fetchall()
      conn.close()
    except Exception:
      pass

    if not orders:
      lo.addWidget(QLabel("Aucune commande enregistrée pour ce client."))
      return w

    tbl = QTableWidget(len(orders), 5)
    tbl.setHorizontalHeaderLabels(["Référence", "Statut", "Poids kg", "Date prévue", "Livré"])
    tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
    tbl.verticalHeader().setVisible(False)
    tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    tbl.setStyleSheet(
      f"QTableWidget{{background:{C['input']};color:{C['text']};gridline-color:{C['border']};}}"
      f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};border:1px solid {C['border']};}}"
    )
    STATUS_COLORS = {
      "delivered": "#00FF88", "failed": "#FF4757",
      "pending":  "#8899AA", "assigned": "#FFB800",
    }
    for i, o in enumerate(orders):
      tbl.setItem(i, 0, QTableWidgetItem(o["reference"] or ""))
      s_it = QTableWidgetItem(o["status"] or "")
      s_it.setForeground(QColor(STATUS_COLORS.get(o["status"] or "", C["text"])))
      tbl.setItem(i, 1, s_it)
      tbl.setItem(i, 2, QTableWidgetItem(str(o["quantity_kg"] or 0)))
      tbl.setItem(i, 3, QTableWidgetItem(str(o["scheduled_date"] or "")))
      tbl.setItem(i, 4, QTableWidgetItem(str(o["actual_arrival"] or "")))
    lo.addWidget(tbl)
    return w

  # ── Save / get_data ───────────────────────────────────────────────

  def _on_save(self):
    if not self._name.text().strip():
      QMessageBox.warning(self, "Validation", "Le nom est obligatoire.")
      return
    r1 = _hhmm_to_min(self._ready1.text())
    d1 = _hhmm_to_min(self._due1.text())
    if r1 >= d1 and d1 != 0:
      QMessageBox.warning(self, "Validation",
                "Créneau 1 : l'heure de début doit être avant la fin.")
      return
    self.accept()

  def get_data(self) -> dict:
    vr = self._veh_req.currentText()
    return {
      # Général
      "name":     self._name.text().strip(),
      "company_name": self._company.text().strip(),
      "client_type":  self._ctype.currentText(),
      "archived":   1 if self._status.currentText() == "archivé" else 0,
      "tags":     self._tags.get_tags(),
      # Adresse
      "address":    self._address.text().strip(),
      "latitude":   self._lat.value(),
      "longitude":   self._lon.value(),
      "access_code":  self._access_code.text().strip(),
      "instructions": self._instructions.toPlainText().strip(),
      # Livraison
      "demand_kg":   self._demand_kg.value(),
      "demand_m3":   self._demand_m3.value(),
      "service_time": self._svc_time.value(),
      "ready_time":  _hhmm_to_min(self._ready1.text()),
      "due_time":   _hhmm_to_min(self._due1.text()) or 1440,
      "time_window2_start": self._ready2.text().strip() or None,
      "time_window2_end":  self._due2.text().strip()  or None,
      "adr_class":     self._adr.currentText() or None,
      "vehicle_requirement":vr if vr != "Aucun" else None,
      "punctuality_factor": round(self._punct.value() / 2, 1),
      "delay_penalty_per_hour": self._penalty.value(),
      # Contact
      "contact": self._contact.text().strip(),
      "phone":  self._phone.text().strip(),
      "email":  self._email.text().strip(),
      "notes":  self._notes.toPlainText().strip(),
      "preferred_driver_id": self._pref_driver.currentData(),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MAP DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

_MAP_HTML = """<!DOCTYPE html><html><head><meta charset='utf-8'/>
<link rel='stylesheet' href='https://unpkg.com/leaflet@1.9.4/dist/leaflet.css'/>
<script src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js'></script>
<style>html,body,#map{width:100%;height:100%;margin:0;padding:0;}</style></head>
<body><div id='map'></div><script>
var TC={supermarche:'#00D4FF',restaurant:'#FF6B6B',bureau:'#96CEB4',
    pharmacie:'#C3A6FF',particulier:'#FFD93D',default:'#8899AA'};
var map=L.map('map').setView([33.5731,-7.5898],12);
L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',
 {attribution:'OpenStreetMap'}).addTo(map);
var data=CLIENT_JSON;var bnds=[];
data.forEach(function(c){
 var col=TC[c.t]||TC.default;
 var icon=L.divIcon({html:'<div style="background:'+col+
  ';width:12px;height:12px;border-radius:50%;border:2px solid #fff;'+
  'box-shadow:0 1px 3px rgba(0,0,0,.5)"></div>',iconSize:[12,12],iconAnchor:[6,6]});
 L.marker([c.a,c.o],{icon:icon}).addTo(map).bindPopup(
  '<b>'+c.n+'</b><br>'+c.t+' | prio '+c.p);
 bnds.push([c.a,c.o]);
});
if(bnds.length) map.fitBounds(bnds,{padding:[30,30]});
</script></body></html>"""


class _MapDialog(QDialog):
  def __init__(self, clients: list, parent=None):
    super().__init__(parent)
    self.setWindowTitle(f"Carte clients — {len(clients)} points")
    self.resize(820, 620)
    self.setStyleSheet(_dialog_qss() + f"QDialog{{background:{C['bg']};color:{C['text']};}}")
    lo = QVBoxLayout(self)
    lo.setContentsMargins(0, 0, 0, 0)
    lo.setSpacing(0)

    if HAS_WEB:
      view = QWebEngineView()
      data = _json.dumps([
        {"n": (c.get("name") or "")[:40],
         "t": (c.get("client_type") or "standard").lower(),
         "a": float(c.get("latitude") or 0),
         "o": float(c.get("longitude") or 0),
         "p": c.get("priority") or 3}
        for c in clients
        if float(c.get("latitude") or 0) != 0
      ])
      view.setHtml(_MAP_HTML.replace("CLIENT_JSON", data))
      lo.addWidget(view, 1)
    else:
      msg = QLabel(
        f"{len(clients)} clients — {WEBENGINE_FALLBACK_SHORT}"
      )
      msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
      msg.setStyleSheet(f"color:{C['text2']};font-size:14px;")
      lo.addWidget(msg, 1)

    # Légende + close
    legend = QHBoxLayout()
    legend.setContentsMargins(14, 6, 14, 6)
    for t, col in list(_TYPE_COLORS.items())[:6]:
      dot = QLabel()
      dot.setFixedSize(10, 10)
      dot.setStyleSheet(f"background:{col};border-radius:5px;")
      legend.addWidget(dot)
      legend.addWidget(QLabel(t))
      legend.addSpacing(10)
    legend.addStretch()
    close_btn = QPushButton("Fermer")
    close_btn.setObjectName("secondaryBtn")
    close_btn.clicked.connect(self.accept)
    legend.addWidget(close_btn)
    lo.addLayout(legend)


# ═══════════════════════════════════════════════════════════════════════════════
# ANOMALY DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class _AnomalyDialog(QDialog):
  def __init__(self, anomalies: list, parent=None):
    super().__init__(parent)
    n = sum(len(a["issues"]) for a in anomalies)
    self.setWindowTitle(f"Anomalies clients — {len(anomalies)} clients, {n} problèmes")
    self.resize(720, 480)
    self.setStyleSheet(_dialog_qss() + f"QDialog{{background:{C['bg']};color:{C['text']};}}")
    lo = QVBoxLayout(self)
    lo.setContentsMargins(16, 16, 16, 12)
    lo.setSpacing(10)

    if not anomalies:
      lo.addWidget(QLabel("Aucune anomalie détectée. Les données clients semblent cohérentes."))
    else:
      lo.addWidget(QLabel(
        f"<b style='color:{C['warning']}'>{len(anomalies)} clients</b>"
        f" présentent <b style='color:{C['danger']}'>{n} anomalie(s)</b>."
      ))
      tbl = QTableWidget(n, 4)
      tbl.setHorizontalHeaderLabels(["ID", "Client", "Sévérité", "Description"])
      tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
      tbl.verticalHeader().setVisible(False)
      tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
      tbl.setStyleSheet(
        f"QTableWidget{{background:{C['input']};color:{C['text']};gridline-color:{C['border']};}}"
        f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};border:1px solid {C['border']};}}"
      )
      SEV = {"high": C["danger"], "medium": C["warning"], "low": C["text2"]}
      row = 0
      for a in anomalies:
        for sev, desc in a["issues"]:
          tbl.setItem(row, 0, QTableWidgetItem(str(a["id"])))
          tbl.setItem(row, 1, QTableWidgetItem(a["name"]))
          s_it = QTableWidgetItem(sev)
          s_it.setForeground(QColor(SEV.get(sev, C["text"])))
          tbl.setItem(row, 2, s_it)
          tbl.setItem(row, 3, QTableWidgetItem(desc))
          row += 1
      lo.addWidget(tbl)

    btn_row = QHBoxLayout()
    btn_row.addStretch()
    close = QPushButton("Fermer")
    close.setObjectName("secondaryBtn")
    close.clicked.connect(self.accept)
    btn_row.addWidget(close)
    lo.addLayout(btn_row)


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT REPORT DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class _ImportReportDialog(QDialog):
  def __init__(self, report: dict, parent=None):
    super().__init__(parent)
    self.setWindowTitle("Rapport d'import clients")
    self.resize(500, 380)
    self.setStyleSheet(_dialog_qss() + f"QDialog{{background:{C['bg']};color:{C['text']};}}")
    lo = QVBoxLayout(self)
    lo.setContentsMargins(20, 20, 20, 12)
    lo.setSpacing(12)

    summary = QLabel(
      f"<b style='color:{C['success']}'> {report['created']} clients créés</b><br>"
      f"<b style='color:{C['accent']}'>Reset {report['updated']} clients mis à jour</b><br>"
      f"<b style='color:{C['danger']}'> {report['errors']} erreurs</b>"
    )
    summary.setTextFormat(Qt.TextFormat.RichText)
    lo.addWidget(summary)

    if report.get("error_list"):
      lo.addWidget(QLabel("Détail des erreurs :"))
      err = QTextEdit()
      err.setReadOnly(True)
      err.setMaximumHeight(180)
      err.setStyleSheet(
        f"QTextEdit{{background:{C['input']};color:{C['danger']};"
        f"border:1px solid {C['border']};border-radius:5px;}}"
      )
      err.setPlainText("\n".join(report["error_list"][:50]))
      lo.addWidget(err)

    btn_row = QHBoxLayout()
    btn_row.addStretch()
    ok = QPushButton("Fermer")
    ok.setObjectName("primaryBtn")
    ok.clicked.connect(self.accept)
    btn_row.addWidget(ok)
    lo.addLayout(btn_row)


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT MAPPING DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

_DB_FIELDS = [
  ("name",      "Nom *"),
  ("company_name",  "Entreprise"),
  ("address",    "Adresse"),
  ("latitude",    "Latitude"),
  ("longitude",   "Longitude"),
  ("demand_kg",   "Demande kg"),
  ("demand_m3",   "Volume m³"),
  ("ready_time",   "Créneau début (min)"),
  ("due_time",    "Créneau fin (min)"),
  ("service_time",  "Durée visite (min)"),
  ("priority",    "Priorité (1-5)"),
  ("client_type",  "Type"),
  ("phone",     "Téléphone"),
  ("email",     "Email"),
  ("tags",      "Tags"),
]
_ALIASES: dict[str, list[str]] = {
  "name":     ["nom", "name", "client", "customer"],
  # Pas d'alias "x"/"y" seuls : trop ambigu avec des colonnes texte ou codes.
  "latitude":   ["lat", "latitude", "ycoord"],
  "longitude":  ["lon", "long", "longitude", "xcoord"],
  "demand_kg":  ["demand", "demande", "poids", "kg", "quantity", "demand_kg"],
  "demand_m3":  ["demand_m3", "volume_m3", "m3"],
  "address":   ["adresse", "address", "addr"],
  "phone":    ["tel", "phone", "telephone", "contact_phone"],
  "email":    ["email", "mail", "contact_email"],
  "client_type": ["type", "client_type", "categorie"],
  "company_name": ["entreprise", "company", "societe", "raison_sociale"],
  "priority":   ["priorite", "priority", "prio"],
  "ready_time":  ["ready", "ready_time", "heure_debut", "debut"],
  "due_time":   ["due", "due_time", "heure_fin", "fin"],
  "service_time": ["service", "service_time", "duree", "unloading_time"],
  "tags":     ["tags", "etiquettes"],
}


class _ImportMappingDialog(QDialog):
  """Import : preview 5 lignes + mapping colonnes CSV → champs DB + option géocodage."""

  def __init__(self, filepath: str, parent=None):
    super().__init__(parent)
    self.filepath = filepath
    self.col_map: dict = {}
    self.geocode = False
    self.setWindowTitle("Import clients — Mapping colonnes")
    self.setMinimumSize(700, 560)
    self.setModal(True)
    self.setStyleSheet(
      _dialog_qss()
      + f"QLabel{{background:transparent;color:{C['text']};}}"
      f"QComboBox{{background:{C['input']};color:{C['text']};border:1px solid {C['border']};"
      "border-radius:4px;padding:3px 6px;}"
      f"QTableWidget{{background:{C['input']};color:{C['text']};gridline-color:{C['border']};}}"
      f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};border:1px solid {C['border']};}}"
      f"QCheckBox{{color:{C['text']};background:transparent;spacing:8px;}}"
      f"QScrollArea{{border:none;background:transparent;}}"
    )
    self._headers, self._preview = _read_headers_and_preview(filepath)
    self._combos: dict[str, QComboBox] = {}
    self._setup_ui()
    self._auto_map()

  def _setup_ui(self):
    lo = QVBoxLayout(self)
    lo.setContentsMargins(16, 16, 16, 12)
    lo.setSpacing(10)

    lo.addWidget(QLabel(
      f"<b>Fichier :</b> {os.path.basename(self.filepath)} "
      f"— {len(self._headers)} colonnes détectées"
    ))

    # Preview
    if self._preview:
      lo.addWidget(QLabel("Aperçu (5 premières lignes) :"))
      prev = QTableWidget(min(5, len(self._preview)), len(self._headers))
      prev.setHorizontalHeaderLabels(self._headers)
      prev.setMaximumHeight(140)
      prev.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
      prev.horizontalHeader().setSectionResizeMode(
        QHeaderView.ResizeMode.ResizeToContents)
      for r, row in enumerate(self._preview[:5]):
        for c, col in enumerate(self._headers):
          prev.setItem(r, c, QTableWidgetItem(str(row.get(col) or "")))
      lo.addWidget(prev)

    lo.addWidget(QLabel("Associer colonnes fichier → champs CityPulse :"))

    scroll = QScrollArea()
    scroll.setWidgetResizable(True)
    scroll.setMaximumHeight(230)
    inner = QWidget()
    fl = QFormLayout(inner)
    fl.setSpacing(6)
    fl.setContentsMargins(8, 8, 8, 8)

    opts = ["(ignorer)"] + self._headers
    for db_field, label in _DB_FIELDS:
      cb = QComboBox()
      cb.addItems(opts)
      fl.addRow(QLabel(label), cb)
      self._combos[db_field] = cb
    scroll.setWidget(inner)
    lo.addWidget(scroll)

    self._geocode_cb = QCheckBox(
      "Géocoder les adresses manquantes (Nominatim — ~1s/ligne)"
    )
    lo.addWidget(self._geocode_cb)

    btn_row = QHBoxLayout()
    btn_row.addStretch()
    cancel = QPushButton("Annuler")
    cancel.setObjectName("secondaryBtn")
    cancel.setFixedHeight(36)
    cancel.setCursor(Qt.CursorShape.PointingHandCursor)
    cancel.clicked.connect(self.reject)
    ok = QPushButton("Importer")
    ok.setObjectName("primaryBtn")
    ok.setFixedHeight(36)
    ok.setCursor(Qt.CursorShape.PointingHandCursor)
    ok.clicked.connect(self._accept)
    btn_row.addWidget(cancel)
    btn_row.addWidget(ok)
    lo.addLayout(btn_row)

  def _auto_map(self):
    header_lower = {h.lower(): h for h in self._headers}
    for db_field, alts in _ALIASES.items():
      if db_field not in self._combos:
        continue
      for alt in alts:
        if alt in header_lower:
          idx = self._combos[db_field].findText(header_lower[alt])
          if idx >= 0:
            self._combos[db_field].setCurrentIndex(idx)
          break

  def _preview_cell(self, preview_row: dict, col: str):
    if col not in preview_row:
      return None
    v = preview_row[col]
    if v in (None, "", "None"):
      return None
    return v

  def _accept(self):
    self.col_map = {
      f: cb.currentText()
      for f, cb in self._combos.items()
      if cb.currentIndex() > 0
    }
    if "name" not in self.col_map:
      QMessageBox.warning(self, "Mapping incomplet",
                "Associez au moins la colonne 'Nom *'.")
      return
    bad_msgs: list[str] = []
    for field in _IMPORT_NUMERIC_FIELDS:
      if field not in self.col_map:
        continue
      col = self.col_map[field]
      for ri, preview_row in enumerate(self._preview[:5]):
        raw = self._preview_cell(preview_row, col)
        if raw is None:
          continue
        _, bad = _import_parse_float(raw, 0.0)
        if bad:
          bad_msgs.append(
            f"Aperçu ligne {ri + 1}, «{field}» : {str(raw)[:40]!r}"
          )
          break
    if bad_msgs:
      QMessageBox.warning(
        self,
        "Valeurs non numériques",
        "Des colonnes mappées sur des champs numériques contiennent du texte "
        "dans l’aperçu. Corrigez le mapping ou le fichier.\n\n"
        + "\n".join(bad_msgs[:5]),
      )
      return
    self.geocode = self._geocode_cb.isChecked()
    self.accept()


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTS WIDGET — PAGE PRINCIPALE
# ═══════════════════════════════════════════════════════════════════════════════

class ClientsWidget(QWidget):
  PAGE_SIZE = 100

  def __init__(self, main_window):
    super().__init__()
    self.main_window = main_window
    self._page   = 0
    self._total  = 0
    self._search  = ""
    self._filter_prio_max = 5
    self._threads: list = []
    self._setup_ui()

  # ── UI construction ───────────────────────────────────────────────

  def _setup_ui(self):
    root = QVBoxLayout(self)
    root.setContentsMargins(4, 4, 4, 8)
    root.setSpacing(14)

    # ── SectionHeader ──────────────────────────────────────────────
    self._header = SectionHeader(
      title="Gestion des Clients",
      subtitle="Livraisons, fenêtres horaires, géocodage, anomalies",
      action_text="+ Ajouter",
      action_callback=self._add_client,
    )
    root.addWidget(self._header)

    # ── Toolbar ────────────────────────────────────────────────────
    toolbar = QHBoxLayout()
    toolbar.setSpacing(6)

    self._search_bar = SearchBar(placeholder="Rechercher (nom, entreprise, tél, tags)…")
    self._search_bar.setMaximumWidth(310)
    self._search_bar.search_changed.connect(self._on_search)
    toolbar.addWidget(self._search_bar)
    toolbar.addSpacing(6)

    _BTN_STYLE = (
      f"QPushButton{{background:{C['input']};color:{C['text']};"
      f"border:1px solid {C['border']};border-radius:5px;"
      "font-size:12px;padding:4px 10px;}}"
      f"QPushButton:hover{{background:{C['hover']};border-color:{C['accent']};}}"
    )
    for attr, txt, tip, fn in [
      ("_btn_import",    "Importer",   "Importer CSV ou Excel",          self._import_data),
      ("_btn_export",    "Exporter",   "Exporter CSV / Excel / JSON",    self._export_menu),
      ("_btn_map",       "Vue Carte",  "Afficher sur carte Leaflet",     self._show_map),
      ("_btn_anomalies", "Anomalies",  "Détecter anomalies (z-score)",   self._detect_anomalies),
    ]:
      btn = QPushButton(txt)
      btn.setFixedHeight(30)
      btn.setToolTip(tip)
      btn.setStyleSheet(_BTN_STYLE)
      btn.clicked.connect(fn)
      toolbar.addWidget(btn)
      setattr(self, attr, btn)

    toolbar.addSpacing(8)

    # Batch ops
    self._batch_combo = QComboBox()
    self._batch_combo.addItems([
      "Action lot…", "Archiver sélection", "Exporter sélection CSV",
    ])
    self._batch_combo.setFixedWidth(165)
    self._batch_combo.setStyleSheet(_BTN_STYLE)
    toolbar.addWidget(self._batch_combo)

    self._batch_apply_btn = QPushButton("Appliquer")
    self._batch_apply_btn.setObjectName("primaryBtn")
    self._batch_apply_btn.setFixedHeight(30)
    self._batch_apply_btn.clicked.connect(self._apply_batch)
    toolbar.addWidget(self._batch_apply_btn)

    toolbar.addStretch()
    self._count_lbl = QLabel("0 clients")
    self._count_lbl.setStyleSheet(f"color:{C['text2']};font-size:12px;")
    toolbar.addWidget(self._count_lbl)
    toolbar.addSpacing(4)
    _hb = QPushButton()
    _hb.setFixedSize(30, 30)
    _hb.setToolTip("Aide — Clients")
    _hb.setCursor(Qt.CursorShape.PointingHandCursor)
    apply_action_button(_hb, "help-circle", "#7FA8C0", "#1A2E4A", "#1A3A5C", 18)
    _hb.clicked.connect(lambda: show_help(self, "clients"))
    toolbar.addWidget(_hb)

    root.addLayout(toolbar)

    # ── Filtres (CollapsibleSection) ───────────────────────────────
    self._filter_section = CollapsibleSection("Filtres avancés", collapsed=True)
    f_lo = self._filter_section._inner

    row1 = QHBoxLayout()
    row1.addWidget(QLabel("Type :"))
    self._type_cbs: dict[str, QCheckBox] = {}
    for t in ["supermarche", "restaurant", "bureau", "pharmacie", "particulier", "standard"]:
      cb = QCheckBox(t)
      cb.stateChanged.connect(self._on_filter_changed)
      self._type_cbs[t] = cb
      row1.addWidget(cb)
    row1.addStretch()
    f_lo.addLayout(row1)

    row2 = QHBoxLayout()
    row2.addWidget(QLabel("Priorité max :"))
    self._prio_slider = QSlider(Qt.Orientation.Horizontal)
    self._prio_slider.setRange(1, 5)
    self._prio_slider.setValue(5)
    self._prio_slider.setFixedWidth(120)
    self._prio_lbl = QLabel("5")
    self._prio_slider.valueChanged.connect(lambda v: (
      self._prio_lbl.setText(str(v)),
      setattr(self, "_filter_prio_max", v),
      self._on_filter_changed(),
    ))
    row2.addWidget(self._prio_slider)
    row2.addWidget(self._prio_lbl)
    row2.addSpacing(20)
    row2.addWidget(QLabel("Tag :"))
    self._tag_filter = QLineEdit()
    self._tag_filter.setPlaceholderText("ex: alimentaire")
    self._tag_filter.setFixedWidth(130)
    self._tag_filter.setStyleSheet(
      f"QLineEdit{{background:{C['input']};color:{C['text']};border:1px solid {C['border']};"
      "border-radius:5px;padding:3px 6px;font-size:12px;}}"
    )
    self._tag_filter.textChanged.connect(self._on_filter_changed)
    row2.addWidget(self._tag_filter)
    row2.addSpacing(16)
    save_f = QPushButton("Sauvegarder filtre")
    save_f.setFixedHeight(26)
    save_f.setStyleSheet(_BTN_STYLE)
    save_f.clicked.connect(lambda: show_toast(
      self.window(), "Filtre sauvegardé (session uniquement).", "info"))
    row2.addWidget(save_f)
    row2.addStretch()
    f_lo.addLayout(row2)
    root.addWidget(self._filter_section)

    # ── Table ──────────────────────────────────────────────────────
    self._table = QTableWidget()
    self._table.setColumnCount(10)
    self._table.setHorizontalHeaderLabels([
      "ID", "Nom", "Entreprise", "Tél",
      "Demande kg", "Créneaux", "Priorité", "Tags", "Statut", "Actions",
    ])
    hdr = self._table.horizontalHeader()
    hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
    hdr.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
    for col, w in [(0,50),(2,140),(3,110),(4,88),(5,120),(6,80),(8,100),(9,120)]:
      self._table.setColumnWidth(col, w)

    self._table.verticalHeader().setVisible(False)
    self._table.verticalHeader().setDefaultSectionSize(36)
    self._table.setAlternatingRowColors(True)
    self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
    self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
    self._table.setSortingEnabled(False)
    self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
    self._table.customContextMenuRequested.connect(self._context_menu)
    self._table.doubleClicked.connect(self._on_double_click)
    self._table.setStyleSheet(
      f"QTableWidget{{background:{C['bg']};color:{C['text']};"
      f"gridline-color:{C['border']};border:none;"
      "alternate-background-color:#0F2035;}"
      f"QTableWidget::item:selected{{background:{C['hover']};color:{C['accent']};}}"
      f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};"
      f"border:1px solid {C['border']};padding:4px 6px;font-size:11px;font-weight:600;}}"
    )

    self._empty = EmptyState(
      title="Aucun client",
      subtitle="Importez un fichier CSV/Excel ou ajoutez un client manuellement.",
      action_text="+ Ajouter un client",
      action_callback=self._add_client,
    )
    self._stack = QStackedWidget()
    self._stack.addWidget(self._table)
    self._stack.addWidget(self._empty)
    root.addWidget(self._stack, 1)

    # ── Pagination ─────────────────────────────────────────────────
    self._pagination = PaginationBar(page_size=self.PAGE_SIZE)
    self._pagination.page_changed.connect(self._on_page_changed)
    root.addWidget(self._pagination)

    self._overlay = LoadingOverlay(self)

  # ── Data loading ──────────────────────────────────────────────────

  def _build_where(self) -> tuple[str, list]:
    clauses = ["archived=0"]
    params: list = []

    if self._search:
      s = f"%{self._search}%"
      clauses.append(
        "(name LIKE ? OR COALESCE(company_name,'') LIKE ? "
        " OR COALESCE(phone,'') LIKE ? OR COALESCE(tags,'') LIKE ? )"
      )
      params += [s, s, s, s]

    sel = [k for k, cb in self._type_cbs.items() if cb.isChecked()]
    if sel:
      clauses.append(f"client_type IN ({','.join('?'*len(sel))})")
      params += sel

    if self._filter_prio_max < 5:
      clauses.append("priority <= ")
      params.append(self._filter_prio_max)

    tag = self._tag_filter.text().strip() if hasattr(self, "_tag_filter") else ""
    if tag:
      clauses.append("COALESCE(tags,'') LIKE ? ")
      params.append(f"%{tag}%")

    return " WHERE " + " AND ".join(clauses), params

  def retranslate_ui(self, lang: str):
    from app.i18n import tr
    if hasattr(self, "_header"):
        self._header.set_title(tr("section.clients", lang))
        self._header.set_subtitle(tr("clients.subtitle", lang))
    if hasattr(self, "_btn_import"):
        self._btn_import.setText(tr("clients.btn.import", lang))
        self._btn_export.setText(tr("clients.btn.export", lang))
        self._btn_map.setText(tr("clients.btn.map", lang))
        self._btn_anomalies.setText(tr("clients.btn.anomalies", lang))
    if hasattr(self, "_batch_combo") and self._batch_combo.count() > 0:
        self._batch_combo.setItemText(0, tr("clients.batch.default", lang))
    if hasattr(self, "_batch_apply_btn"):
        self._batch_apply_btn.setText(tr("clients.btn.apply", lang))
    if hasattr(self, "_filter_section"):
        self._filter_section.set_title(tr("clients.filter.title", lang))

  def refresh_data(self):
    where, params = self._build_where()
    conn = get_connection()
    self._total = conn.execute(
      f"SELECT COUNT(*) FROM clients{where}", params
    ).fetchone()[0]

    offset = self._page * self.PAGE_SIZE
    rows = conn.execute(
      f"SELECT * FROM clients{where}"
      f" ORDER BY priority, name LIMIT {self.PAGE_SIZE} OFFSET {offset}",
      params,
    ).fetchall()
    conn.close()

    self._pagination.update_total(self._total)
    self._count_lbl.setText(f"{self._total} clients")
    self._fill_table(rows)
    self._stack.setCurrentIndex(0 if self._total else 1)

  def _fill_table(self, rows):
    self._table.blockSignals(True)
    self._table.setRowCount(len(rows))

    for r, row in enumerate(rows):
      def _item(val, align=None, color=None) -> QTableWidgetItem:
        it = QTableWidgetItem(str(val) if val is not None else "")
        it.setFlags(Qt.ItemFlag(it.flags().value & ~Qt.ItemFlag.ItemIsEditable.value))
        if align:
          it.setTextAlignment(align)
        if color:
          it.setForeground(QColor(color))
        return it

      R = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter

      # 0 ID
      self._table.setItem(r, 0, _item(row["id"], align=R))

      # 1 Nom — store client_id in UserRole
      name_it = _item(row["name"])
      name_it.setData(Qt.ItemDataRole.UserRole, row["id"])
      self._table.setItem(r, 1, name_it)

      # 2 Entreprise
      company = ""
      try:
        company = row["company_name"] or ""
      except Exception:
        pass
      self._table.setItem(r, 2, _item(company))

      # 3 Tél
      phone = ""
      try:
        phone = row["phone"] or ""
      except Exception:
        pass
      self._table.setItem(r, 3, _item(phone))

      # 4 Demande kg
      self._table.setItem(r, 4, _item(
        f"{float(row['demand_kg']):.1f}" if row["demand_kg"] is not None else "0",
        align=R,
      ))

      # 5 Créneaux
      tw = (f"{_min_to_hhmm(row['ready_time'])}"
         f"–{_min_to_hhmm(row['due_time'])}")
      self._table.setItem(r, 5, _item(tw))

      # 6 Priorité 
      self._table.setItem(r, 6, _item(
        _priority_stars(row["priority"]), color="#FFB800"
      ))

      # 7 Tags
      tags = ""
      try:
        tags = row["tags"] or ""
      except Exception:
        pass
      self._table.setItem(r, 7, _item(tags))

      # 8 Statut (type)
      ctype = row.get("client_type") or "standard"
      self._table.setItem(r, 8, _item(ctype, color=_type_color(ctype)))

      # 9 Actions
      self._table.setCellWidget(r, 9, self._make_actions(row["id"]))

    self._table.blockSignals(False)

  def _make_actions(self, cid: int) -> QWidget:
    w = QWidget()
    lo = QHBoxLayout(w)
    lo.setContentsMargins(4, 2, 4, 2)
    lo.setSpacing(3)
    for lucide_key, tip, fn, bg, fg, hbg in [
      ("pencil",   "Modifier",          lambda _, i=cid: self._edit_client(i),        C["hover"], C["accent"],  C["panel"]),
      ("map",      "Voir sur carte",    lambda _, i=cid: self._show_one_map(i),       C["hover"], "#96CEB4",    C["panel"]),
      ("globe",    "Compte web",        lambda _, i=cid: self._create_web_account(i), C["hover"], "#7EC8E3",    C["panel"]),
      ("archive",  "Archiver",          lambda _, i=cid: self._delete_client(i),      C["hover"], C["danger"],  "#3A1020"),
    ]:
      btn = QPushButton()
      btn.setFixedSize(28, 28)
      btn.setToolTip(tip)
      btn.setCursor(Qt.CursorShape.PointingHandCursor)
      apply_action_button(btn, lucide_key, fg, bg, hbg, icon_px=16)
      btn.clicked.connect(fn)
      lo.addWidget(btn)
    return w

  def _create_web_account(self, cid: int):
    conn = get_connection()
    row = conn.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not row:
      return
    name = row["name"] or ""
    parts = name.strip().split()
    first = parts[0] if parts else "Client"
    last = " ".join(parts[1:]) if len(parts) > 1 else str(cid)
    email = row["email"] if "email" in row.keys() else ""
    svc = get_django_service()
    if not svc.base_url or not svc.secret_key:
      from PyQt6.QtWidgets import QMessageBox
      QMessageBox.warning(self, "Compte web",
        "Django non configure.\nAllez dans Parametres > Django et renseignez l'URL et la cle.")
      return
    result = svc.create_web_user(
      desktop_id=cid, role="client",
      first_name=first, last_name=last, email=email or "",
    )
    self._show_credentials_dialog(result, "client", name)

  def _show_credentials_dialog(self, result: dict, role: str, display_name: str):
    from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QLineEdit, QPushButton, QFormLayout
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QClipboard
    from PyQt6.QtWidgets import QApplication
    if not result.get("ok"):
      from PyQt6.QtWidgets import QMessageBox
      QMessageBox.warning(self, "Compte web", f"Erreur : {result.get('error','inconnue')}")
      return
    dlg = QDialog(self)
    dlg.setWindowTitle("Compte web cree")
    dlg.setFixedWidth(380)
    dlg.setStyleSheet(_dialog_qss() + f"QDialog{{background:{C['bg']};color:{C['text']};}}")
    lo = QVBoxLayout(dlg)
    lo.setSpacing(12)
    status_lbl = QLabel("Compte cree" if result.get("created") else "Compte mis a jour")
    status_lbl.setStyleSheet(f"color:{'#00FF88' if result.get('created') else '#FFB800'};font-weight:bold;font-size:13px;")
    lo.addWidget(status_lbl)
    lo.addWidget(QLabel(f"Utilisateur : <b>{display_name}</b>  ({role})"))
    form = QFormLayout()
    form.setSpacing(8)
    u_edit = QLineEdit(result.get("username", ""))
    u_edit.setReadOnly(True)
    u_edit.setStyleSheet(f"background:{C['panel']};color:{C['text']};border:1px solid {C['border']};border-radius:5px;padding:4px 8px;")
    p_edit = QLineEdit(result.get("_password", ""))
    p_edit.setReadOnly(True)
    p_edit.setStyleSheet(u_edit.styleSheet())
    form.addRow("Identifiant :", u_edit)
    form.addRow("Mot de passe :", p_edit)
    lo.addLayout(form)
    note = QLabel("Transmettez ces identifiants a la personne.\nLe mot de passe ne sera plus affiche.")
    note.setStyleSheet(f"color:{C['text2']};font-size:11px;")
    note.setWordWrap(True)
    lo.addWidget(note)
    url_lbl = QLabel(f"URL : {result.get('_url', '') or 'http://127.0.0.1:8000'}")
    url_lbl.setStyleSheet(f"color:{C['accent']};font-size:11px;")
    lo.addWidget(url_lbl)
    copy_btn = QPushButton("Copier les identifiants")
    copy_btn.setObjectName("secondaryBtn")
    def _copy():
      QApplication.clipboard().setText(
        f"Identifiant : {u_edit.text()}\nMot de passe : {p_edit.text()}")
      copy_btn.setText("Copie !")
    copy_btn.clicked.connect(_copy)
    lo.addWidget(copy_btn)
    close_btn = QPushButton("Fermer")
    close_btn.setObjectName("primaryBtn")
    close_btn.clicked.connect(dlg.accept)
    lo.addWidget(close_btn)
    dlg.exec()

  # ── Signals / slots ───────────────────────────────────────────────

  def _on_search(self, text: str):
    self._search = text
    self._page = 0
    self.refresh_data()

  def _on_filter_changed(self, *_):
    self._filter_prio_max = self._prio_slider.value()
    self._page = 0
    self.refresh_data()

  def _on_page_changed(self, page: int, offset: int, limit: int):
    self._page = page
    self.refresh_data()

  def _on_double_click(self, idx):
    item = self._table.item(idx.row(), 1)
    if item:
      cid = item.data(Qt.ItemDataRole.UserRole)
      if cid:
        self._edit_client(cid)

  def _context_menu(self, pos: QPoint):
    row = self._table.rowAt(pos.y())
    if row < 0:
      return
    item = self._table.item(row, 1)
    cid = item.data(Qt.ItemDataRole.UserRole) if item else None
    if not cid:
      return

    menu = QMenu(self)
    menu.setStyleSheet(
      f"QMenu{{background:{C['panel']};color:{C['text']};border:1px solid {C['border']};"
      "border-radius:6px;padding:4px;}"
      f"QMenu::item{{padding:6px 18px;border-radius:4px;}}"
      f"QMenu::item:selected{{background:{C['hover']};}}"
    )
    for label, fn in [
      (" Modifier",      lambda: self._edit_client(cid)),
      (" Voir sur carte",   lambda: self._show_one_map(cid)),
      (" Dupliquer",     lambda: self._duplicate_client(cid)),
      (None, None),
      (" Archiver",     lambda: self._delete_client(cid)),
    ]:
      if label is None:
        menu.addSeparator()
      else:
        act = QAction(label, self)
        act.triggered.connect(fn)
        menu.addAction(act)
    menu.exec(self._table.viewport().mapToGlobal(pos))

  # ── CRUD ──────────────────────────────────────────────────────────

  def _save_client_fields(self, conn, cid: int, data: dict):
    """Mise à jour des colonnes étendues (idempotent si colonne absente)."""
    for col, val in [
      ("company_name", data.get("company_name")),
      ("tags",     data.get("tags")),
      ("access_code", data.get("access_code")),
      ("notes",    data.get("notes")),
      ("preferred_driver_id",  data.get("preferred_driver_id")),
      ("punctuality_factor",   data.get("punctuality_factor")),
      ("delay_penalty_per_hour", data.get("delay_penalty_per_hour")),
      ("vehicle_requirement",  data.get("vehicle_requirement")),
      ("adr_class",       data.get("adr_class")),
      ("time_window2_start",   data.get("time_window2_start")),
      ("time_window2_end",    data.get("time_window2_end")),
    ]:
      if val is not None:
        try:
          conn.execute(f"UPDATE clients SET {col}= ? WHERE id= ?", (val, cid))
        except Exception:
          pass

  def _add_client(self):
    dlg = _ClientDialog(self)
    if dlg.exec() != QDialog.DialogCode.Accepted:
      return
    data = dlg.get_data()
    conn = get_connection()
    cur = conn.execute("""
      INSERT INTO clients
      (name,address,latitude,longitude,demand_kg,demand_m3,
       ready_time,due_time,service_time,priority,client_type,
       contact,phone,email,instructions,archived,created_at)
      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,datetime('now'))
    """, (
      data["name"], data["address"], data["latitude"], data["longitude"],
      data["demand_kg"], data["demand_m3"], data["ready_time"], data["due_time"],
      data["service_time"], 3, data["client_type"],
      data["contact"], data["phone"], data["email"], data["instructions"],
    ))
    self._save_client_fields(conn, cur.lastrowid, data)
    conn.commit()
    conn.close()
    log_action("CLIENT_CREATE", f"Client '{data['name']}' créé")
    show_toast(self.window(), f"Client '{data['name']}' créé", "success")
    self.refresh_data()

  def _edit_client(self, cid: int):
    conn = get_connection()
    row = conn.execute("SELECT * FROM clients WHERE id= ?", (cid,)).fetchone()
    conn.close()
    if not row:
      return
    dlg = _ClientDialog(self, dict(row))
    if dlg.exec() != QDialog.DialogCode.Accepted:
      return
    data = dlg.get_data()
    conn = get_connection()
    conn.execute("""
      UPDATE clients SET
      name= ?,address= ?,latitude= ?,longitude= ?,demand_kg= ?,demand_m3= ?,
      ready_time= ?,due_time= ?,service_time= ?,client_type= ?,
      contact= ?,phone= ?,email= ?,instructions= ?,archived= ?,
      updated_at=datetime('now') WHERE id=?
    """, (
      data["name"], data["address"], data["latitude"], data["longitude"],
      data["demand_kg"], data["demand_m3"], data["ready_time"], data["due_time"],
      data["service_time"], data["client_type"],
      data["contact"], data["phone"], data["email"], data["instructions"],
      data["archived"], cid,
    ))
    self._save_client_fields(conn, cid, data)
    conn.commit()
    conn.close()
    log_action("CLIENT_UPDATE", f"Client #{cid} modifié")
    show_toast(self.window(), "Client mis à jour", "success")
    self.refresh_data()

  def _delete_client(self, cid: int):
    if not ConfirmDialog.ask(
      self,
      "Archiver client",
      "Archiver ce client (suppression logique)",
      "warning",
      "Archiver",
    ):
      return
    conn = get_connection()
    conn.execute("UPDATE clients SET archived=1, updated_at=datetime('now') WHERE id= ?", (cid,))
    conn.commit()
    conn.close()
    log_action("CLIENT_DELETE", f"Client #{cid} archivé")
    show_toast(self.window(), "Client archivé", "info")
    self.refresh_data()

  def _duplicate_client(self, cid: int):
    conn = get_connection()
    row = conn.execute("SELECT * FROM clients WHERE id= ?", (cid,)).fetchone()
    if not row:
      conn.close()
      return
    conn.execute("""
      INSERT INTO clients
      (name,address,latitude,longitude,demand_kg,demand_m3,ready_time,due_time,
       service_time,priority,client_type,contact,phone,email,instructions,
       archived,created_at)
      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,datetime('now'))
    """, (
      f"{row['name']} (copie)", row.get("address"), row["latitude"], row["longitude"],
      row["demand_kg"], row.get("demand_m3", 0), row["ready_time"], row["due_time"],
      row["service_time"], row["priority"], row.get("client_type"),
      row.get("contact"), row.get("phone"), row.get("email"), row.get("instructions"),
    ))
    conn.commit()
    conn.close()
    log_action("CLIENT_DUPLICATE", f"Client #{cid} dupliqué")
    show_toast(self.window(), "Client dupliqué", "success")
    self.refresh_data()

  # ── Batch ops ─────────────────────────────────────────────────────

  def _selected_ids(self) -> list[int]:
    ids: list[int] = []
    seen: set = set()
    for idx in self._table.selectedIndexes():
      item = self._table.item(idx.row(), 1)
      if item:
        cid = item.data(Qt.ItemDataRole.UserRole)
        if cid and cid not in seen:
          ids.append(cid)
          seen.add(cid)
    return ids

  def _apply_batch(self):
    op = self._batch_combo.currentText()
    if op == "Action lot…":
      show_toast(self.window(), "Choisissez une action dans la liste.", "info")
      return
    ids = self._selected_ids()
    if not ids:
      show_toast(self.window(), "Sélectionnez au moins un client.", "info")
      return

    if op == "Archiver sélection":
      if not ConfirmDialog.ask(
        self,
        "Archiver",
        f"Archiver {len(ids)} clients ",
        "warning",
        "Archiver",
      ):
        return
      conn = get_connection()
      conn.execute(
        f"UPDATE clients SET archived=1 WHERE id IN ({','.join('?'*len(ids))})",
        ids
      )
      conn.commit()
      conn.close()
      log_action("CLIENT_BATCH_ARCHIVE", f"{len(ids)} clients archivés")
      show_toast(self.window(), f"{len(ids)} clients archivés", "success")
      self.refresh_data()

    elif "Exporter sélection" in op:
      path, _ = QFileDialog.getSaveFileName(
        self, "Exporter sélection", "clients_selection.csv", "CSV (*.csv)")
      if path:
        self._export_ids_csv(ids, path)

  def _export_ids_csv(self, ids: list, path: str):
    conn = get_connection()
    rows = conn.execute(
      f"SELECT * FROM clients WHERE id IN ({','.join('?'*len(ids))})", ids
    ).fetchall()
    conn.close()
    COLS = ["id","name","company_name","address","latitude","longitude",
        "demand_kg","demand_m3","ready_time","due_time","service_time",
        "priority","client_type","phone","email","tags"]
    with open(path, "w", newline="", encoding="utf-8") as f:
      w = csv.writer(f)
      w.writerow(COLS)
      for row in rows:
        w.writerow([row.get(c) for c in COLS])
    log_action("CLIENT_EXPORT", f"{len(rows)} clients → {path}")
    show_toast(self.window(), f"{len(rows)} clients exportés", "success")

  # ── Import ────────────────────────────────────────────────────────

  def _import_data(self):
    path, _ = QFileDialog.getOpenFileName(
      self, "Importer clients", "",
      "Fichiers supportés (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls)",
    )
    if not path:
      return
    dlg = _ImportMappingDialog(path, self)
    if dlg.exec() != QDialog.DialogCode.Accepted:
      return
    self._overlay.start("Import en cours…")
    t = _ImportThread(path, dlg.col_map, dlg.geocode, self)
    t.progress.connect(lambda msg, c, tot: self._overlay.start(f"{msg} ({c}/{tot})"))
    t.finished.connect(self._on_import_done)
    self._threads.append(t)
    t.start()

  def _on_import_done(self, report: dict):
    self._overlay.stop()
    self.refresh_data()
    _ImportReportDialog(report, self).exec()

  # ── Export ────────────────────────────────────────────────────────

  def _export_menu(self):
    menu = QMenu(self)
    menu.setStyleSheet(
      f"QMenu{{background:{C['panel']};color:{C['text']};border:1px solid {C['border']};"
      "border-radius:6px;padding:4px;}"
      f"QMenu::item{{padding:6px 18px;border-radius:4px;}}"
      f"QMenu::item:selected{{background:{C['hover']};}}"
    )
    for label, fn in [
      ("Exporter CSV",  self._export_csv),
      ("Exporter Excel", self._export_excel),
      ("Exporter JSON", self._export_json),
    ]:
      act = QAction(label, self)
      act.triggered.connect(fn)
      menu.addAction(act)
    btn = self.sender()
    menu.exec(btn.mapToGlobal(btn.rect().bottomLeft()) if btn else QCursor.pos())

  def _all_rows(self) -> list:
    conn = get_connection()
    rows = conn.execute(
      "SELECT * FROM clients WHERE archived=0 ORDER BY priority, name"
    ).fetchall()
    conn.close()
    return rows

  def _export_csv(self):
    path, _ = QFileDialog.getSaveFileName(self, "Exporter CSV", "clients.csv", "CSV (*.csv)")
    if not path:
      return
    rows = self._all_rows()
    COLS = ["id","name","company_name","address","latitude","longitude",
        "demand_kg","demand_m3","ready_time","due_time","service_time",
        "priority","client_type","phone","email","tags"]
    with open(path, "w", newline="", encoding="utf-8") as f:
      w = csv.writer(f)
      w.writerow(COLS)
      for row in rows:
        w.writerow([row.get(c) for c in COLS])
    log_action("CLIENT_EXPORT_CSV", f"{len(rows)} → {path}")
    show_toast(self.window(), f"{len(rows)} clients exportés (CSV)", "success")

  def _export_excel(self):
    if not HAS_OPENPYXL:
      show_toast(self.window(), "openpyxl requis pour Excel.", "error")
      return
    path, _ = QFileDialog.getSaveFileName(
      self, "Exporter Excel", "clients.xlsx", "Excel (*.xlsx)")
    if not path:
      return
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    COLS = ["id","name","company_name","address","latitude","longitude",
        "demand_kg","demand_m3","ready_time","due_time","service_time",
        "priority","client_type","phone","email","tags"]
    ws.append(COLS)
    for row in self._all_rows():
      ws.append([row.get(c) for c in COLS])
    wb.save(path)
    log_action("CLIENT_EXPORT_XLS", f"Excel → {path}")
    show_toast(self.window(), "Export Excel terminé", "success")

  def _export_json(self):
    path, _ = QFileDialog.getSaveFileName(
      self, "Exporter JSON", "clients.json", "JSON (*.json)")
    if not path:
      return
    data = [dict(r) for r in self._all_rows()]
    with open(path, "w", encoding="utf-8") as f:
      _json.dump(data, f, ensure_ascii=False, indent=2)
    log_action("CLIENT_EXPORT_JSON", f"JSON → {path}")
    show_toast(self.window(), f"{len(data)} clients exportés (JSON)", "success")

  # ── Map ───────────────────────────────────────────────────────────

  def _show_map(self):
    conn = get_connection()
    clients = [dict(r) for r in conn.execute(
      "SELECT id,name,latitude,longitude,client_type,priority"
      " FROM clients WHERE archived=0"
    ).fetchall()]
    conn.close()
    if not clients:
      show_toast(self.window(), "Aucun client à afficher.", "info")
      return
    _MapDialog(clients, self).exec()

  def _show_one_map(self, cid: int):
    conn = get_connection()
    row = conn.execute("SELECT * FROM clients WHERE id= ?", (cid,)).fetchone()
    conn.close()
    if row:
      _MapDialog([dict(row)], self).exec()

  # ── Anomalies ─────────────────────────────────────────────────────

  def _detect_anomalies(self):
    self._overlay.start("Analyse des données clients…")
    t = _AnomalyThread(self)
    t.finished.connect(self._on_anomalies)
    self._threads.append(t)
    t.start()

  def _on_anomalies(self, anomalies: list):
    self._overlay.stop()
    _AnomalyDialog(anomalies, self).exec()
    if anomalies:
      log_action("CLIENT_ANOMALY",
            f"{len(anomalies)} clients avec anomalies détectées")
