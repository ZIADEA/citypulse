"""
orders_widget.py — Gestion des commandes CityPulse Logistics v1.0
=================================================================
• 5 KPICards : En attente | Assignées | En cours | Livrées aujourd'hui | Échecs
• Tableau paginé (StatusBadge) + filtres + recherche
• Dialogue 4 onglets : Commande | Marchandises | Créneaux | Assignation
• Commandes récurrentes : CRUD templates + Générer semaine
• Import/Export CSV/Excel | Actions en lot
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import csv
import io
import logging
import re
from datetime import date, timedelta, datetime

# ── PyQt6 ────────────────────────────────────────────────────────────────────
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog,
    QFormLayout, QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox,
    QDateEdit, QTextEdit, QCheckBox, QMessageBox, QFrame,
    QStackedWidget, QTabWidget, QAbstractItemView, QMenu,
    QFileDialog, QSizePolicy, QScrollArea, QGroupBox, QCompleter,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QColor, QFont, QAction

# ── Local ─────────────────────────────────────────────────────────────────────
from ..database.db_manager import get_connection, log_action
from ..services.report_service import ReportService, REPORTLAB_OK
from .toast import show_toast
from .help_dialog import show_help
from .components import SectionHeader, SearchBar, KPICard, StatusBadge, PaginationBar, ConfirmDialog
from .components.confirm_dialog import _dialog_qss
from .lucide_icons import apply_action_button

try:
    import openpyxl as _openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
C = {
    "bg":    "#0D1B2A", "panel":  "#112240", "input":  "#1A2E4A",
    "accent":"#00D4FF", "success":"#00FF88", "warning":"#FFB800",
    "danger":"#FF4757", "text":   "#E8F4FD", "text2":  "#8899AA",
    "border":"#1E3A5F", "hover":  "#1A3A5C",
}
PAGE_SIZE = 80

_GRP_QSS = (
    f"QGroupBox{{color:{C['text2']};border:1px solid {C['border']};"
    "border-radius:6px;margin-top:16px;padding-top:12px;padding-bottom:8px;}}"
    f"QGroupBox::title{{subcontrol-origin:margin;subcontrol-position:top left;"
    f"left:10px;padding:0 4px;color:{C['accent']};font-weight:700;font-size:11px;}}"
)

# ── Status mapping ─────────────────────────────────────────────────────────────
_STATUS_LABEL = {
    "pending":     "En attente",
    "assigned":    "Assignée",
    "in_progress": "En cours",
    "delivered":   "Livrée",
    "failed":      "Échec",
    "cancelled":   "Annulée",
}
_STATUS_VARIANT = {
    "pending":     "warning",
    "assigned":    "info",
    "in_progress": "info",
    "delivered":   "success",
    "failed":      "danger",
    "cancelled":   "danger",
    # Valeurs parfois présentes en base / imports (non canoniques)
    "success":     "success",
    "completed":   "success",
    "complete":    "success",
    "in_transit":  "info",
}

def _normalize_order_status(raw) -> str:
    """Statut canonique pour affichage / KPI (gère alias et casses)."""
    s = (raw or "").strip().lower()
    if not s:
        return "pending"
    aliases = {
        "success": "delivered",
        "completed": "delivered",
        "complete": "delivered",
        "in_transit": "in_progress",
        "shipping": "in_progress",
        "failure": "failed",
    }
    return aliases.get(s, s)


def _priority_stars(p) -> str:
    """Même logique que la page Clients : priorité 1 = le plus urgent (5 étoiles affichées)."""
    try:
        p = max(1, min(5, int(p)))
    except (TypeError, ValueError):
        p = 3
    filled = 6 - p
    return "★" * filled + "☆" * (5 - filled)


_OP_TYPE = {
    "delivery": "Livraison",
    "pickup":   "Collecte",
    "collection": "Collecte",  # alias (ex. données démo / imports)
    "exchange": "Échange",
    "return":   "Retour",
}
_TEMP_OPTIONS   = ["ambient", "chilled", "frozen"]
_ADR_CLASSES    = ["", "1", "2", "3", "4", "4.1", "4.2", "4.3", "5.1", "5.2", "6", "7", "8", "9"]
_GOODS_CATS     = ["standard", "alimentaire", "médical", "chimique", "frais", "surgelé",
                   "fragile", "ADR", "valeur déclarée", "volumineux"]
_RECUR_TYPES    = ["daily", "weekly", "biweekly", "monthly"]
_RECUR_LABELS   = {"daily":"Quotidien","weekly":"Hebdomadaire","biweekly":"Bi-mensuel","monthly":"Mensuel"}
_WEEKDAYS       = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
_PRIORITIES     = [(1,"★★★★★ Critique"),(2,"★★★★ Haute"),(3,"★★★ Normale"),
                   (4,"★★ Basse"),(5,"★ Minimale")]

# ── Shared style snippets ──────────────────────────────────────────────────────
_INP_STYLE = (
    f"QLineEdit,QTextEdit,QSpinBox,QDoubleSpinBox,QDateEdit,QComboBox{{"
    f"background:{C['input']};color:{C['text']};border:1px solid {C['border']};"
    "border-radius:5px;padding:4px 8px;}"
    f"QComboBox::drop-down{{border:none;}}"
    f"QComboBox QAbstractItemView{{background:{C['panel']};color:{C['text']};"
    f"border:1px solid {C['border']};}}"
)
# Véhicule / chauffeur (commande) : flèche lisible + zone cliquable pour ouvrir la liste BDD
_ORDER_ASSIGN_COMBO_QSS = (
    f"QComboBox#orderAssignVehicle::drop-down,QComboBox#orderAssignDriver::drop-down{{"
    "subcontrol-origin:padding;subcontrol-position:top right;width:26px;"
    f"border-left:1px solid {C['border']};background:{C['panel']};}}"
)

_ORDER_UNASSIGNED_FK = 0


def _wire_order_assign_combo(cb: QComboBox) -> None:
    """Combo éditable branché sur le modèle : saisie filtre la liste issue de la BDD."""
    cb.setEditable(True)
    cb.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
    cb.setMaxVisibleItems(25)
    cb.setMinimumHeight(34)
    compl = QCompleter(cb.model(), cb)
    compl.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
    compl.setFilterMode(Qt.MatchFlag.MatchContains)
    compl.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
    cb.setCompleter(compl)

    def _on_compl_pick(text: str):
        idx = cb.findText(text)
        if idx >= 0:
            cb.setCurrentIndex(idx)
            _sync_assign_combo_line_edit(cb)

    compl.activated.connect(_on_compl_pick)


def _sync_assign_combo_line_edit(cb: QComboBox) -> None:
    """Qt : sur QComboBox éditable, le texte du lineEdit peut rester vide après setCurrentIndex."""
    if not cb.isEditable():
        return
    le = cb.lineEdit()
    if le is None:
        return
    le.setText(cb.currentText())
    le.setCursorPosition(0)


def _norm_order_fk(val):
    if val is None or val == _ORDER_UNASSIGNED_FK:
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _set_assign_combo_by_id(combo: QComboBox, raw_id, missing_label: str | None) -> None:
    combo.setCurrentIndex(0)
    kid = _norm_order_fk(raw_id)
    if kid is None:
        _sync_assign_combo_line_edit(combo)
        return
    idx = combo.findData(kid)
    if idx < 0:
        idx = combo.findData(str(kid))
    if idx < 0:
        combo.addItem(missing_label or f"#{kid} (réf. BDD)", kid)
        idx = combo.count() - 1
    combo.setCurrentIndex(idx)
    _sync_assign_combo_line_edit(combo)


def _fk_from_assign_combo(cb: QComboBox):
    """Retourne l'id BDD après choix liste / complétion, ou None si non assigné."""
    d = cb.currentData()
    n = _norm_order_fk(d)
    if n is not None:
        return n
    t = (cb.currentText() or "").strip()
    idx = cb.findText(t)
    if idx >= 0:
        return _norm_order_fk(cb.itemData(idx))
    return None


def _ref_prefix() -> str:
    return date.today().strftime("ORD-%Y%m%d-")


def _next_ref(conn) -> str:
    prefix = _ref_prefix()
    row = conn.execute(
        "SELECT COUNT(*) FROM orders WHERE reference LIKE ? ", (prefix + "%",)
    ).fetchone()
    n = (row[0] if row else 0) + 1
    return f"{prefix}{n:04d}"


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT THREAD
# ═══════════════════════════════════════════════════════════════════════════════

class _ImportThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)

    def __init__(self, path: str, col_map: dict, parent=None):
        super().__init__(parent)
        self.path = path; self.col_map = col_map

    def run(self):
        created = updated = errors = 0
        try:
            if self.path.endswith((".xlsx", ".xls")) and HAS_OPENPYXL:
                wb = _openpyxl.load_workbook(self.path, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                raw_rows = [
                    {headers[j]: (str(cell.value) if cell.value is not None else "")
                     for j, cell in enumerate(row)}
                    for row in ws.iter_rows(min_row=2)
                ]
                wb.close()
            else:
                with open(self.path, newline="", encoding="utf-8-sig") as f:
                    raw_rows = list(csv.DictReader(f))

            cm = self.col_map
            conn = get_connection()
            for i, row in enumerate(raw_rows, 1):
                try:
                    ref = str(row.get(cm.get("reference", "reference"), "") or "").strip()
                    client_name = str(row.get(cm.get("client", "client"), "") or "").strip()
                    kg   = float(row.get(cm.get("quantity_kg", "quantity_kg"), 0) or 0)
                    sched= str(row.get(cm.get("scheduled_date", "scheduled_date"), "") or "").strip()
                    notes= str(row.get(cm.get("notes", "notes"), "") or "").strip()

                    c_id = None
                    if client_name:
                        cr = conn.execute(
                            "SELECT id FROM clients WHERE name LIKE ? AND archived=0 LIMIT 1",
                            (f"%{client_name}%",),
                        ).fetchone()
                        if cr: c_id = cr[0]
                    if c_id is None:
                        cr = conn.execute("SELECT id FROM clients WHERE archived=0 LIMIT 1").fetchone()
                        if cr: c_id = cr[0]
                    if c_id is None:
                        errors += 1; continue

                    if not ref:
                        ref = _next_ref(conn)

                    existing = conn.execute(
                        "SELECT id FROM orders WHERE reference= ?", (ref,)
                    ).fetchone()
                    if existing:
                        conn.execute("UPDATE orders SET quantity_kg= ?,delivery_notes= ?,"
                                     "scheduled_date= ? WHERE id= ?",
                                     (kg, notes, sched, existing[0]))
                        updated += 1
                    else:
                        conn.execute("""
                            INSERT INTO orders
                            (reference,client_id,quantity_kg,delivery_notes,scheduled_date,status)
                            VALUES (?,?,?,?,?,'pending')
                        """, (ref, c_id, kg, notes, sched))
                        created += 1
                    if i % 20 == 0:
                        self.progress.emit(f"{i} lignes traitées…")
                except Exception as e:
                    errors += 1
                    logger.debug("Import row %d error: %s", i, e)

            conn.commit(); conn.close()
            self.finished.emit({"created": created, "updated": updated, "errors": errors})
        except Exception as e:
            self.finished.emit({"created": 0, "updated": 0, "errors": 1, "msg": str(e)})


# ═══════════════════════════════════════════════════════════════════════════════
# GENERATE WEEK THREAD
# ═══════════════════════════════════════════════════════════════════════════════

class _GenerateWeekThread(QThread):
    finished = pyqtSignal(int)
    error    = pyqtSignal(str)

    def __init__(self, depot_id=None, parent=None):
        super().__init__(parent)
        self.depot_id = depot_id

    def run(self):
        try:
            conn = get_connection()
            templates = conn.execute(
                "SELECT * FROM recurring_order_templates WHERE is_active=1"
            ).fetchall()
            today = date.today()
            monday = today - timedelta(days=today.weekday())
            created = 0
            _en_day = {"Mon":0,"Tue":1,"Wed":2,"Thu":3,"Fri":4,"Sat":5,"Sun":6}
            for tpl in templates:
                days_str = tpl["recurrence_days"] or ""
                active_days: list[int] = []
                # Essai format JSON ["Mon","Wed",...] (ancien format démo)
                try:
                    import json as _j
                    parsed = _j.loads(days_str)
                    if isinstance(parsed, list):
                        active_days = [_en_day[d] for d in parsed if d in _en_day]
                except Exception:
                    pass
                # Fallback format entiers "0,2,4"
                if not active_days:
                    try:
                        active_days = [int(x) for x in days_str.split(",") if x.strip().isdigit()]
                    except Exception:
                        pass
                # Défaut lun-ven si rien de lisible
                if not active_days:
                    active_days = list(range(5))

                for offset in active_days:
                    day_date = monday + timedelta(days=offset)
                    sched = day_date.strftime("%Y-%m-%d")
                    if conn.execute(
                        "SELECT 1 FROM orders WHERE client_id= ? AND scheduled_date= ? AND is_recurring=1 LIMIT 1",
                        (tpl["client_id"], sched),
                    ).fetchone():
                        continue
                    ref = _next_ref(conn)
                    conn.execute("""
                        INSERT INTO orders
                        (reference,client_id,quantity_kg,volume_m3,units_count,
                         goods_category,time_window_start,time_window_end,
                         visit_duration_minutes,priority,delivery_notes,
                         status,is_recurring,scheduled_date,depot_id)
                        VALUES (?,?,?,?,?,?,?,?,?,?,'Généré automatiquement',
                        'pending',1,?,?)
                    """, (
                        ref, tpl["client_id"],
                        tpl["quantity_kg"] or 0, tpl["volume_m3"] or 0,
                        tpl["units_count"] or 1, tpl["goods_category"] or "standard",
                        tpl["time_window_start"], tpl["time_window_end"],
                        tpl["visit_duration_minutes"] or 15,
                        tpl["priority"] or 5, sched, self.depot_id,
                    ))
                    created += 1
            conn.commit(); conn.close()
            log_action("ORDERS_GENERATE_WEEK", f"{created} commandes récurrentes créées")
            self.finished.emit(created)
        except Exception as e:
            self.error.emit(str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# RECURRING TEMPLATES DIALOG — CRUD
# ═══════════════════════════════════════════════════════════════════════════════

class _RecurringDialog(QDialog):

    def __init__(self, parent=None, gen_week_cb=None):
        super().__init__(parent)
        self._gen_week_cb = gen_week_cb
        self.setWindowTitle("Templates récurrents")
        self.resize(860, 520)
        self.setStyleSheet(_dialog_qss() + f"QDialog{{background:{C['bg']};color:{C['text']};}}")
        self._setup_ui()
        self._refresh()

    def _setup_ui(self):
        lo = QVBoxLayout(self)
        lo.setContentsMargins(16, 16, 16, 12)
        lo.setSpacing(8)

        hdr = QHBoxLayout()
        t = QLabel("Templates de commandes récurrentes")
        t.setStyleSheet(f"color:{C['text']};font-size:14px;font-weight:700;")
        hdr.addWidget(t); hdr.addStretch()
        add_btn = QPushButton("+ Nouveau template"); add_btn.setObjectName("primaryBtn")
        add_btn.setFixedHeight(32); add_btn.clicked.connect(self._add)
        hdr.addWidget(add_btn)
        lo.addLayout(hdr)

        self._table = QTableWidget()
        self._table.setColumnCount(8)
        self._table.setHorizontalHeaderLabels([
            "Nom", "Client", "Récurrence", "Jours", "Créneaux", "kg", "Actif", "Actions"
        ])
        self._table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(36)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setStyleSheet(
            f"QTableWidget{{background:{C['bg']};color:{C['text']};"
            f"gridline-color:{C['border']};border:1px solid {C['border']};alternate-background-color:#0F2035;}}"
            f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};"
            f"border:1px solid {C['border']};padding:4px 6px;font-size:11px;}}"
        )
        lo.addWidget(self._table, 1)

        footer = QHBoxLayout()
        if self._gen_week_cb:
            gen_btn = QPushButton(" Générer la semaine")
            gen_btn.setObjectName("primaryBtn")
            gen_btn.setFixedHeight(32)
            gen_btn.setToolTip("Crée les commandes de la semaine courante depuis les templates actifs")
            gen_btn.clicked.connect(self._on_gen_week)
            footer.addWidget(gen_btn)
        footer.addStretch()
        close_btn = QPushButton("Fermer"); close_btn.setObjectName("secondaryBtn")
        close_btn.setFixedHeight(32)
        close_btn.clicked.connect(self.accept)
        footer.addWidget(close_btn)
        lo.addLayout(footer)

    def _refresh(self):
        conn = get_connection()
        rows = conn.execute("""
            SELECT t.*, c.name as client_name
            FROM recurring_order_templates t
            LEFT JOIN clients c ON c.id=t.client_id
            ORDER BY t.name
        """).fetchall()
        conn.close()
        self._table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            def _it(v): i = QTableWidgetItem(str(v or ""));
            i_it = lambda v: QTableWidgetItem(str(v or ""))

            def _set(col, val, color=None):
                it = QTableWidgetItem(str(val or ""))
                it.setFlags(Qt.ItemFlag(it.flags().value & ~Qt.ItemFlag.ItemIsEditable.value))
                if color: it.setForeground(QColor(color))
                it.setData(Qt.ItemDataRole.UserRole, row["id"])
                self._table.setItem(r, col, it)

            _set(0, row["name"])
            _set(1, row["client_name"] or f"Client #{row['client_id']}", C["text2"])
            _set(2, _RECUR_LABELS.get(row["recurrence_type"] or "", row["recurrence_type"] or ""))
            days_str = row["recurrence_days"] or ""
            days_labels = ""
            try:
                import json as _j
                parsed = _j.loads(days_str)
                if isinstance(parsed, list):
                    # format ancien: ["Mon","Tue",...]
                    _en = {"Mon":0,"Tue":1,"Wed":2,"Thu":3,"Fri":4,"Sat":5,"Sun":6}
                    days_labels = ",".join(
                        _WEEKDAYS[_en[d]] for d in parsed if d in _en
                    )
            except Exception:
                pass
            if not days_labels:
                # format attendu: "0,1,2,..."
                try:
                    days_labels = ",".join(
                        _WEEKDAYS[int(d)] for d in days_str.split(",")
                        if d.strip().isdigit() and 0 <= int(d) < 7
                    )
                except Exception:
                    days_labels = days_str
            _set(3, days_labels)
            _set(4, f"{row.get('time_window_start') or ''}–{row.get('time_window_end') or ''}")
            _set(5, f"{row.get('quantity_kg', 0) or 0:.0f}")

            active_cell = QTableWidgetItem(" Actif" if row.get("is_active") else " Inactif")
            active_cell.setFlags(Qt.ItemFlag(active_cell.flags().value & ~Qt.ItemFlag.ItemIsEditable.value))
            active_cell.setForeground(QColor(C["success"] if row.get("is_active") else C["danger"]))
            self._table.setItem(r, 6, active_cell)

            self._table.setCellWidget(r, 7, self._make_actions(row["id"], dict(row)))

    def _on_gen_week(self):
        if self._gen_week_cb:
            self._gen_week_cb()

    def _make_actions(self, tid: int, row: dict) -> QWidget:
        w = QWidget(); lo = QHBoxLayout(w); lo.setContentsMargins(3,1,3,1); lo.setSpacing(3)
        for lucide_key, tip, fn, fg, hbg in [
            ("pencil", "Modifier",  lambda _, i=tid, d=row: self._edit(i, d), C["accent"], C["panel"]),
            ("trash-2", "Supprimer", lambda _, i=tid: self._delete(i), C["danger"], "#3A1020"),
        ]:
            btn = QPushButton(); btn.setFixedSize(28, 28)
            btn.setToolTip(tip); btn.setCursor(Qt.CursorShape.PointingHandCursor)
            apply_action_button(btn, lucide_key, fg, C["hover"], hbg, icon_px=16)
            btn.clicked.connect(fn); lo.addWidget(btn)
        return w

    def _add(self):
        dlg = _TemplateEditDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._refresh()

    def _edit(self, tid: int, data: dict):
        dlg = _TemplateEditDialog(self, data)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._refresh()

    def _delete(self, tid: int):
        if not ConfirmDialog.ask(self, "Supprimer", "Supprimer ce template ", "danger"):
            return
        conn = get_connection()
        conn.execute("DELETE FROM recurring_order_templates WHERE id= ?", (tid,))
        conn.commit(); conn.close()
        log_action("TEMPLATE_DELETE", f"Template #{tid} supprimé")
        self._refresh()


# ═══════════════════════════════════════════════════════════════════════════════
# TEMPLATE EDIT DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class _TemplateEditDialog(QDialog):

    def __init__(self, parent=None, data: dict = None):
        super().__init__(parent)
        self.data = data or {}
        self.setWindowTitle("Modifier template" if data else "Nouveau template")
        self.resize(520, 480)
        self.setStyleSheet(
            _dialog_qss()
            + f"QDialog{{background:{C['bg']};color:{C['text']};}}" + _INP_STYLE +
            f"QLabel{{background:transparent;color:{C['text']};}}"
            f"QGroupBox{{color:{C['text2']};border:1px solid {C['border']};"
            "border-radius:5px;margin-top:10px;padding-top:8px;}"
            f"QGroupBox::title{{subcontrol-origin:margin;left:8px;padding:0 4px;"
            f"color:{C['accent']};font-weight:700;}}"
        )
        self._setup_ui()

    def _lbl2(self, t):
        l = QLabel(t); l.setStyleSheet(f"color:{C['text2']};font-size:11px;background:transparent;")
        return l

    def _setup_ui(self):
        lo = QVBoxLayout(self); fl = QFormLayout()
        fl.setSpacing(8); fl.setContentsMargins(12, 12, 12, 8)
        d = self.data

        self._name = QLineEdit(d.get("name", ""))
        self._name.setPlaceholderText("Nom du template *")

        conn = get_connection()
        clients = conn.execute(
            "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
        ).fetchall()
        conn.close()
        self._client = QComboBox()
        self._client.setEditable(True)
        for c in clients:
            self._client.addItem(c["name"], c["id"])
        if d.get("client_id"):
            idx = self._client.findData(d["client_id"])
            if idx >= 0: self._client.setCurrentIndex(idx)

        self._rtype = QComboBox()
        for k, v in _RECUR_LABELS.items():
            self._rtype.addItem(v, k)
        if d.get("recurrence_type"):
            idx = self._rtype.findData(d["recurrence_type"])
            if idx >= 0: self._rtype.setCurrentIndex(idx)

        # Day checkboxes
        days_grp = QGroupBox("Jours actifs")
        days_grp.setStyleSheet(_GRP_QSS)
        days_lo = QHBoxLayout(days_grp)
        days_lo.setContentsMargins(8,12,8,4); days_lo.setSpacing(8)
        saved_days = set()
        try:
            saved_days = {int(x) for x in (d.get("recurrence_days") or "0,1,2,3,4").split(",") if x.strip().isdigit()}
        except Exception:
            saved_days = {0,1,2,3,4}
        self._day_cbs = []
        for i, day in enumerate(_WEEKDAYS):
            cb = QCheckBox(day); cb.setChecked(i in saved_days)
            cb.setStyleSheet(f"color:{C['text']};background:transparent;")
            days_lo.addWidget(cb); self._day_cbs.append(cb)

        self._kg   = QDoubleSpinBox(); self._kg.setRange(0,99999); self._kg.setDecimals(1)
        self._kg.setValue(float(d.get("quantity_kg") or 0))
        self._m3   = QDoubleSpinBox(); self._m3.setRange(0,9999);  self._m3.setDecimals(2)
        self._m3.setValue(float(d.get("volume_m3") or 0))
        self._tw_start = QLineEdit(d.get("time_window_start") or "08:00"); self._tw_start.setFixedWidth(60)
        self._tw_end   = QLineEdit(d.get("time_window_end") or "18:00");   self._tw_end.setFixedWidth(60)
        tw_row = QHBoxLayout()
        tw_row.addWidget(self._tw_start); tw_row.addWidget(QLabel("→")); tw_row.addWidget(self._tw_end)
        tw_row.addStretch()
        self._active = QCheckBox("Template actif")
        self._active.setChecked(bool(d.get("is_active", 1)))
        self._active.setStyleSheet(f"color:{C['text']};background:transparent;")

        fl.addRow(self._lbl2("Nom *"),       self._name)
        fl.addRow(self._lbl2("Client"),      self._client)
        fl.addRow(self._lbl2("Récurrence"),  self._rtype)
        fl.addRow(self._lbl2("Jours"),       days_grp)
        fl.addRow(self._lbl2("Quantité kg"), self._kg)
        fl.addRow(self._lbl2("Volume m³"),   self._m3)
        fl.addRow(self._lbl2("Créneaux"),    tw_row)
        fl.addRow("",                        self._active)

        lo.addLayout(fl)
        lo.addStretch(1)

        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("Annuler"); cancel.setObjectName("secondaryBtn")
        cancel.setFixedHeight(32); cancel.clicked.connect(self.reject)
        save = QPushButton("Sauvegarder"); save.setObjectName("primaryBtn")
        save.setFixedHeight(32); save.setMinimumWidth(110)
        save.clicked.connect(self._on_save)
        btn_row.addWidget(cancel); btn_row.addWidget(save)
        lo.addLayout(btn_row)

    def _on_save(self):
        if not self._name.text().strip():
            QMessageBox.warning(self, "Validation", "Nom obligatoire."); return
        days = ",".join(str(i) for i, cb in enumerate(self._day_cbs) if cb.isChecked())
        cid = self._client.currentData()
        vals = {
            "name":              self._name.text().strip(),
            "client_id":         cid,
            "recurrence_type":   self._rtype.currentData(),
            "recurrence_days":   days,
            "quantity_kg":       self._kg.value(),
            "volume_m3":         self._m3.value(),
            "time_window_start": self._tw_start.text().strip(),
            "time_window_end":   self._tw_end.text().strip(),
            "is_active":         int(self._active.isChecked()),
        }
        try:
            conn = get_connection()
            if self.data.get("id"):
                conn.execute("""
                    UPDATE recurring_order_templates SET
                    name= ?,client_id= ?,recurrence_type= ?,recurrence_days= ?,
                    quantity_kg= ?,volume_m3= ?,time_window_start= ?,
                    time_window_end= ?,is_active= ? WHERE id=?
                """, (*vals.values(), self.data["id"]))
                log_action("TEMPLATE_UPDATE", f"Template #{self.data['id']} modifié")
            else:
                conn.execute("""
                    INSERT INTO recurring_order_templates
                    (name,client_id,recurrence_type,recurrence_days,quantity_kg,
                     volume_m3,time_window_start,time_window_end,is_active)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, tuple(vals.values()))
                log_action("TEMPLATE_CREATE", f"Template '{vals['name']}' créé")
            conn.commit(); conn.close()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# ORDER DIALOG — 4 onglets
# ═══════════════════════════════════════════════════════════════════════════════

class _OrderDialog(QDialog):

    def __init__(self, parent=None, order: dict = None):
        super().__init__(parent)
        self.order = order or {}
        self.setWindowTitle("Modifier commande" if order else "Nouvelle commande")
        self.setMinimumSize(680, 560)
        self.resize(720, 580)
        self.setModal(True)
        self.setStyleSheet(
            _dialog_qss()
            + f"QDialog{{background:{C['bg']};color:{C['text']};}}"
            f"QTabWidget::pane{{background:{C['panel']};border:1px solid {C['border']};border-radius:6px;}}"
            f"QTabBar::tab{{background:{C['input']};color:{C['text2']};padding:8px 14px;"
            "border-top-left-radius:4px;border-top-right-radius:4px;margin-right:2px;font-size:12px;}"
            f"QTabBar::tab:selected{{background:{C['accent']};color:{C['bg']};font-weight:700;}}"
            f"QTabBar::tab:hover{{background:{C['hover']};}}" + _INP_STYLE +
            f"QLabel{{background:transparent;color:{C['text']};}}"
            f"QGroupBox{{color:{C['text2']};border:1px solid {C['border']};"
            "border-radius:5px;margin-top:10px;padding-top:8px;}"
            f"QGroupBox::title{{subcontrol-origin:margin;left:8px;padding:0 4px;"
            f"color:{C['accent']};font-weight:700;}}"
            + _ORDER_ASSIGN_COMBO_QSS
        )
        self._load_lists()
        self._setup_ui()

    def _load_lists(self):
        conn = get_connection()
        self._clients  = conn.execute(
            "SELECT id,name FROM clients WHERE archived=0 ORDER BY name").fetchall()
        self._vehicles = conn.execute(
            "SELECT id,registration,brand,allowed_adr FROM vehicles ORDER BY registration").fetchall()
        self._drivers  = conn.execute(
            "SELECT id,first_name,last_name FROM drivers ORDER BY last_name").fetchall()
        self._depots   = conn.execute(
            "SELECT id,name FROM depots ORDER BY name").fetchall()
        # Charger le transporteur si la commande est sous-traitée
        self._carrier_name = None
        cid = self.order.get("carrier_id")
        if cid:
            try:
                row = conn.execute(
                    "SELECT name FROM carriers WHERE id=?", (cid,)
                ).fetchone()
                if row:
                    self._carrier_name = row["name"]
            except Exception:
                pass
        conn.close()

    def _lbl(self, t): l = QLabel(t); l.setStyleSheet(f"color:{C['text2']};font-size:11px;background:transparent;"); return l
    def _le(self, v="", ph=""): w = QLineEdit(str(v) if v else ""); w.setPlaceholderText(ph); return w

    def _setup_ui(self):
        lo = QVBoxLayout(self); lo.setContentsMargins(16,16,16,12); lo.setSpacing(12)
        tabs = QTabWidget()
        tabs.addTab(self._tab_commande(),    "  Commande  ")
        tabs.addTab(self._tab_marchandises(),"  Marchandises  ")
        tabs.addTab(self._tab_creneaux(),    "  Créneaux  ")
        tabs.addTab(self._tab_assignation(), "  Assignation  ")
        lo.addWidget(tabs, 1)

        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("Annuler"); cancel.setObjectName("secondaryBtn")
        cancel.setFixedHeight(34); cancel.clicked.connect(self.reject)
        save   = QPushButton("Sauvegarder"); save.setObjectName("primaryBtn")
        save.setFixedHeight(34); save.setMinimumWidth(120); save.clicked.connect(self._on_save)
        btn_row.addWidget(cancel); btn_row.addWidget(save)
        lo.addLayout(btn_row)

    # ── Tab 0 : Commande ──────────────────────────────────────────────
    def _tab_commande(self) -> QWidget:
        w = QWidget(); fl = QFormLayout(w); fl.setSpacing(10); fl.setContentsMargins(16,16,16,8)
        o = self.order

        conn = get_connection()
        auto_ref = _next_ref(conn); conn.close()
        self._ref    = self._le(o.get("reference") or auto_ref)
        self._ref.setPlaceholderText("Référence auto-générée")

        self._client = QComboBox(); self._client.setEditable(True)
        for c in self._clients:
            self._client.addItem(c["name"], c["id"])
        if o.get("client_id"):
            idx = self._client.findData(o["client_id"])
            if idx >= 0: self._client.setCurrentIndex(idx)

        self._op_type = QComboBox()
        for k, v in _OP_TYPE.items(): self._op_type.addItem(v, k)
        if o.get("operation_type"):
            idx = self._op_type.findData(o["operation_type"])
            if idx >= 0: self._op_type.setCurrentIndex(idx)

        self._status = QComboBox()
        for k, v in _STATUS_LABEL.items(): self._status.addItem(v, k)
        if o.get("status"):
            idx = self._status.findData(o["status"])
            if idx >= 0: self._status.setCurrentIndex(idx)

        self._date = QDateEdit(); self._date.setCalendarPopup(True); self._date.setDisplayFormat("dd/MM/yyyy")
        sched = o.get("scheduled_date") or date.today().isoformat()
        try: self._date.setDate(QDate.fromString(sched, "yyyy-MM-dd"))
        except Exception: self._date.setDate(QDate.currentDate())

        self._priority = QComboBox()
        for val, label in _PRIORITIES: self._priority.addItem(label, val)
        try:
            p = int(round(float(o.get("priority", 3) or 3)))
        except (TypeError, ValueError):
            p = 3
        p = max(1, min(5, p))
        idx = self._priority.findData(p)
        self._priority.setCurrentIndex(max(0, idx))

        fl.addRow(self._lbl("Référence"),  self._ref)
        fl.addRow(self._lbl("Client *"),   self._client)
        fl.addRow(self._lbl("Type"),       self._op_type)
        fl.addRow(self._lbl("Statut"),     self._status)
        fl.addRow(self._lbl("Date prévue"),self._date)
        fl.addRow(self._lbl("Priorité"),   self._priority)
        return w

    # ── Tab 1 : Marchandises ──────────────────────────────────────────
    def _tab_marchandises(self) -> QWidget:
        w = QWidget(); fl = QFormLayout(w); fl.setSpacing(10); fl.setContentsMargins(16,16,16,8)
        o = self.order

        self._kg    = QDoubleSpinBox(); self._kg.setRange(0,99999); self._kg.setDecimals(1)
        self._kg.setValue(float(o.get("quantity_kg") or 0))
        self._m3    = QDoubleSpinBox(); self._m3.setRange(0,9999);  self._m3.setDecimals(2)
        self._m3.setValue(float(o.get("volume_m3") or 0))
        self._units = QSpinBox();       self._units.setRange(0, 99999)
        try:
            uc = int(round(float(o.get("units_count") or 1)))
        except (TypeError, ValueError):
            uc = 1
        self._units.setValue(max(0, min(99999, uc)))

        self._cat = QComboBox()
        for c in _GOODS_CATS: self._cat.addItem(c)
        idx = self._cat.findText(o.get("goods_category") or "standard")
        if idx >= 0: self._cat.setCurrentIndex(idx)

        self._temp = QComboBox()
        for t in _TEMP_OPTIONS: self._temp.addItem(t)
        idx = self._temp.findText(o.get("temperature_required") or "ambient")
        if idx >= 0: self._temp.setCurrentIndex(idx)

        self._adr = QComboBox()
        for a in _ADR_CLASSES: self._adr.addItem(a if a else "— aucun ADR")
        idx = self._adr.findText(o.get("adr_class") or "")
        self._adr.setCurrentIndex(max(0, idx))

        self._value = QDoubleSpinBox(); self._value.setRange(0,9999999); self._value.setDecimals(2)
        self._value.setPrefix("€ "); self._value.setValue(float(o.get("declared_value") or 0))

        fl.addRow(self._lbl("Quantité (kg)"),    self._kg)
        fl.addRow(self._lbl("Volume (m³)"),       self._m3)
        fl.addRow(self._lbl("Unités / palettes"), self._units)
        fl.addRow(self._lbl("Catégorie"),         self._cat)
        fl.addRow(self._lbl("Température"),       self._temp)
        fl.addRow(self._lbl("Classe ADR"),        self._adr)
        fl.addRow(self._lbl("Valeur déclarée"),   self._value)
        return w

    # ── Tab 2 : Créneaux ──────────────────────────────────────────────
    def _tab_creneaux(self) -> QWidget:
        w = QWidget(); fl = QFormLayout(w); fl.setSpacing(10); fl.setContentsMargins(16,16,16,8)
        o = self.order

        tw1_row = QHBoxLayout()
        self._tw1s = self._le(o.get("time_window_start") or "08:00", "HH:MM"); self._tw1s.setFixedWidth(70)
        self._tw1e = self._le(o.get("time_window_end") or "18:00",   "HH:MM"); self._tw1e.setFixedWidth(70)
        tw1_row.addWidget(self._tw1s); tw1_row.addWidget(QLabel("→")); tw1_row.addWidget(self._tw1e); tw1_row.addStretch()

        tw2_row = QHBoxLayout()
        self._tw2s = self._le(o.get("time_window2_start") or "", "HH:MM (optionnel)"); self._tw2s.setFixedWidth(90)
        self._tw2e = self._le(o.get("time_window2_end") or "",   "HH:MM (optionnel)"); self._tw2e.setFixedWidth(90)
        tw2_row.addWidget(self._tw2s); tw2_row.addWidget(QLabel("→")); tw2_row.addWidget(self._tw2e); tw2_row.addStretch()

        self._visit_dur = QSpinBox(); self._visit_dur.setRange(1, 240); self._visit_dur.setSuffix(" min")
        try:
            vd = int(round(float(o.get("visit_duration_minutes") or 15)))
        except (TypeError, ValueError):
            vd = 15
        self._visit_dur.setValue(max(1, min(240, vd)))

        self._notes = QTextEdit()
        self._notes.setMaximumHeight(64); self._notes.setPlaceholderText("Instructions de livraison…")
        self._notes.setText(o.get("delivery_notes") or "")
        self._notes.setStyleSheet(
            f"QTextEdit{{background:{C['input']};color:{C['text']};"
            f"border:1px solid {C['border']};border-radius:5px;padding:4px;}}"
        )

        self._access = self._le(o.get("access_instructions") or "", "Code/instructions d'accès")

        fl.addRow(self._lbl("Créneau 1 *"),         tw1_row)
        fl.addRow(self._lbl("Créneau 2 (opt.)"),    tw2_row)
        fl.addRow(self._lbl("Durée visite"),         self._visit_dur)
        fl.addRow(self._lbl("Instructions"),         self._notes)
        fl.addRow(self._lbl("Code accès"),           self._access)
        return w

    # ── Tab 3 : Assignation ───────────────────────────────────────────
    def _tab_assignation(self) -> QWidget:
        w = QWidget(); root_lo = QVBoxLayout(w); root_lo.setContentsMargins(0,0,0,0); root_lo.setSpacing(0)

        # Bannière sous-traitance
        if self._carrier_name:
            banner = QLabel(f"  🚛  Sous-traité à : {self._carrier_name}")
            banner.setStyleSheet(
                f"background:{C['panel']};color:{C['accent']};"
                f"font-size:12px;font-weight:600;padding:8px 16px;"
                f"border-bottom:1px solid {C['border']};"
            )
            root_lo.addWidget(banner)

        inner = QWidget(); fl = QFormLayout(inner); fl.setSpacing(10); fl.setContentsMargins(16,16,16,8)
        root_lo.addWidget(inner)
        o = self.order

        self._vehicle = QComboBox()
        self._vehicle.setObjectName("orderAssignVehicle")
        self._vehicle.addItem("— Non assigné", _ORDER_UNASSIGNED_FK)
        for v in self._vehicles:
            reg = (v.get("registration") or "").strip()
            br = (v.get("brand") or "").strip()
            if reg and br:
                lbl = f"{reg} · {br}"
            else:
                lbl = reg or br or f"Véhicule #{v['id']}"
            self._vehicle.addItem(lbl, int(v["id"]))
        _wire_order_assign_combo(self._vehicle)
        self._vehicle.lineEdit().setPlaceholderText("Rechercher un véhicule (immat., marque)…")
        _set_assign_combo_by_id(
            self._vehicle, o.get("vehicle_id"),
            missing_label=(
                f"Véhicule #{o.get('vehicle_id')} (non listé)"
                if o.get("vehicle_id") else None
            ),
        )
        self._vehicle.currentIndexChanged.connect(
            lambda *_: _sync_assign_combo_line_edit(self._vehicle)
        )
        self._vehicle.currentIndexChanged.connect(self._check_compat)

        self._compat_lbl = QLabel("")
        self._compat_lbl.setStyleSheet(f"font-size:11px;background:transparent;")

        self._driver = QComboBox()
        self._driver.setObjectName("orderAssignDriver")
        self._driver.addItem("— Non assigné", _ORDER_UNASSIGNED_FK)
        for dr in self._drivers:
            fn = (dr.get("first_name") or "").strip()
            ln = (dr.get("last_name") or "").strip()
            nm = f"{fn} {ln}".strip() or f"Chauffeur #{dr['id']}"
            self._driver.addItem(nm, int(dr["id"]))
        _wire_order_assign_combo(self._driver)
        self._driver.lineEdit().setPlaceholderText("Rechercher un chauffeur…")
        _set_assign_combo_by_id(
            self._driver, o.get("driver_id"),
            missing_label=(
                f"Chauffeur #{o.get('driver_id')} (non listé)"
                if o.get("driver_id") else None
            ),
        )
        self._driver.currentIndexChanged.connect(
            lambda *_: _sync_assign_combo_line_edit(self._driver)
        )

        self._depot = QComboBox()
        self._depot.addItem("— Non défini", None)
        for dp in self._depots:
            self._depot.addItem(dp["name"], dp["id"])
        if o.get("depot_id"):
            idx = self._depot.findData(o["depot_id"])
            if idx >= 0: self._depot.setCurrentIndex(idx)

        fl.addRow(self._lbl("Véhicule"),     self._vehicle)
        fl.addRow("",                        self._compat_lbl)
        fl.addRow(self._lbl("Chauffeur"),    self._driver)
        fl.addRow(self._lbl("Dépôt"),        self._depot)
        return w

    def _check_compat(self):
        vid = _fk_from_assign_combo(self._vehicle)
        if not vid:
            self._compat_lbl.setText("")
            return
        adr_needed = ""
        try: adr_needed = getattr(self, "_adr", None) and self._adr.currentText() or ""
        except Exception: pass
        temp_needed = ""
        try: temp_needed = self._temp.currentText()
        except Exception: pass

        issues = []
        try:
            conn = get_connection(); veh = conn.execute(
                "SELECT allowed_adr,temperature_capability FROM vehicles WHERE id= ?", (vid,)
            ).fetchone(); conn.close()
            if veh:
                if adr_needed and adr_needed not in ("", "— aucun ADR"):
                    allowed = str(veh[0] or "")
                    if adr_needed not in allowed and allowed not in ("all", "*"):
                        issues.append(f"ADR {adr_needed} non autorisé sur ce véhicule")
                if temp_needed == "frozen":
                    tc = str(veh[1] or "").lower()
                    if "frozen" not in tc and "surgel" not in tc:
                        issues.append("Véhicule non équipé pour surgelés")
        except Exception:
            pass
        if issues:
            self._compat_lbl.setText(" " + " | ".join(issues))
            self._compat_lbl.setStyleSheet(f"color:{C['warning']};font-size:11px;background:transparent;")
        else:
            self._compat_lbl.setText(" Compatible")
            self._compat_lbl.setStyleSheet(f"color:{C['success']};font-size:11px;background:transparent;")

    # ── Save ──────────────────────────────────────────────────────────
    def _on_save(self):
        if not self._client.currentData():
            QMessageBox.warning(self, "Validation", "Sélectionnez un client."); return
        self.accept()

    def get_data(self) -> dict:
        adr_raw = ""
        try:
            adr_raw = self._adr.currentText()
            if adr_raw.startswith("—"): adr_raw = ""
        except Exception: pass
        return {
            "reference":             self._ref.text().strip(),
            "client_id":             self._client.currentData(),
            "operation_type":        self._op_type.currentData(),
            "status":                self._status.currentData(),
            "scheduled_date":        self._date.date().toString("yyyy-MM-dd"),
            "priority":              self._priority.currentData(),
            "quantity_kg":           self._kg.value(),
            "volume_m3":             self._m3.value(),
            "units_count":           self._units.value(),
            "goods_category":        self._cat.currentText(),
            "temperature_required":  self._temp.currentText(),
            "adr_class":             adr_raw,
            "declared_value":        self._value.value(),
            "time_window_start":     self._tw1s.text().strip(),
            "time_window_end":       self._tw1e.text().strip(),
            "time_window2_start":    self._tw2s.text().strip() or None,
            "time_window2_end":      self._tw2e.text().strip() or None,
            "visit_duration_minutes":self._visit_dur.value(),
            "delivery_notes":        self._notes.toPlainText().strip(),
            "access_instructions":   self._access.text().strip(),
            "vehicle_id":            _fk_from_assign_combo(self._vehicle),
            "driver_id":             _fk_from_assign_combo(self._driver),
            "depot_id":              self._depot.currentData(),
        }


# ═══════════════════════════════════════════════════════════════════════════════
# ORDERS WIDGET — Page principale
# ═══════════════════════════════════════════════════════════════════════════════

class OrdersWidget(QWidget):

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self._page       = 0
        self._all_rows   = []
        self._threads: list = []
        self._setup_ui()

    # ── UI ────────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(4, 4, 4, 8)
        root.setSpacing(14)

        self._header = SectionHeader(
            title="Gestion des Commandes",
            subtitle="Créez, assignez et suivez toutes vos commandes de livraison",
            action_text="+ Nouvelle commande",
            action_callback=self._add_order,
        )
        root.addWidget(self._header)

        # ── KPI Cards ─────────────────────────────────────────────────
        kpi_row = QHBoxLayout(); kpi_row.setSpacing(14)
        self._kpi_pending  = KPICard("En attente",          "0", icon="Attente")
        self._kpi_assigned = KPICard("Assignées",           "0", icon="")
        self._kpi_progress = KPICard("En cours",            "0", icon="")
        self._kpi_done     = KPICard("Livrées aujourd'hui", "0", icon="")
        self._kpi_failed   = KPICard("Échecs",              "0", icon="")
        for kpi in [self._kpi_pending, self._kpi_assigned, self._kpi_progress,
                    self._kpi_done, self._kpi_failed]:
            kpi_row.addWidget(kpi)
        root.addLayout(kpi_row)

        # ── Toolbar ───────────────────────────────────────────────────
        toolbar = QHBoxLayout(); toolbar.setSpacing(6)

        self._search = SearchBar(placeholder="Référence, client, notes…")
        self._search.setMaximumWidth(260)
        self._search.search_changed.connect(self._on_search)
        toolbar.addWidget(self._search)

        self._filter_status = QComboBox(); self._filter_status.setFixedWidth(130)
        self._filter_status.addItem("Tous les statuts", "")
        for k, v in _STATUS_LABEL.items(): self._filter_status.addItem(v, k)
        self._filter_status.currentIndexChanged.connect(self._on_search)
        self._filter_status.setStyleSheet(
            f"QComboBox{{background:{C['input']};color:{C['text']};"
            f"border:1px solid {C['border']};border-radius:5px;padding:4px 8px;}}"
            f"QComboBox QAbstractItemView{{background:{C['panel']};color:{C['text']};"
            f"border:1px solid {C['border']};}}"
        )
        toolbar.addWidget(self._filter_status)

        _S = (
            f"QPushButton{{background:{C['input']};color:{C['text']};"
            f"border:1px solid {C['border']};border-radius:5px;font-size:12px;padding:4px 10px;}}"
            f"QPushButton:hover{{background:{C['hover']};border-color:{C['accent']};}}"
        )
        for label, tip, fn in [
            (" Templates récurrents", "CRUD templates",  self._show_recurring),
            (" Exporter",            "Export CSV/Excel", self._export),
            (" Importer",            "Import CSV/Excel", self._import),
        ]:
            btn = QPushButton(label); btn.setFixedHeight(30); btn.setToolTip(tip)
            btn.setStyleSheet(_S); btn.clicked.connect(fn)
            toolbar.addWidget(btn)

        toolbar.addStretch()
        self._count_lbl = QLabel("0 commandes")
        self._count_lbl.setStyleSheet(f"color:{C['text2']};font-size:12px;")
        toolbar.addWidget(self._count_lbl)
        toolbar.addSpacing(4)
        _hb = QPushButton()
        _hb.setFixedSize(30, 30)
        _hb.setToolTip("Aide — Commandes")
        _hb.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_action_button(_hb, "help-circle", "#7FA8C0", "#1A2E4A", "#1A3A5C", 18)
        _hb.clicked.connect(lambda: show_help(self, "orders"))
        toolbar.addWidget(_hb)
        root.addLayout(toolbar)

        # ── Batch actions bar ─────────────────────────────────────────
        self._batch_bar = QFrame()
        self._batch_bar.setVisible(False)
        self._batch_bar.setStyleSheet(
            f"QFrame{{background:{C['panel']};border:1px solid {C['accent']};"
            "border-radius:5px;padding:4px;}}"
        )
        bb = QHBoxLayout(self._batch_bar); bb.setContentsMargins(8,4,8,4); bb.setSpacing(6)
        self._batch_lbl = QLabel("0 sélectionnée(s)")
        self._batch_lbl.setStyleSheet(f"color:{C['accent']};font-weight:600;font-size:12px;background:transparent;")
        bb.addWidget(self._batch_lbl)
        for label, fn in [
            (" Marquer livrées", self._batch_deliver),
            ("⟳ Réassigner",      self._batch_reassign),
            (" Supprimer",       self._batch_delete),
        ]:
            b = QPushButton(label); b.setObjectName("secondaryBtn")
            b.setFixedHeight(26); b.clicked.connect(fn); bb.addWidget(b)
        bb.addStretch()
        root.addWidget(self._batch_bar)

        # ── Table ─────────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setColumnCount(9)
        self._table.setHorizontalHeaderLabels([
            "Réf", "Client", "Type", "Statut", "Date",
            "kg", "ADR", "Priorité", "Actions",
        ])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed);   self._table.setColumnWidth(0, 130)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for col, w in [(2,90),(3,110),(4,90),(5,60),(6,55),(7,80),(8,105)]:
            self._table.setColumnWidth(col, w)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(36)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setSortingEnabled(True)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._context_menu)
        self._table.doubleClicked.connect(self._on_dblclick)
        self._table.itemSelectionChanged.connect(self._on_selection_changed)
        self._table.setStyleSheet(
            f"QTableWidget{{background:{C['bg']};color:{C['text']};"
            f"gridline-color:{C['border']};border:none;alternate-background-color:#0F2035;}}"
            f"QTableWidget::item:selected{{background:{C['hover']};color:{C['accent']};}}"
            f"QHeaderView::section{{background:{C['panel']};color:{C['text2']};"
            f"border:1px solid {C['border']};padding:4px 6px;font-size:11px;font-weight:600;}}"
        )
        root.addWidget(self._table, 1)

        # ── Pagination ────────────────────────────────────────────────
        self._pagination = PaginationBar(page_size=PAGE_SIZE)
        self._pagination.page_changed.connect(lambda pg, off, lim: self._on_page(pg))
        root.addWidget(self._pagination)

    def retranslate_ui(self, lang: str):
        from app.i18n import tr
        if hasattr(self, "_header"):
            self._header.set_title(tr("section.orders", lang))

    # ── Data ──────────────────────────────────────────────────────────

    def refresh_data(self):
        self._page = 0
        self._load_kpis()
        self._load_rows()

    def _load_kpis(self):
        try:
            today = date.today().isoformat()
            conn = get_connection()
            cur = conn.cursor()
            n_pending = cur.execute(
                """SELECT COUNT(*) FROM orders WHERE archived=0 AND (
                    status IS NULL OR trim(status)='' OR lower(trim(status))='pending'
                )"""
            ).fetchone()[0]
            n_assigned = cur.execute(
                """SELECT COUNT(*) FROM orders WHERE archived=0
                   AND lower(trim(coalesce(status,'')))='assigned'"""
            ).fetchone()[0]
            n_progress = cur.execute(
                """SELECT COUNT(*) FROM orders WHERE archived=0
                   AND lower(trim(coalesce(status,''))) IN
                   ('in_progress','in_transit','shipping')"""
            ).fetchone()[0]
            n_done = cur.execute(
                """SELECT COUNT(*) FROM orders WHERE archived=0 AND scheduled_date= ?
                   AND lower(trim(coalesce(status,''))) IN
                   ('delivered','success','completed','complete')""",
                (today,),
            ).fetchone()[0]
            n_failed = cur.execute(
                """SELECT COUNT(*) FROM orders WHERE archived=0
                   AND lower(trim(coalesce(status,''))) IN ('failed','failure')"""
            ).fetchone()[0]
            conn.close()
            self._kpi_pending.set_value(str(n_pending))
            self._kpi_assigned.set_value(str(n_assigned))
            self._kpi_progress.set_value(str(n_progress))
            self._kpi_done.set_value(str(n_done))
            self._kpi_failed.set_value(str(n_failed))
        except Exception as e:
            logger.warning("KPI commandes : %s", e, exc_info=True)

    def _load_rows(self):
        search = ""
        try:
            search = self._search.get_text().strip()
        except Exception:
            pass
        status_filter = self._filter_status.currentData() or ""

        try:
            conn = get_connection()
            params = []
            where = ["o.archived=0"]
            if search:
                s = f"%{search}%"
                where.append("(o.reference LIKE ? OR c.name LIKE ? OR o.delivery_notes LIKE ? )")
                params.extend([s, s, s])
            if status_filter:
                where.append("o.status= ?"); params.append(status_filter)
            sql = (
                "SELECT o.*,c.name AS client_name FROM orders o"
                " LEFT JOIN clients c ON c.id=o.client_id"
                f" WHERE {' AND '.join(where)}"
                " ORDER BY o.scheduled_date DESC, o.priority ASC, o.id DESC"
            )
            self._all_rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
            conn.close()
        except Exception as e:
            logger.debug("Orders load error: %s", e)
            self._all_rows = []

        total = len(self._all_rows)
        pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        self._page = min(self._page, pages - 1)
        self._count_lbl.setText(
            f"{total} commande{'s' if total != 1 else ''}"
        )
        self._pagination.update_total(total)
        self._fill_table()

    def _fill_table(self):
        start = self._page * PAGE_SIZE
        rows  = self._all_rows[start:start + PAGE_SIZE]
        self._table.setSortingEnabled(False)
        self._table.blockSignals(True)
        self._table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            def _it(val, color=None):
                it = QTableWidgetItem(str(val) if val is not None else "")
                it.setFlags(Qt.ItemFlag(it.flags().value & ~Qt.ItemFlag.ItemIsEditable.value))
                if color: it.setForeground(QColor(color))
                return it

            ref_it = _it(row.get("reference") or "")
            ref_it.setData(Qt.ItemDataRole.UserRole, row["id"])
            ref_it.setFont(QFont("Consolas", 9))
            self._table.setItem(r, 0, ref_it)
            self._table.setItem(r, 1, _it(row.get("client_name") or ""))

            op_key = str(row.get("operation_type") or "delivery").strip().lower()
            op_lbl = _OP_TYPE.get(op_key) or (row.get("operation_type") or "—")
            self._table.setItem(r, 2, _it(op_lbl, C["text2"]))

            # StatusBadge via setCellWidget (statut normalisé : success → Livrée, etc.)
            st = _normalize_order_status(row.get("status"))
            _var = _STATUS_VARIANT.get(st, "neutral")
            _lbl = _STATUS_LABEL.get(st, st)
            badge = StatusBadge(_var, _lbl)
            cw = QWidget(); cl = QHBoxLayout(cw); cl.setContentsMargins(4,2,4,2)
            cl.addWidget(badge); cl.addStretch()
            self._table.setCellWidget(r, 3, cw)

            self._table.setItem(r, 4, _it(row.get("scheduled_date") or ""))
            self._table.setItem(r, 5, _it(f"{float(row.get('quantity_kg') or 0):.0f}"))
            self._table.setItem(r, 6, _it(row.get("adr_class") or "—", C["warning"] if row.get("adr_class") else None))
            self._table.setItem(
                r, 7, _it(_priority_stars(row.get("priority")), color="#FFB800"),
            )
            self._table.setCellWidget(r, 8, self._make_actions(row["id"]))

        self._table.blockSignals(False)
        self._table.setSortingEnabled(True)

    def _on_search(self, *_):
        self._page = 0
        self._load_rows()

    def _on_page(self, p: int):
        self._page = p
        self._fill_table()

    def _on_selection_changed(self):
        sel = self._table.selectionModel().selectedRows()
        n = len(sel)
        self._batch_bar.setVisible(n > 1)
        self._batch_lbl.setText(f"{n} sélectionnée(s)")

    def _on_dblclick(self, idx):
        item = self._table.item(idx.row(), 0)
        if item:
            oid = item.data(Qt.ItemDataRole.UserRole)
            if oid: self._edit_order(oid)

    # ── Action widgets ────────────────────────────────────────────────

    def _export_delivery_note(self, oid: int):
        if not REPORTLAB_OK:
            QMessageBox.warning(self, "PDF", "Installez reportlab (pip install reportlab).")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Bon de livraison", f"BL_{oid}.pdf", "PDF (*.pdf)")
        if not path:
            return
        try:
            ReportService().generate_delivery_note(oid, path)
            show_toast(self.window(), "Bon de livraison généré.", "success")
            log_action("ORDER_DELIVERY_NOTE", f"order={oid} → {path}")
        except Exception as e:
            QMessageBox.warning(self, "Bon de livraison", str(e))

    def _make_actions(self, oid: int) -> QWidget:
        w = QWidget(); lo = QHBoxLayout(w); lo.setContentsMargins(3,1,3,1); lo.setSpacing(3)
        for lucide_key, tip, fn, fg, hbg in [
            ("pencil", "Modifier",    lambda _, i=oid: self._edit_order(i), C["accent"], C["panel"]),
            ("copy", "Dupliquer",   lambda _, i=oid: self._duplicate(i), C["text2"], C["panel"]),
            ("file-text", "Bon de livraison (PDF)", lambda _, i=oid: self._export_delivery_note(i), C["success"], C["panel"]),
            ("trash-2", "Supprimer",   lambda _, i=oid: self._delete_order(i), C["danger"], "#3A1020"),
        ]:
            btn = QPushButton(); btn.setFixedSize(28, 28)
            btn.setToolTip(tip); btn.setCursor(Qt.CursorShape.PointingHandCursor)
            apply_action_button(btn, lucide_key, fg, C["hover"], hbg, icon_px=16)
            btn.clicked.connect(fn); lo.addWidget(btn)
        return w

    def _context_menu(self, pos):
        row = self._table.rowAt(pos.y())
        if row < 0: return
        item = self._table.item(row, 0)
        oid = item.data(Qt.ItemDataRole.UserRole) if item else None
        if not oid: return

        menu = QMenu(self)
        menu.setStyleSheet(
            f"QMenu{{background:{C['panel']};color:{C['text']};border:1px solid {C['border']};"
            "border-radius:6px;padding:4px;}}"
            f"QMenu::item{{padding:6px 18px;border-radius:4px;}}"
            f"QMenu::item:selected{{background:{C['hover']};}}"
        )
        for label, fn in [
            ("  Modifier",    lambda: self._edit_order(oid)),
            ("  Dupliquer",   lambda: self._duplicate(oid)),
            (None, None),
            ("  Supprimer",   lambda: self._delete_order(oid)),
        ]:
            if label is None: menu.addSeparator()
            else:
                act = QAction(label, self); act.triggered.connect(fn); menu.addAction(act)
        menu.exec(self._table.viewport().mapToGlobal(pos))

    # ── CRUD ──────────────────────────────────────────────────────────

    def _add_order(self):
        dlg = _OrderDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        data = dlg.get_data()
        conn = get_connection()
        if not data["reference"]:
            data["reference"] = _next_ref(conn)
        conn.execute("""
            INSERT INTO orders
            (reference,client_id,operation_type,status,scheduled_date,priority,
             quantity_kg,volume_m3,units_count,goods_category,temperature_required,
             adr_class,declared_value,time_window_start,time_window_end,
             time_window2_start,time_window2_end,visit_duration_minutes,
             delivery_notes,access_instructions,vehicle_id,driver_id,depot_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data["reference"], data["client_id"], data["operation_type"], data["status"],
            data["scheduled_date"], data["priority"], data["quantity_kg"], data["volume_m3"],
            data["units_count"], data["goods_category"], data["temperature_required"],
            data["adr_class"], data["declared_value"], data["time_window_start"],
            data["time_window_end"], data["time_window2_start"], data["time_window2_end"],
            data["visit_duration_minutes"], data["delivery_notes"], data["access_instructions"],
            data["vehicle_id"], data["driver_id"], data["depot_id"],
        ))
        conn.commit(); conn.close()
        log_action("ORDER_CREATE", f"Commande {data['reference']} créée")
        show_toast(self.window(), f"Commande {data['reference']} créée", "success")
        self.refresh_data()

    def _edit_order(self, oid: int):
        conn = get_connection()
        row = conn.execute("SELECT * FROM orders WHERE id= ?", (oid,)).fetchone()
        conn.close()
        if not row: return
        dlg = _OrderDialog(self, dict(row))
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        data = dlg.get_data()
        conn = get_connection()
        conn.execute("""
            UPDATE orders SET
            reference= ?,client_id= ?,operation_type= ?,status= ?,scheduled_date= ?,
            priority= ?,quantity_kg= ?,volume_m3= ?,units_count= ?,goods_category= ?,
            temperature_required= ?,adr_class= ?,declared_value= ?,
            time_window_start= ?,time_window_end= ?,time_window2_start= ?,
            time_window2_end= ?,visit_duration_minutes= ?,delivery_notes= ?,
            access_instructions= ?,vehicle_id= ?,driver_id= ?,depot_id= ?,
            updated_at=datetime('now')
            WHERE id=?
        """, (
            data["reference"], data["client_id"], data["operation_type"], data["status"],
            data["scheduled_date"], data["priority"], data["quantity_kg"], data["volume_m3"],
            data["units_count"], data["goods_category"], data["temperature_required"],
            data["adr_class"], data["declared_value"], data["time_window_start"],
            data["time_window_end"], data["time_window2_start"], data["time_window2_end"],
            data["visit_duration_minutes"], data["delivery_notes"], data["access_instructions"],
            data["vehicle_id"], data["driver_id"], data["depot_id"], oid,
        ))
        conn.commit(); conn.close()
        log_action("ORDER_UPDATE", f"Commande #{oid} modifiée")
        show_toast(self.window(), "Commande mise à jour", "success")
        self.refresh_data()

    def _delete_order(self, oid: int):
        if not ConfirmDialog.ask(self, "Supprimer", "Supprimer cette commande ", "danger"): return
        conn = get_connection()
        conn.execute("UPDATE orders SET archived=1 WHERE id= ?", (oid,))
        conn.commit(); conn.close()
        log_action("ORDER_DELETE", f"Commande #{oid} archivée")
        show_toast(self.window(), "Commande archivée", "info")
        self.refresh_data()

    def _duplicate(self, oid: int):
        conn = get_connection()
        row = conn.execute("SELECT * FROM orders WHERE id= ?", (oid,)).fetchone()
        if not row: conn.close(); return
        d = dict(row)
        ref = _next_ref(conn)
        conn.execute("""
            INSERT INTO orders
            (reference,client_id,operation_type,quantity_kg,volume_m3,
             time_window_start,time_window_end,visit_duration_minutes,
             priority,goods_category,status)
            VALUES (?,?,?,?,?,?,?,?,?,?,'pending')
        """, (ref, d["client_id"], d.get("operation_type","delivery"),
              d.get("quantity_kg",0), d.get("volume_m3",0),
              d.get("time_window_start"), d.get("time_window_end"),
              d.get("visit_duration_minutes",15), d.get("priority",3),
              d.get("goods_category","standard")))
        conn.commit(); conn.close()
        log_action("ORDER_DUPLICATE", f"Commande #{oid} dupliquée → {ref}")
        show_toast(self.window(), f"Dupliquée → {ref}", "success")
        self.refresh_data()

    # ── Batch actions ─────────────────────────────────────────────────

    def _selected_ids(self) -> list:
        ids = []
        for sr in self._table.selectionModel().selectedRows():
            item = self._table.item(sr.row(), 0)
            if item:
                oid = item.data(Qt.ItemDataRole.UserRole)
                if oid: ids.append(oid)
        return ids

    def _batch_deliver(self):
        ids = self._selected_ids()
        if not ids: return
        conn = get_connection()
        conn.execute(
            f"UPDATE orders SET status='delivered' WHERE id IN ({','.join('?'*len(ids))})", ids
        ); conn.commit(); conn.close()
        log_action("ORDER_BATCH_DELIVER", f"{len(ids)} commandes marquées livrées")
        show_toast(self.window(), f"{len(ids)} commandes livrées", "success")
        self.refresh_data()

    def _batch_reassign(self):
        ids = self._selected_ids()
        if not ids: return
        dlg = _BatchReassignDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        vid, did = dlg.get_assignment()
        conn = get_connection()
        if vid:
            conn.execute(f"UPDATE orders SET vehicle_id= ?,status='assigned' WHERE id IN ({','.join('?'*len(ids))})", [vid]+ids)
        if did:
            conn.execute(f"UPDATE orders SET driver_id= ? WHERE id IN ({','.join('?'*len(ids))})", [did]+ids)
        conn.commit(); conn.close()
        log_action("ORDER_BATCH_ASSIGN", f"{len(ids)} commandes réassignées")
        show_toast(self.window(), f"{len(ids)} commandes réassignées", "success")
        self.refresh_data()

    def _batch_delete(self):
        ids = self._selected_ids()
        if not ids: return
        if not ConfirmDialog.ask(self, "Supprimer", f"Archiver {len(ids)} commandes ", "danger"): return
        conn = get_connection()
        conn.execute(f"UPDATE orders SET archived=1 WHERE id IN ({','.join('?'*len(ids))})", ids)
        conn.commit(); conn.close()
        log_action("ORDER_BATCH_DELETE", f"{len(ids)} commandes archivées")
        show_toast(self.window(), f"{len(ids)} commandes archivées", "info")
        self.refresh_data()

    # ── Recurring ─────────────────────────────────────────────────────

    def _show_recurring(self):
        _RecurringDialog(self, gen_week_cb=self._gen_week).exec()

    def _gen_week(self):
        conn = get_connection()
        n_tpl = conn.execute(
            "SELECT COUNT(*) FROM recurring_order_templates WHERE is_active=1"
        ).fetchone()[0]
        conn.close()
        if n_tpl == 0:
            show_toast(self.window(), "Aucun template actif trouvé.", "info"); return
        t = _GenerateWeekThread(parent=self)
        t.finished.connect(lambda n: (
            show_toast(self.window(), f"{n} commandes générées pour la semaine", "success"),
            self.refresh_data(),
        ))
        t.error.connect(lambda e: show_toast(self.window(), f"Erreur: {e}", "error"))
        self._threads.append(t)
        t.start()

    # ── Import / Export ───────────────────────────────────────────────

    def _export(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter commandes", "orders_export.csv",
            "CSV (*.csv);;Excel (*.xlsx)"
        )
        if not path: return
        conn = get_connection()
        rows = conn.execute("""
            SELECT o.*,c.name AS client_name FROM orders o
            LEFT JOIN clients c ON c.id=o.client_id
            WHERE o.archived=0 ORDER BY o.scheduled_date DESC
        """).fetchall()
        conn.close()
        try:
            if path.endswith(".xlsx") and HAS_OPENPYXL:
                wb = _openpyxl.Workbook(); ws = wb.active
                if rows:
                    ws.append(list(rows[0].keys()))
                    for row in rows: ws.append(list(row))
                wb.save(path)
            else:
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    if rows:
                        wr = csv.DictWriter(f, fieldnames=rows[0].keys())
                        wr.writeheader()
                        wr.writerows([dict(r) for r in rows])
            show_toast(self.window(), f"{len(rows)} commandes exportées", "success")
        except Exception as e:
            show_toast(self.window(), f"Erreur export: {e}", "error")

    def _import(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Importer commandes", "",
            "CSV / Excel (*.csv *.xlsx *.xls)"
        )
        if not path: return
        dlg = _ImportMappingDialog(path, self)
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        col_map = dlg.get_mapping()
        t = _ImportThread(path, col_map, parent=self)
        t.progress.connect(lambda m: show_toast(self.window(), m, "info"))
        t.finished.connect(self._on_import_done)
        self._threads.append(t); t.start()

    def _on_import_done(self, report: dict):
        c, u, e = report.get("created",0), report.get("updated",0), report.get("errors",0)
        QMessageBox.information(
            self, "Import terminé",
            f"Créées : {c}\nMises à jour : {u}\nErreurs : {e}"
        )
        log_action("ORDER_IMPORT", f"{c} créées, {u} màj, {e} erreurs")
        self.refresh_data()


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT MAPPING DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class _ImportMappingDialog(QDialog):
    _FIELDS = [
        ("reference",     "Référence"),
        ("client",        "Client (nom)"),
        ("quantity_kg",   "Quantité kg"),
        ("scheduled_date","Date prévue"),
        ("notes",         "Notes"),
    ]

    def __init__(self, path: str, parent=None):
        super().__init__(parent)
        self.path = path
        self.setWindowTitle("Mapping colonnes")
        self.resize(500, 360)
        self.setStyleSheet(
            _dialog_qss()
            + f"QDialog{{background:{C['bg']};color:{C['text']};}}" + _INP_STYLE +
            f"QLabel{{background:transparent;color:{C['text']};}}"
        )
        self._headers = self._get_headers()
        self._combos: dict[str, QComboBox] = {}
        self._setup_ui()

    def _get_headers(self) -> list:
        try:
            if self.path.endswith((".xlsx",".xls")) and HAS_OPENPYXL:
                wb = _openpyxl.load_workbook(self.path, read_only=True, data_only=True)
                ws = wb.active
                h = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                wb.close(); return h
            with open(self.path, newline="", encoding="utf-8-sig") as f:
                return next(csv.reader(f))
        except Exception:
            return []

    def _setup_ui(self):
        lo = QVBoxLayout(self); fl = QFormLayout()
        fl.setSpacing(8); fl.setContentsMargins(16,16,16,8)
        t = QLabel(f"Fichier : {self.path.split('/')[-1].split(chr(92))[-1]}")
        t.setStyleSheet(f"color:{C['text2']};font-size:11px;background:transparent;")
        lo.addWidget(t)
        for fk, flabel in self._FIELDS:
            cb = QComboBox(); cb.addItem("— ignorer", ""); 
            for h in self._headers: cb.addItem(h, h)
            for i in range(cb.count()):
                if cb.itemText(i).lower() == fk.lower() or cb.itemData(i).lower() == fk.lower():
                    cb.setCurrentIndex(i); break
            self._combos[fk] = cb
            fl.addRow(QLabel(flabel), cb)
        lo.addLayout(fl); lo.addStretch()
        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("Annuler"); cancel.setObjectName("secondaryBtn")
        cancel.setFixedHeight(32); cancel.clicked.connect(self.reject)
        ok = QPushButton("Importer"); ok.setObjectName("primaryBtn")
        ok.setFixedHeight(32); ok.clicked.connect(self.accept)
        btn_row.addWidget(cancel); btn_row.addWidget(ok)
        lo.addLayout(btn_row)

    def get_mapping(self) -> dict:
        return {fk: cb.currentData() for fk, cb in self._combos.items() if cb.currentData()}


# ═══════════════════════════════════════════════════════════════════════════════
# BATCH REASSIGN DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class _BatchReassignDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Réassigner")
        self.resize(380, 200)
        self.setStyleSheet(
            _dialog_qss()
            + f"QDialog{{background:{C['bg']};color:{C['text']};}}" + _INP_STYLE +
            f"QLabel{{background:transparent;color:{C['text']};}}"
        )
        conn = get_connection()
        vehicles = conn.execute(
            "SELECT id,registration FROM vehicles ORDER BY registration"
        ).fetchall()
        drivers = conn.execute(
            "SELECT id,first_name,last_name FROM drivers ORDER BY last_name"
        ).fetchall()
        conn.close()

        lo = QVBoxLayout(self); fl = QFormLayout(); fl.setSpacing(8); fl.setContentsMargins(16,12,16,8)
        self._veh = QComboBox(); self._veh.addItem("— Garder", None)
        for v in vehicles: self._veh.addItem(v["registration"], v["id"])
        self._drv = QComboBox(); self._drv.addItem("— Garder", None)
        for d in drivers: self._drv.addItem(f"{d['first_name']} {d['last_name']}", d["id"])
        fl.addRow(QLabel("Véhicule"), self._veh); fl.addRow(QLabel("Chauffeur"), self._drv)
        lo.addLayout(fl); lo.addStretch()
        btn_row = QHBoxLayout(); btn_row.addStretch()
        cancel = QPushButton("Annuler"); cancel.setObjectName("secondaryBtn"); cancel.setFixedHeight(32); cancel.clicked.connect(self.reject)
        ok = QPushButton("Appliquer"); ok.setObjectName("primaryBtn"); ok.setFixedHeight(32); ok.clicked.connect(self.accept)
        btn_row.addWidget(cancel); btn_row.addWidget(ok); lo.addLayout(btn_row)

    def get_assignment(self) -> tuple:
        return self._veh.currentData(), self._drv.currentData()
