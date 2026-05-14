"""
report_service.py — ReportService : PDF / XLSX / JSON
=====================================================
Dépendances : reportlab, openpyxl, matplotlib (optionnel graphiques), qrcode, Pillow.
"""

from __future__ import annotations

import io
import json
import logging
import os
import sqlite3
from datetime import datetime, timedelta, date
from typing import Sequence

logger = logging.getLogger(__name__)

# ── reportlab ────────────────────────────────────────────────────────────────
try:
  from reportlab.lib import colors
  from reportlab.lib.pagesizes import A4
  from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
  from reportlab.lib.units import cm
  from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
  from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, PageBreak, Image as RLImage,
  )
  REPORTLAB_OK = True
except ImportError:
  REPORTLAB_OK = False
  colors = None # type: ignore

# ── openpyxl ─────────────────────────────────────────────────────────────────
try:
  import openpyxl

  OPENPYXL_OK = True
except ImportError:
  OPENPYXL_OK = False

# ── matplotlib ─────────────────────────────────────────────────────────────
try:
  import matplotlib

  matplotlib.use("Agg")
  from matplotlib.figure import Figure
  from matplotlib.backends.backend_agg import FigureCanvasAgg

  HAS_MPL = True
except ImportError:
  HAS_MPL = False

# ── QR ───────────────────────────────────────────────────────────────────────
try:
  import qrcode

  HAS_QR = True
except ImportError:
  HAS_QR = False

from ..database.db_manager import get_connection, log_action

# Couleurs CityPulse (PDF clair pour impression)
CLR_PRIMARY = None
CLR_SECONDARY = None
if REPORTLAB_OK:
  CLR_PRIMARY = colors.HexColor("#1E3A5F")
  CLR_SECONDARY = colors.HexColor("#3B82F6")
  CLR_SUCCESS = colors.HexColor("#22C55E")
  CLR_WARNING = colors.HexColor("#F59E0B")
  CLR_DANGER = colors.HexColor("#EF4444")
  CLR_LIGHT = colors.HexColor("#F9FAFB")
  CLR_BORDER = colors.HexColor("#E5E7EB")
  CLR_TEXT = colors.HexColor("#111827")
  CLR_MUTED = colors.HexColor("#6B7280")


def _require_reportlab():
  if not REPORTLAB_OK:
    raise RuntimeError("reportlab non installé — pip install reportlab")


def _require_openpyxl():
  if not OPENPYXL_OK:
    raise RuntimeError("openpyxl non installé — pip install openpyxl")


def _row_dict(row: sqlite3.Row) -> dict:
  return {k: row[k] for k in row.keys()}


def _file_size_kb(path: str) -> int:
  try:
    return max(1, int(os.path.getsize(path) / 1024))
  except OSError:
    return 0


# ── Internationalisation des rapports ─────────────────────────────────────────
_RL_DATA: dict[str, dict[str, str]] = {
  "fr": {
    "roadbook_title": "Carnet de route chauffeur",
    "fleet_daily_title": "Rapport flotte journalier",
    "kpi_title": "Rapport KPI",
    "algo_cmp_title": "Comparaison des algorithmes",
    "client_report_title": "Fiche client",
    "drv_perf_title": "Performance chauffeurs",
    "rse_title": "Rapport conformité RSE (cadre conducteurs)",
    "carrier_title": "Rapport transporteurs",
    "bl_title": "BON DE LIVRAISON",
    "cmr_title": "LETTRE DE VOITURE (CMR)",
    "adr_title": "FICHE TRANSPORT ADR (synthèse)",
    "manifest_title": "MANIFESTE DE CHARGEMENT",
    "route_sheet_title": "Feuille de Route",
    "legal_terms_title": "Conditions générales d'utilisation (synthèse)",
    "legal_privacy_title": "Politique de traitement des données (synthèse)",
    "legal_default_title": "Document légal",
    "label_driver": "Chauffeur", "label_phone": "Tél.", "label_license": "Permis",
    "label_vehicle": "Véhicule", "label_capacity": "Capacité",
    "label_planned_date": "Date planifiée", "label_generated": "Généré",
    "label_generated_at": "Généré le", "label_doc_no": "N° document",
    "label_order": "Commande", "label_shipper": "Expéditeur (dépôt)",
    "label_recipient": "Destinataire", "label_date": "Date",
    "label_name": "Nom", "label_signature": "Signature",
    "label_reserves": "Réserve / observations", "label_totals": "Totaux",
    "label_fill_rate": "Taux remplissage (kg)", "label_weight": "Poids",
    "label_volume": "Volume", "label_packages": "Colis",
    "label_algo": "Algo", "label_type": "Type", "label_case": "Case",
    "label_signatures": "Signatures", "label_shipper_short": "Expéditeur",
    "label_carrier_short": "Transporteur", "label_recipient_colon": "Destinataire",
    "section_planned_stops": "Arrêts planifiés",
    "section_recent_orders": "Commandes récentes (30)",
    "section_stop_detail": "Détail des arrêts",
    "section_dock_check": "Chef de quai — contrôle chargement",
    "col_order": "Ordre", "col_type": "Type", "col_client": "Client",
    "col_address": "Adresse", "col_order_ref": "Réf. commande",
    "col_planned_arrival": "Arrivée plan.", "col_planned_departure": "Départ plan.",
    "col_duration_min": "Durée (min)", "col_dist_prev_km": "Dist. préc. km",
    "col_load_kg": "Charge kg", "col_status": "Statut",
    "col_reg": "Immat.", "col_driver": "Chauffeur", "col_km": "Km",
    "col_duration_min_short": "Durée min", "col_co2_kg": "CO₂ kg",
    "col_stops": "Arrêts", "col_on_time": "À l'heure",
    "col_metric": "Métrique", "col_period": "Période", "col_prev_period": "S-1 (préc.)",
    "row_optim_runs": "Runs optimisation", "row_avg_dist": "Distance moy. (km)",
    "row_avg_respect": "Respect créneaux moy. %",
    "chart_optim_per_day": "Nombre d'optimisations / jour",
    "col_algo": "Algo", "col_dist_km": "Dist. km", "col_duration": "Durée",
    "col_cost": "Coût", "col_respect_pct": "Respect %", "col_avg_delay": "Retard moy.",
    "col_cpu_ms": "CPU ms", "col_clients": "Clients",
    "col_field": "Champ", "col_value": "Valeur", "col_id": "ID",
    "col_name": "Nom", "col_coords": "Coordonnées", "col_demand_kg": "Demande kg",
    "col_priority": "Priorité", "col_contact": "Contact", "col_phone": "Téléphone",
    "col_ref": "Réf.", "col_date": "Date", "col_kg": "kg",
    "col_n_routes": "Tournées", "col_max_drive": "Max conduite avant pause (min)",
    "col_pause_min": "Pause min", "col_max_h_day": "Max h/j",
    "col_rest_min_day": "Repos min/j", "col_route_duration_min": "Durée routes min",
    "col_alert": "Alerte", "alert_near_limit": " Durée proche plafond",
    "alert_check_breaks": "ℹ Vérifier pauses",
    "col_zones": "Zones", "col_cost_per_km": "€/km", "col_rating": "Note",
    "col_shipments": "Expéditions", "col_delivered": "Livrées", "col_total_cost": "Coût total",
    "col_article_desc": "Article / désignation", "col_qty_ordered": "Qté cmd (kg)",
    "col_qty_delivered": "Qté livrée", "col_observations": "Observations",
    "col_num": "#", "col_reference": "Référence", "col_qty": "Qté",
    "col_weight_kg": "Poids kg", "col_vol_m3": "Vol. m³",
    "col_instructions": "Instructions / accès",
    "col_un_designation": "Désignation officielle ONU", "col_un_no": "N° ONU",
    "col_adr_class": "Classe ADR (colonne commande)", "col_packaging_group": "Groupe d'emballage",
    "col_qty_packages": "Quantité / colis", "col_temperature": "Température",
    "col_time_window": "Fenêtre horaire", "col_arrival": "Arrivée",
    "col_departure": "Départ", "col_load_kg_paren": "Charge (kg)",
    "row_total_dist": "Distance totale", "row_est_duration": "Durée estimée",
    "row_total_load": "Charge totale", "row_est_cost": "Coût estimé",
    "row_time_respect": "Respect horaires", "row_dist_source": "Source distance",
    "val_osrm_source": "Routière réelle (OSRM)", "val_haversine_source": "Estimation (Haversine)",
    "chart_km_by_vehicle": "Kilométrage planifié par véhicule",
    "msg_no_route": "Aucune tournée", "msg_no_order": "Aucune commande",
    "msg_no_carrier": "Aucun transporteur", "msg_no_stop": "Aucun arrêt",
    "cmr_case": "Case", "cmr_c1_title": "Expéditeur (nom, adresse, pays)",
    "cmr_c2_title": "Destinataire (nom, adresse, pays)",
    "cmr_c3_title": "Lieu prévu pour la livraison de la marchandise",
    "cmr_c6_title": "Transporteur (nom, adresse, pays) — 1er transporteur",
    "cmr_c11_title": "Nature de la marchandise", "cmr_c12_title": "Méthode d'emballage",
    "cmr_c13_title": "Nombre, désignation, poids brut kg",
    "cmr_c18_title": "Réserves et observations du transporteur",
    "cmr_c23_title": "Signature et cachet de l'expéditeur",
    "cmr_c24_title": "Signature et cachet du destinataire",
    "rse_subtitle": "Seuils déclarés vs charge planifiée agrégée (durée routes). Analyse indicative — validation métier requise.",
    "legal_body": ("Le présent document est un modèle informatif généré par CityPulse Logistics. "
      "Il ne constitue pas un conseil juridique. Adaptez-le avec votre service juridique "
      "avant toute diffusion. Données traitées : planification, géolocalisation, "
      "identifiants métier. Durée de conservation selon politique interne et obligations légales."),
    "manifest_certif": "Je certifie que le chargement correspond au présent manifeste.",
    "bl_sig_text": "Signature du destinataire (précédée de la mention « Bon pour accord »)",
    "adr_declaration": ("Le transporteur déclare que la marchandise est acceptée pour transport conformément aux prescriptions "
      "de l'ADR. L'expéditeur certifie que la classification, l'emballage, l'étiquetage et la documentation "
      "sont conformes aux réglementations en vigueur. <i>Texte type — à adapter selon votre exploitation.</i>"),
    "footer_confidential": "Document confidentiel — CityPulse Logistics — Vérifier les créneaux horaires avant départ. En cas d'incident, contacter le dispatch.",
    "footer_generated": "Document généré le", "footer_keep": "Conserver pour traçabilité.",
    "footer_algo": "Algorithme", "footer_confidential_route_2": "Ce document est confidentiel.",
    "footer_generated_by": "Document généré par CityPulse Logistics",
  },
  "en": {
    "roadbook_title": "Driver Roadbook", "fleet_daily_title": "Daily Fleet Report",
    "kpi_title": "KPI Report", "algo_cmp_title": "Algorithm Comparison",
    "client_report_title": "Client Profile", "drv_perf_title": "Driver Performance",
    "rse_title": "RSE Compliance Report (drivers)", "carrier_title": "Carrier Report",
    "bl_title": "DELIVERY NOTE", "cmr_title": "CONSIGNMENT NOTE (CMR)",
    "adr_title": "ADR TRANSPORT DOCUMENT (summary)", "manifest_title": "LOADING MANIFEST",
    "route_sheet_title": "Route Sheet",
    "legal_terms_title": "Terms of Use (summary)",
    "legal_privacy_title": "Data Processing Policy (summary)", "legal_default_title": "Legal Document",
    "label_driver": "Driver", "label_phone": "Tel.", "label_license": "License",
    "label_vehicle": "Vehicle", "label_capacity": "Capacity",
    "label_planned_date": "Planned date", "label_generated": "Generated",
    "label_generated_at": "Generated on", "label_doc_no": "Document N°",
    "label_order": "Order", "label_shipper": "Shipper (depot)",
    "label_recipient": "Recipient", "label_date": "Date",
    "label_name": "Name", "label_signature": "Signature",
    "label_reserves": "Reserves / remarks", "label_totals": "Totals",
    "label_fill_rate": "Fill rate (kg)", "label_weight": "Weight",
    "label_volume": "Volume", "label_packages": "Packages",
    "label_algo": "Algo", "label_type": "Type", "label_case": "Box",
    "label_signatures": "Signatures", "label_shipper_short": "Shipper",
    "label_carrier_short": "Carrier", "label_recipient_colon": "Recipient",
    "section_planned_stops": "Planned stops",
    "section_recent_orders": "Recent orders (30)", "section_stop_detail": "Stop detail",
    "section_dock_check": "Dock supervisor — loading check",
    "col_order": "Order", "col_type": "Type", "col_client": "Client",
    "col_address": "Address", "col_order_ref": "Order ref.",
    "col_planned_arrival": "Planned arr.", "col_planned_departure": "Planned dep.",
    "col_duration_min": "Duration (min)", "col_dist_prev_km": "Prev. dist. km",
    "col_load_kg": "Load kg", "col_status": "Status",
    "col_reg": "Plate", "col_driver": "Driver", "col_km": "Km",
    "col_duration_min_short": "Duration min", "col_co2_kg": "CO₂ kg",
    "col_stops": "Stops", "col_on_time": "On time",
    "col_metric": "Metric", "col_period": "Period", "col_prev_period": "W-1 (prev.)",
    "row_optim_runs": "Optimization runs", "row_avg_dist": "Avg distance (km)",
    "row_avg_respect": "Avg time slot respect %",
    "chart_optim_per_day": "Number of optimizations / day",
    "col_algo": "Algo", "col_dist_km": "Dist. km", "col_duration": "Duration",
    "col_cost": "Cost", "col_respect_pct": "Respect %", "col_avg_delay": "Avg delay",
    "col_cpu_ms": "CPU ms", "col_clients": "Clients",
    "col_field": "Field", "col_value": "Value", "col_id": "ID",
    "col_name": "Name", "col_coords": "Coordinates", "col_demand_kg": "Demand kg",
    "col_priority": "Priority", "col_contact": "Contact", "col_phone": "Phone",
    "col_ref": "Ref.", "col_date": "Date", "col_kg": "kg",
    "col_n_routes": "Routes", "col_max_drive": "Max drive before break (min)",
    "col_pause_min": "Min break", "col_max_h_day": "Max h/day",
    "col_rest_min_day": "Min rest/day", "col_route_duration_min": "Route duration min",
    "col_alert": "Alert", "alert_near_limit": " Near daily limit",
    "alert_check_breaks": "ℹ Check breaks",
    "col_zones": "Zones", "col_cost_per_km": "€/km", "col_rating": "Rating",
    "col_shipments": "Shipments", "col_delivered": "Delivered", "col_total_cost": "Total cost",
    "col_article_desc": "Article / description", "col_qty_ordered": "Qty ord. (kg)",
    "col_qty_delivered": "Qty delivered", "col_observations": "Remarks",
    "col_num": "#", "col_reference": "Reference", "col_qty": "Qty",
    "col_weight_kg": "Weight kg", "col_vol_m3": "Vol. m³",
    "col_instructions": "Instructions / access",
    "col_un_designation": "Official UN designation", "col_un_no": "UN N°",
    "col_adr_class": "ADR class (order column)", "col_packaging_group": "Packing group",
    "col_qty_packages": "Quantity / packages", "col_temperature": "Temperature",
    "col_time_window": "Time window", "col_arrival": "Arrival",
    "col_departure": "Departure", "col_load_kg_paren": "Load (kg)",
    "row_total_dist": "Total distance", "row_est_duration": "Estimated duration",
    "row_total_load": "Total load", "row_est_cost": "Estimated cost",
    "row_time_respect": "Time slot respect", "row_dist_source": "Distance source",
    "val_osrm_source": "Real road (OSRM)", "val_haversine_source": "Estimate (Haversine)",
    "chart_km_by_vehicle": "Planned km by vehicle",
    "msg_no_route": "No routes", "msg_no_order": "No orders",
    "msg_no_carrier": "No carriers", "msg_no_stop": "No stops",
    "cmr_case": "Box", "cmr_c1_title": "Sender (name, address, country)",
    "cmr_c2_title": "Consignee (name, address, country)",
    "cmr_c3_title": "Place designated for delivery of the goods",
    "cmr_c6_title": "Carrier (name, address, country) — 1st carrier",
    "cmr_c11_title": "Nature of goods", "cmr_c12_title": "Packing method",
    "cmr_c13_title": "Number, description, gross weight kg",
    "cmr_c18_title": "Carrier's reservations and observations",
    "cmr_c23_title": "Sender's signature and stamp",
    "cmr_c24_title": "Consignee's signature and stamp",
    "rse_subtitle": "Declared thresholds vs aggregated planned load (route duration). Indicative analysis — business validation required.",
    "legal_body": ("This document is an informational template generated by CityPulse Logistics. "
      "It does not constitute legal advice. Adapt it with your legal department "
      "before any distribution. Data processed: planning, geolocation, "
      "business identifiers. Retention period according to internal policy and legal obligations."),
    "manifest_certif": "I certify that the loading corresponds to this manifest.",
    "bl_sig_text": "Recipient's signature (preceded by the mention 'Approved')",
    "adr_declaration": ("The carrier declares that the goods are accepted for transport in accordance with ADR regulations. "
      "The sender certifies that classification, packaging, labelling and documentation "
      "comply with current regulations. <i>Standard text — adapt to your operations.</i>"),
    "footer_confidential": "Confidential — CityPulse Logistics — Check time windows before departure. In case of incident, contact dispatch.",
    "footer_generated": "Generated on", "footer_keep": "Keep for traceability.",
    "footer_algo": "Algorithm", "footer_confidential_route_2": "This document is confidential.",
    "footer_generated_by": "Document generated by CityPulse Logistics",
  },
  "es": {
    "roadbook_title": "Hoja de ruta del conductor", "fleet_daily_title": "Informe diario de flota",
    "kpi_title": "Informe KPI", "algo_cmp_title": "Comparación de algoritmos",
    "client_report_title": "Ficha del cliente", "drv_perf_title": "Rendimiento de conductores",
    "rse_title": "Informe de conformidad RSE (conductores)", "carrier_title": "Informe de transportistas",
    "bl_title": "ALBARÁN DE ENTREGA", "cmr_title": "CARTA DE PORTE (CMR)",
    "adr_title": "DOCUMENTO DE TRANSPORTE ADR (resumen)", "manifest_title": "MANIFIESTO DE CARGA",
    "route_sheet_title": "Hoja de Ruta",
    "legal_terms_title": "Condiciones generales de uso (resumen)",
    "legal_privacy_title": "Política de protección de datos (resumen)", "legal_default_title": "Documento legal",
    "label_driver": "Conductor", "label_phone": "Tel.", "label_license": "Licencia",
    "label_vehicle": "Vehículo", "label_capacity": "Capacidad",
    "label_planned_date": "Fecha planificada", "label_generated": "Generado",
    "label_generated_at": "Generado el", "label_doc_no": "N° documento",
    "label_order": "Pedido", "label_shipper": "Remitente (depósito)",
    "label_recipient": "Destinatario", "label_date": "Fecha",
    "label_name": "Nombre", "label_signature": "Firma",
    "label_reserves": "Reservas / observaciones", "label_totals": "Totales",
    "label_fill_rate": "Tasa de llenado (kg)", "label_weight": "Peso",
    "label_volume": "Volumen", "label_packages": "Bultos",
    "label_algo": "Algo", "label_type": "Tipo", "label_case": "Casilla",
    "label_signatures": "Firmas", "label_shipper_short": "Remitente",
    "label_carrier_short": "Transportista", "label_recipient_colon": "Destinatario",
    "section_planned_stops": "Paradas planificadas",
    "section_recent_orders": "Pedidos recientes (30)", "section_stop_detail": "Detalle de paradas",
    "section_dock_check": "Jefe de muelle — control de carga",
    "col_order": "Orden", "col_type": "Tipo", "col_client": "Cliente",
    "col_address": "Dirección", "col_order_ref": "Ref. pedido",
    "col_planned_arrival": "Llegada plan.", "col_planned_departure": "Salida plan.",
    "col_duration_min": "Duración (min)", "col_dist_prev_km": "Dist. ant. km",
    "col_load_kg": "Carga kg", "col_status": "Estado",
    "col_reg": "Matrícula", "col_driver": "Conductor", "col_km": "Km",
    "col_duration_min_short": "Duración min", "col_co2_kg": "CO₂ kg",
    "col_stops": "Paradas", "col_on_time": "A tiempo",
    "col_metric": "Métrica", "col_period": "Período", "col_prev_period": "S-1 (ant.)",
    "row_optim_runs": "Optimizaciones", "row_avg_dist": "Distancia prom. (km)",
    "row_avg_respect": "Ventana horaria prom. %",
    "chart_optim_per_day": "Número de optimizaciones / día",
    "col_algo": "Algo", "col_dist_km": "Dist. km", "col_duration": "Duración",
    "col_cost": "Coste", "col_respect_pct": "Respeto %", "col_avg_delay": "Retraso prom.",
    "col_cpu_ms": "CPU ms", "col_clients": "Clientes",
    "col_field": "Campo", "col_value": "Valor", "col_id": "ID",
    "col_name": "Nombre", "col_coords": "Coordenadas", "col_demand_kg": "Demanda kg",
    "col_priority": "Prioridad", "col_contact": "Contacto", "col_phone": "Teléfono",
    "col_ref": "Ref.", "col_date": "Fecha", "col_kg": "kg",
    "col_n_routes": "Rutas", "col_max_drive": "Conducción máx. antes de pausa (min)",
    "col_pause_min": "Pausa mín.", "col_max_h_day": "Máx h/día",
    "col_rest_min_day": "Descanso mín/día", "col_route_duration_min": "Duración rutas min",
    "col_alert": "Alerta", "alert_near_limit": " Cerca del límite diario",
    "alert_check_breaks": "ℹ Verificar pausas",
    "col_zones": "Zonas", "col_cost_per_km": "€/km", "col_rating": "Nota",
    "col_shipments": "Expediciones", "col_delivered": "Entregadas", "col_total_cost": "Coste total",
    "col_article_desc": "Artículo / descripción", "col_qty_ordered": "Cant. ped. (kg)",
    "col_qty_delivered": "Cant. entregada", "col_observations": "Observaciones",
    "col_num": "#", "col_reference": "Referencia", "col_qty": "Cant.",
    "col_weight_kg": "Peso kg", "col_vol_m3": "Vol. m³",
    "col_instructions": "Instrucciones / acceso",
    "col_un_designation": "Denominación oficial ONU", "col_un_no": "N° ONU",
    "col_adr_class": "Clase ADR (columna pedido)", "col_packaging_group": "Grupo de embalaje",
    "col_qty_packages": "Cantidad / bultos", "col_temperature": "Temperatura",
    "col_time_window": "Ventana horaria", "col_arrival": "Llegada",
    "col_departure": "Salida", "col_load_kg_paren": "Carga (kg)",
    "row_total_dist": "Distancia total", "row_est_duration": "Duración estimada",
    "row_total_load": "Carga total", "row_est_cost": "Coste estimado",
    "row_time_respect": "Respeto horario", "row_dist_source": "Fuente distancia",
    "val_osrm_source": "Ruta real (OSRM)", "val_haversine_source": "Estimación (Haversine)",
    "chart_km_by_vehicle": "Km planificados por vehículo",
    "msg_no_route": "Sin rutas", "msg_no_order": "Sin pedidos",
    "msg_no_carrier": "Sin transportistas", "msg_no_stop": "Sin paradas",
    "cmr_case": "Casilla", "cmr_c1_title": "Remitente (nombre, dirección, país)",
    "cmr_c2_title": "Destinatario (nombre, dirección, país)",
    "cmr_c3_title": "Lugar previsto para la entrega de la mercancía",
    "cmr_c6_title": "Transportista (nombre, dirección, país) — 1er transportista",
    "cmr_c11_title": "Naturaleza de la mercancía", "cmr_c12_title": "Método de embalaje",
    "cmr_c13_title": "Número, descripción, peso bruto kg",
    "cmr_c18_title": "Reservas y observaciones del transportista",
    "cmr_c23_title": "Firma y sello del remitente",
    "cmr_c24_title": "Firma y sello del destinatario",
    "rse_subtitle": "Umbrales declarados vs carga planificada agregada (duración rutas). Análisis indicativo — validación empresarial requerida.",
    "legal_body": ("Este documento es una plantilla informativa generada por CityPulse Logistics. "
      "No constituye asesoramiento jurídico. Adáptelo con su departamento legal "
      "antes de cualquier distribución. Datos tratados: planificación, geolocalización, "
      "identificadores empresariales. Período de conservación según política interna y obligaciones legales."),
    "manifest_certif": "Certifico que la carga corresponde al presente manifiesto.",
    "bl_sig_text": "Firma del destinatario (precedida de la mención «Conforme»)",
    "adr_declaration": ("El transportista declara que la mercancía es aceptada para transporte de conformidad con las prescripciones ADR. "
      "El remitente certifica que la clasificación, el embalaje, el etiquetado y la documentación "
      "son conformes a las reglamentaciones vigentes. <i>Texto tipo — a adaptar según su explotación.</i>"),
    "footer_confidential": "Confidencial — CityPulse Logistics — Verificar ventanas horarias antes de salir. En caso de incidente, contactar despacho.",
    "footer_generated": "Generado el", "footer_keep": "Conservar para trazabilidad.",
    "footer_algo": "Algoritmo", "footer_confidential_route_2": "Este documento es confidencial.",
    "footer_generated_by": "Documento generado por CityPulse Logistics",
  },
  "de": {
    "roadbook_title": "Fahrerrouten-Heft", "fleet_daily_title": "Täglicher Flottenbericht",
    "kpi_title": "KPI-Bericht", "algo_cmp_title": "Algorithmenvergleich",
    "client_report_title": "Kundenprofil", "drv_perf_title": "Fahrerleistung",
    "rse_title": "CSR-Konformitätsbericht (Fahrer)", "carrier_title": "Spediteursbericht",
    "bl_title": "LIEFERSCHEIN", "cmr_title": "FRACHTBRIEF (CMR)",
    "adr_title": "ADR-TRANSPORTDOKUMENT (Zusammenfassung)", "manifest_title": "LADELISTE",
    "route_sheet_title": "Tourenblatt",
    "legal_terms_title": "Allgemeine Nutzungsbedingungen (Zusammenfassung)",
    "legal_privacy_title": "Datenschutzrichtlinie (Zusammenfassung)", "legal_default_title": "Rechtsdokument",
    "label_driver": "Fahrer", "label_phone": "Tel.", "label_license": "Führerschein",
    "label_vehicle": "Fahrzeug", "label_capacity": "Kapazität",
    "label_planned_date": "Geplantes Datum", "label_generated": "Erstellt",
    "label_generated_at": "Erstellt am", "label_doc_no": "Dokument-Nr.",
    "label_order": "Auftrag", "label_shipper": "Absender (Depot)",
    "label_recipient": "Empfänger", "label_date": "Datum",
    "label_name": "Name", "label_signature": "Unterschrift",
    "label_reserves": "Vorbehalte / Bemerkungen", "label_totals": "Gesamt",
    "label_fill_rate": "Auslastungsgrad (kg)", "label_weight": "Gewicht",
    "label_volume": "Volumen", "label_packages": "Pakete",
    "label_algo": "Algo", "label_type": "Typ", "label_case": "Feld",
    "label_signatures": "Unterschriften", "label_shipper_short": "Absender",
    "label_carrier_short": "Spediteur", "label_recipient_colon": "Empfänger",
    "section_planned_stops": "Geplante Haltestellen",
    "section_recent_orders": "Letzte Aufträge (30)", "section_stop_detail": "Stopdetails",
    "section_dock_check": "Ladeaufsicht — Ladekontrolle",
    "col_order": "Reihenfolge", "col_type": "Typ", "col_client": "Kunde",
    "col_address": "Adresse", "col_order_ref": "Auftragsref.",
    "col_planned_arrival": "Geplante Ank.", "col_planned_departure": "Geplante Abf.",
    "col_duration_min": "Dauer (min)", "col_dist_prev_km": "Vorh. Dist. km",
    "col_load_kg": "Ladung kg", "col_status": "Status",
    "col_reg": "Kfz-Kennz.", "col_driver": "Fahrer", "col_km": "Km",
    "col_duration_min_short": "Dauer min", "col_co2_kg": "CO₂ kg",
    "col_stops": "Stopps", "col_on_time": "Pünktlich",
    "col_metric": "Kennzahl", "col_period": "Zeitraum", "col_prev_period": "W-1 (vor.)",
    "row_optim_runs": "Optimierungsläufe", "row_avg_dist": "Ø Distanz (km)",
    "row_avg_respect": "Ø Zeitfenster %",
    "chart_optim_per_day": "Anzahl Optimierungen / Tag",
    "col_algo": "Algo", "col_dist_km": "Dist. km", "col_duration": "Dauer",
    "col_cost": "Kosten", "col_respect_pct": "Einhaltung %", "col_avg_delay": "Ø Verspätung",
    "col_cpu_ms": "CPU ms", "col_clients": "Kunden",
    "col_field": "Feld", "col_value": "Wert", "col_id": "ID",
    "col_name": "Name", "col_coords": "Koordinaten", "col_demand_kg": "Bedarf kg",
    "col_priority": "Priorität", "col_contact": "Kontakt", "col_phone": "Telefon",
    "col_ref": "Ref.", "col_date": "Datum", "col_kg": "kg",
    "col_n_routes": "Touren", "col_max_drive": "Max. Fahrdauer vor Pause (min)",
    "col_pause_min": "Min. Pause", "col_max_h_day": "Max h/Tag",
    "col_rest_min_day": "Min. Ruhezeit/Tag", "col_route_duration_min": "Tourendauer min",
    "col_alert": "Alarm", "alert_near_limit": " Nahe Tageslimit",
    "alert_check_breaks": "ℹ Pausen prüfen",
    "col_zones": "Zonen", "col_cost_per_km": "€/km", "col_rating": "Bewertung",
    "col_shipments": "Sendungen", "col_delivered": "Geliefert", "col_total_cost": "Gesamtkosten",
    "col_article_desc": "Artikel / Bezeichnung", "col_qty_ordered": "Best. Menge (kg)",
    "col_qty_delivered": "Gelieferte Menge", "col_observations": "Bemerkungen",
    "col_num": "#", "col_reference": "Referenz", "col_qty": "Menge",
    "col_weight_kg": "Gewicht kg", "col_vol_m3": "Vol. m³",
    "col_instructions": "Hinweise / Zugang",
    "col_un_designation": "Offizielle UN-Bezeichnung", "col_un_no": "UN-Nr.",
    "col_adr_class": "ADR-Klasse (Auftragsspalte)", "col_packaging_group": "Verpackungsgruppe",
    "col_qty_packages": "Menge / Pakete", "col_temperature": "Temperatur",
    "col_time_window": "Zeitfenster", "col_arrival": "Ankunft",
    "col_departure": "Abfahrt", "col_load_kg_paren": "Ladung (kg)",
    "row_total_dist": "Gesamtdistanz", "row_est_duration": "Geschätzte Dauer",
    "row_total_load": "Gesamtladung", "row_est_cost": "Geschätzte Kosten",
    "row_time_respect": "Zeitfenstereinhaltung", "row_dist_source": "Distanzquelle",
    "val_osrm_source": "Echte Straße (OSRM)", "val_haversine_source": "Schätzung (Haversine)",
    "chart_km_by_vehicle": "Geplante km pro Fahrzeug",
    "msg_no_route": "Keine Touren", "msg_no_order": "Keine Aufträge",
    "msg_no_carrier": "Keine Spediteure", "msg_no_stop": "Keine Stopps",
    "cmr_case": "Feld", "cmr_c1_title": "Absender (Name, Adresse, Land)",
    "cmr_c2_title": "Empfänger (Name, Adresse, Land)",
    "cmr_c3_title": "Vorgesehener Lieferort",
    "cmr_c6_title": "Frachtführer (Name, Adresse, Land) — 1. Frachtführer",
    "cmr_c11_title": "Art der Güter", "cmr_c12_title": "Verpackungsart",
    "cmr_c13_title": "Anzahl, Beschreibung, Bruttogewicht kg",
    "cmr_c18_title": "Vorbehalte und Bemerkungen des Frachtführers",
    "cmr_c23_title": "Unterschrift und Stempel des Absenders",
    "cmr_c24_title": "Unterschrift und Stempel des Empfängers",
    "rse_subtitle": "Erklärte Schwellenwerte vs. aggregierte geplante Last (Tourdauer). Indikativer Bericht — betriebliche Validierung erforderlich.",
    "legal_body": ("Dieses Dokument ist eine von CityPulse Logistics generierte Informationsvorlage. "
      "Es stellt keine Rechtsberatung dar. Passen Sie es mit Ihrer Rechtsabteilung an, "
      "bevor Sie es verteilen. Verarbeitete Daten: Planung, Geolokalisierung, "
      "Geschäftskennungen. Aufbewahrungsfrist gemäß interner Richtlinie und gesetzlichen Pflichten."),
    "manifest_certif": "Ich bestätige, dass die Ladung diesem Manifest entspricht.",
    "bl_sig_text": "Unterschrift des Empfängers (mit dem Vermerk «Einverstanden»)",
    "adr_declaration": ("Der Frachtführer erklärt, dass die Güter gemäß ADR-Vorschriften zur Beförderung angenommen werden. "
      "Der Absender bestätigt, dass Einstufung, Verpackung, Kennzeichnung und Dokumentation "
      "den geltenden Vorschriften entsprechen. <i>Mustertext — an Ihren Betrieb anpassen.</i>"),
    "footer_confidential": "Vertraulich — CityPulse Logistics — Zeitfenster vor Abfahrt prüfen. Bei Vorfall Disposition kontaktieren.",
    "footer_generated": "Erstellt am", "footer_keep": "Für Rückverfolgbarkeit aufbewahren.",
    "footer_algo": "Algorithmus", "footer_confidential_route_2": "Dieses Dokument ist vertraulich.",
    "footer_generated_by": "Dokument erstellt von CityPulse Logistics",
  },
  "ar": {
    "roadbook_title": "دفتر طريق السائق", "fleet_daily_title": "تقرير الأسطول اليومي",
    "kpi_title": "تقرير مؤشرات الأداء", "algo_cmp_title": "مقارنة الخوارزميات",
    "client_report_title": "بطاقة العميل", "drv_perf_title": "أداء السائقين",
    "rse_title": "تقرير الامتثال الاجتماعي (السائقون)", "carrier_title": "تقرير الناقلين",
    "bl_title": "وصل التسليم", "cmr_title": "بوليصة الشحن (CMR)",
    "adr_title": "وثيقة نقل المواد الخطرة (ملخص)", "manifest_title": "بيان التحميل",
    "route_sheet_title": "ورقة المسار",
    "legal_terms_title": "شروط الاستخدام (ملخص)",
    "legal_privacy_title": "سياسة حماية البيانات (ملخص)", "legal_default_title": "وثيقة قانونية",
    "label_driver": "السائق", "label_phone": "هاتف", "label_license": "رخصة القيادة",
    "label_vehicle": "المركبة", "label_capacity": "السعة",
    "label_planned_date": "التاريخ المخطط", "label_generated": "تاريخ الإنشاء",
    "label_generated_at": "تم الإنشاء في", "label_doc_no": "رقم الوثيقة",
    "label_order": "الطلب", "label_shipper": "المُرسِل (المستودع)",
    "label_recipient": "المُستلِم", "label_date": "التاريخ",
    "label_name": "الاسم", "label_signature": "التوقيع",
    "label_reserves": "تحفظات / ملاحظات", "label_totals": "المجاميع",
    "label_fill_rate": "نسبة التعبئة (كجم)", "label_weight": "الوزن",
    "label_volume": "الحجم", "label_packages": "الطرود",
    "label_algo": "الخوارزمية", "label_type": "النوع", "label_case": "الخانة",
    "label_signatures": "التوقيعات", "label_shipper_short": "المُرسِل",
    "label_carrier_short": "الناقل", "label_recipient_colon": "المُستلِم",
    "section_planned_stops": "المحطات المخططة",
    "section_recent_orders": "الطلبات الأخيرة (30)", "section_stop_detail": "تفاصيل المحطات",
    "section_dock_check": "مشرف الرصيف — مراقبة التحميل",
    "col_order": "الترتيب", "col_type": "النوع", "col_client": "العميل",
    "col_address": "العنوان", "col_order_ref": "مرجع الطلب",
    "col_planned_arrival": "الوصول المخطط", "col_planned_departure": "المغادرة المخططة",
    "col_duration_min": "المدة (دقيقة)", "col_dist_prev_km": "المسافة السابقة كم",
    "col_load_kg": "الحمولة كجم", "col_status": "الحالة",
    "col_reg": "لوحة الترخيص", "col_driver": "السائق", "col_km": "كم",
    "col_duration_min_short": "مدة دقيقة", "col_co2_kg": "CO₂ كجم",
    "col_stops": "المحطات", "col_on_time": "في الوقت المحدد",
    "col_metric": "المؤشر", "col_period": "الفترة", "col_prev_period": "الأسبوع السابق",
    "row_optim_runs": "تشغيلات التحسين", "row_avg_dist": "متوسط المسافة (كم)",
    "row_avg_respect": "متوسط احترام الفترات %",
    "chart_optim_per_day": "عدد التحسينات / يوم",
    "col_algo": "الخوارزمية", "col_dist_km": "مسافة كم", "col_duration": "المدة",
    "col_cost": "التكلفة", "col_respect_pct": "الاحترام %", "col_avg_delay": "متوسط التأخير",
    "col_cpu_ms": "CPU ms", "col_clients": "العملاء",
    "col_field": "الحقل", "col_value": "القيمة", "col_id": "المعرّف",
    "col_name": "الاسم", "col_coords": "الإحداثيات", "col_demand_kg": "الطلب كجم",
    "col_priority": "الأولوية", "col_contact": "جهة الاتصال", "col_phone": "الهاتف",
    "col_ref": "المرجع", "col_date": "التاريخ", "col_kg": "كجم",
    "col_n_routes": "المسارات", "col_max_drive": "الحد الأقصى للقيادة قبل الاستراحة (دقيقة)",
    "col_pause_min": "استراحة دقيقة", "col_max_h_day": "الحد اليومي (ساعة)",
    "col_rest_min_day": "الراحة اليومية (دقيقة)", "col_route_duration_min": "مدة المسار دقيقة",
    "col_alert": "تنبيه", "alert_near_limit": " قريب من الحد اليومي",
    "alert_check_breaks": "ℹ مراجعة الاستراحات",
    "col_zones": "المناطق", "col_cost_per_km": "€/كم", "col_rating": "التقييم",
    "col_shipments": "الشحنات", "col_delivered": "المُسلَّمة", "col_total_cost": "التكلفة الإجمالية",
    "col_article_desc": "المادة / الوصف", "col_qty_ordered": "الكمية المطلوبة (كجم)",
    "col_qty_delivered": "الكمية المُسلَّمة", "col_observations": "الملاحظات",
    "col_num": "#", "col_reference": "المرجع", "col_qty": "الكمية",
    "col_weight_kg": "الوزن كجم", "col_vol_m3": "الحجم م³",
    "col_instructions": "التعليمات / الوصول",
    "col_un_designation": "التسمية الأممية الرسمية", "col_un_no": "رقم ONU",
    "col_adr_class": "فئة ADR", "col_packaging_group": "مجموعة التعبئة",
    "col_qty_packages": "الكمية / الطرود", "col_temperature": "درجة الحرارة",
    "col_time_window": "النافذة الزمنية", "col_arrival": "الوصول",
    "col_departure": "المغادرة", "col_load_kg_paren": "الحمولة (كجم)",
    "row_total_dist": "المسافة الإجمالية", "row_est_duration": "المدة التقديرية",
    "row_total_load": "الحمولة الإجمالية", "row_est_cost": "التكلفة التقديرية",
    "row_time_respect": "احترام الجداول الزمنية", "row_dist_source": "مصدر المسافة",
    "val_osrm_source": "طريق حقيقي (OSRM)", "val_haversine_source": "تقدير (Haversine)",
    "chart_km_by_vehicle": "الكيلومترات المخططة لكل مركبة",
    "msg_no_route": "لا توجد مسارات", "msg_no_order": "لا توجد طلبات",
    "msg_no_carrier": "لا يوجد ناقلون", "msg_no_stop": "لا توجد محطات",
    "cmr_case": "الخانة", "cmr_c1_title": "المُرسِل (الاسم، العنوان، الدولة)",
    "cmr_c2_title": "المُستلِم (الاسم، العنوان، الدولة)",
    "cmr_c3_title": "مكان التسليم المقرر للبضائع",
    "cmr_c6_title": "الناقل (الاسم، العنوان، الدولة) — الناقل الأول",
    "cmr_c11_title": "طبيعة البضاعة", "cmr_c12_title": "طريقة التعبئة",
    "cmr_c13_title": "العدد والوصف والوزن الإجمالي بالكجم",
    "cmr_c18_title": "تحفظات وملاحظات الناقل",
    "cmr_c23_title": "توقيع وختم المُرسِل", "cmr_c24_title": "توقيع وختم المُستلِم",
    "rse_subtitle": "العتبات المُعلَنة مقابل الحمل المُخطَّط المُجمَّع (مدة المسار). تحليل إرشادي — التحقق التجاري مطلوب.",
    "legal_body": ("هذه الوثيقة نموذج إعلامي صادر عن CityPulse Logistics. "
      "لا تشكّل استشارة قانونية. يرجى تكييفها مع قسمكم القانوني "
      "قبل أي توزيع. البيانات المعالَجة: التخطيط، تحديد الموقع الجغرافي، "
      "المعرّفات المهنية. مدة الاحتفاظ وفق السياسة الداخلية والالتزامات القانونية."),
    "manifest_certif": "أشهد أن الشحنة تطابق هذا البيان.",
    "bl_sig_text": "توقيع المُستلِم (مسبوقاً بعبارة «موافق»)",
    "adr_declaration": ("يُعلن الناقل أن البضاعة مقبولة للنقل وفقاً لأحكام ADR. "
      "يُؤكد المُرسِل أن التصنيف والتعبئة والتغليف والوثائق "
      "متوافقة مع اللوائح المعمول بها. <i>نص نموذجي — يُكيَّف حسب استغلالكم.</i>"),
    "footer_confidential": "سري — CityPulse Logistics — التحقق من النوافذ الزمنية قبل المغادرة. في حالة حادث، اتصل بمركز التوزيع.",
    "footer_generated": "تم الإنشاء في", "footer_keep": "احتفظ به للتتبع.",
    "footer_algo": "الخوارزمية", "footer_confidential_route_2": "هذه الوثيقة سرية.",
    "footer_generated_by": "وثيقة صادرة عن CityPulse Logistics",
  },
}


def _RL(lang: str, key: str) -> str:
  """Report label lookup with French fallback."""
  return _RL_DATA.get(lang, _RL_DATA["fr"]).get(key) or _RL_DATA["fr"].get(key, key)


def _history_insert(report_type: str, parameters: dict | None, file_path: str, user_id: int | None = None):
  try:
    conn = get_connection()
    conn.execute(
      """INSERT INTO reports_history (report_type, parameters_json, file_path, file_size_kb, generated_by)
        VALUES (?,?,?,?,?)""",
      (report_type, json.dumps(parameters or {}, ensure_ascii=False), file_path, _file_size_kb(file_path), user_id),
    )
    conn.commit()
    conn.close()
  except Exception:
    logger.exception("reports_history insert failed")
  log_action("REPORT_GENERATED", f"{report_type} → {file_path}")


def _qr_flowable(payload: str, size_cm: float = 2.2):
  if not REPORTLAB_OK:
    return Spacer(1, 0.1 * cm)
  if not HAS_QR:
    return Paragraph(f"<i>QR: {payload[:40]}…</i>", getSampleStyleSheet()["Normal"])
  buf = io.BytesIO()
  img = qrcode.make(payload, box_size=2, border=1)
  img.save(buf, format="PNG")
  buf.seek(0)
  return RLImage(buf, width=size_cm * cm, height=size_cm * cm)


def _mpl_figure_to_image(fig: "Figure", width_cm: float = 16, height_cm: float = 8):
  buf = io.BytesIO()
  canvas = FigureCanvasAgg(fig)
  canvas.print_png(buf, dpi=120)
  buf.seek(0)
  return RLImage(buf, width=width_cm * cm, height=height_cm * cm)


def _page_footer(canvas, doc):
  canvas.saveState()
  canvas.setFont("Helvetica", 8)
  canvas.setFillColor(CLR_MUTED)
  canvas.drawRightString(A4[0] - 1.8 * cm, 1.2 * cm, f"Page {doc.page} — CityPulse Logistics")
  canvas.restoreState()


class _FakeRow:
  """Row factice (interface sqlite3.Row) pour _row_dict."""

  def __init__(self, d: dict):
    self._d = d

  def keys(self):
    return self._d.keys()

  def __getitem__(self, k):
    return self._d[k]


class ReportService:
  """Service centralisé de génération de rapports (PDF, XLSX, JSON)."""

  # ── 1. Carnet de route chauffeur ────────────────────────────────────────

  def generate_driver_roadbook(self, route_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    conn = get_connection()
    route = conn.execute(
      """SELECT r.*, v.registration, v.type as vtype, v.capacity_kg,
           d.first_name as dfname, d.last_name as dlname, d.phone as dphone,
           d.license_number
        FROM routes r
        JOIN vehicles v ON r.vehicle_id = v.id
        LEFT JOIN drivers d ON r.driver_id = d.id
        WHERE r.id = ?""",
      (route_id,),
    ).fetchone()
    if not route:
      conn.close()
      raise ValueError(f"Tournée introuvable : id={route_id}")

    stops = conn.execute(
      """SELECT rs.*, o.reference as order_ref, c.name as client_name, c.address as client_addr,
           c.latitude as clat, c.longitude as clon, c.demand_kg
        FROM route_stops rs
        LEFT JOIN orders o ON rs.order_id = o.id
        LEFT JOIN clients c ON o.client_id = c.id
        WHERE rs.route_id = ? ORDER BY rs.stop_order""",
      (route_id,),
    ).fetchall()
    conn.close()

    r = _row_dict(route)
    driver_name = f"{r.get('dfname') or ''} {r.get('dlname') or ''}".strip() or "—"
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    doc = SimpleDocTemplate(
      output_path,
      pagesize=A4,
      leftMargin=1.5 * cm,
      rightMargin=1.5 * cm,
      topMargin=1.6 * cm,
      bottomMargin=1.8 * cm,
      title=f"{L('roadbook_title')} — {r.get('registration')}",
      author="CityPulse Logistics",
    )
    story = []
    styles = getSampleStyleSheet()

    # Logo texte + chauffeur + véhicule
    hdr = Table(
      [
        [
          Paragraph(
            f"<b><font color='#1E3A5F' size='14'>CityPulse Logistics</font></b><br/>"
            f"<font color='#3B82F6' size='11'>{L('roadbook_title')}</font>",
            styles["Normal"],
          ),
          Paragraph(
            f"<b>{L('label_driver')} :</b> {driver_name}<br/>"
            f"<b>{L('label_phone')} :</b> {r.get('dphone') or '—'}<br/>"
            f"<b>{L('label_license')} :</b> {r.get('license_number') or '—'}<br/>"
            f"<b>{L('label_vehicle')} :</b> {r.get('registration')} ({r.get('vtype')})<br/>"
            f"<b>{L('label_capacity')} :</b> {r.get('capacity_kg') or 0:.0f} kg<br/>"
            f"<b>{L('label_planned_date')} :</b> {r.get('planned_date') or '—'}<br/>"
            f"<b>{L('label_generated')} :</b> {now_str}",
            ParagraphStyle("meta", fontSize=9, textColor=CLR_TEXT),
          ),
        ]
      ],
      colWidths=[7.5 * cm, 9.5 * cm],
    )
    hdr.setStyle(
      TableStyle(
        [
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("BOX", (0, 0), (-1, -1), 0.5, CLR_BORDER),
          ("BACKGROUND", (0, 0), (-1, -1), CLR_LIGHT),
          ("LEFTPADDING", (0, 0), (-1, -1), 8),
          ("TOPPADDING", (0, 0), (-1, -1), 8),
        ]
      )
    )
    story.append(hdr)
    story.append(Spacer(1, 0.35 * cm))

    story.append(
      Paragraph(
        f"<b>{L('section_planned_stops')}</b>",
        ParagraphStyle("h2", fontSize=11, textColor=CLR_PRIMARY, spaceAfter=6, fontName="Helvetica-Bold"),
      )
    )

    for s in stops:
      sd = _row_dict(s)
      payload = json.dumps(
        {
          "stop_id": sd.get("id"),
          "route_id": route_id,
          "order": sd.get("order_ref"),
          "client": sd.get("client_name"),
        },
        ensure_ascii=False,
      )
      row_tbl = Table(
        [
          [
            _qr_flowable(payload, 2.0),
            Table(
              [
                [L("col_order"), str(sd.get("stop_order", ""))],
                [L("col_type"), str(sd.get("stop_type", ""))],
                [L("col_client"), str(sd.get("client_name") or "—")[:40]],
                [L("col_address"), str(sd.get("client_addr") or "—")[:55]],
                [L("col_order_ref"), str(sd.get("order_ref") or "—")],
                [L("col_planned_arrival"), str(sd.get("planned_arrival") or "—")],
                [L("col_planned_departure"), str(sd.get("planned_departure") or "—")],
                [L("col_duration_min"), str(sd.get("duration_min") or "")],
                [L("col_dist_prev_km"), f"{sd.get('distance_from_prev_km') or 0:.2f}"],
                [L("col_load_kg"), f"{sd.get('demand_kg') or 0:.1f}"],
                [L("col_status"), str(sd.get("status") or "pending")],
              ],
              colWidths=[3.2 * cm, 10.5 * cm],
            ),
          ]
        ],
        colWidths=[2.4 * cm, 14.5 * cm],
      )
      row_tbl.setStyle(
        TableStyle(
          [
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOX", (0, 0), (-1, -1), 0.5, CLR_BORDER),
            ("BACKGROUND", (1, 0), (1, 0), colors.white),
            ("ROWBACKGROUNDS", (1, 0), (1, -1), [CLR_LIGHT, colors.white]),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (1, 0), (1, -1), 0.25, CLR_BORDER),
          ]
        )
      )
      story.append(row_tbl)
      story.append(Spacer(1, 0.2 * cm))

    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=CLR_BORDER))
    story.append(
      Paragraph(
        f"<font size='8' color='#6B7280'>{L('footer_confidential')}</font>",
        ParagraphStyle("ft", alignment=TA_CENTER, fontSize=8, textColor=CLR_MUTED),
      )
    )

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("driver_roadbook", {"route_id": route_id}, output_path)
    return output_path

  # ── 2. Rapport flotte journalier ────────────────────────────────────────

  def generate_fleet_daily_report(self, date_str: str, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    if not HAS_MPL:
      logger.warning("matplotlib absent — rapport sans graphique")

    conn = get_connection()
    rows = conn.execute(
      """SELECT r.*, v.registration,
           d.first_name || ' ' || d.last_name as driver_full
        FROM routes r
        JOIN vehicles v ON r.vehicle_id = v.id
        LEFT JOIN drivers d ON r.driver_id = d.id
        WHERE r.planned_date = ? ORDER BY v.registration""",
      (date_str,),
    ).fetchall()
    conn.close()

    doc = SimpleDocTemplate(
      output_path,
      pagesize=A4,
      leftMargin=1.5 * cm,
      rightMargin=1.5 * cm,
      topMargin=1.5 * cm,
      bottomMargin=1.8 * cm,
      title=f"Flotte du {date_str}",
    )
    story = []
    story.append(
      Paragraph(
        f"<b><font size='18' color='#1E3A5F'>{L('fleet_daily_title')}</font></b><br/>"
        f"<font size='12'>{L('label_date')} : {date_str}</font>",
        getSampleStyleSheet()["Title"],
      )
    )
    story.append(Spacer(1, 0.5 * cm))

    if HAS_MPL and rows:
      regs = [(_row_dict(x)["registration"] or "")[:12] for x in rows]
      kms = [float(_row_dict(x).get("total_km") or 0) for x in rows]
      fig = Figure(figsize=(8, 3.5))
      ax = fig.add_subplot(111)
      ax.bar(regs, kms, color="#3B82F6")
      ax.set_ylabel("km")
      ax.set_title(L("chart_km_by_vehicle"))
      ax.tick_params(axis="x", rotation=35)
      fig.tight_layout()
      story.append(_mpl_figure_to_image(fig, width_cm=17, height_cm=7))
      story.append(Spacer(1, 0.4 * cm))

    data = [[L("col_reg"), L("col_driver"), L("col_status"), L("col_km"), L("col_duration_min_short"), L("col_co2_kg"), L("col_stops"), L("col_on_time")]]
    for x in rows:
      d = _row_dict(x)
      data.append(
        [
          str(d.get("registration") or ""),
          str(d.get("driver_full") or "—")[:22],
          str(d.get("status") or ""),
          f"{d.get('total_km') or 0:.1f}",
          f"{d.get('total_duration_min') or 0:.0f}",
          f"{d.get('co2_kg') or 0:.2f}",
          str(d.get("stops_count") or "—"),
          str(d.get("on_time_count") or "—"),
        ]
      )
    if len(data) == 1:
      data.append(["—", L("msg_no_route"), "", "", "", "", "", ""])

    t = Table(data, repeatRows=1)
    t.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
          ("FONTSIZE", (0, 0), (-1, -1), 8),
          ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
          ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CLR_LIGHT]),
        ]
      )
    )
    story.append(t)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("fleet_daily", {"date": date_str}, output_path)
    return output_path

  # ── 3. Rapport KPI période + S-1 ────────────────────────────────────────

  def generate_kpi_report(
    self, start_date: str, end_date: str, output_path: str, fmt: str = "pdf", lang: str = "fr"
  ) -> str:
    fmt = (fmt or "pdf").lower()
    L = lambda k: _RL(lang, k)
    d0 = datetime.strptime(start_date, "%Y-%m-%d").date()
    d1 = datetime.strptime(end_date, "%Y-%m-%d").date()
    span = (d1 - d0).days + 1
    prev_end = d0 - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span - 1)

    conn = get_connection()
    cur = conn.execute(
      """SELECT date(created_at) as d,
           COUNT(*) as n,
           AVG(total_distance) as dist,
           AVG(respect_rate) as resp
        FROM algo_results
        WHERE date(created_at) BETWEEN ? AND ?
        GROUP BY date(created_at) ORDER BY d""",
      (start_date, end_date),
    ).fetchall()

    prev = conn.execute(
      """SELECT COUNT(*) as n,
           AVG(total_distance) as dist,
           AVG(respect_rate) as resp
        FROM algo_results
        WHERE date(created_at) BETWEEN ? AND ?""",
      (prev_start.isoformat(), prev_end.isoformat()),
    ).fetchone()
    conn.close()

    cur_agg = {"n": 0, "dist": 0.0, "resp": 0.0}
    days = []
    counts = []
    for row in cur:
      rd = _row_dict(row)
      days.append(str(rd.get("d", "")))
      counts.append(int(rd.get("n") or 0))
      cur_agg["n"] += int(rd.get("n") or 0)
    if cur:
      cur_agg["dist"] = sum(float(_row_dict(x).get("dist") or 0) for x in cur) / max(len(cur), 1)
      cur_agg["resp"] = sum(float(_row_dict(x).get("resp") or 0) for x in cur) / max(len(cur), 1)

    prev_d = _row_dict(prev) if prev else {}
    prev_agg = {
      "n": int(prev_d.get("n") or 0),
      "dist": float(prev_d.get("dist") or 0),
      "resp": float(prev_d.get("resp") or 0),
    }

    params = {"start": start_date, "end": end_date, "fmt": fmt}

    if fmt == "xlsx":
      _require_openpyxl()
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "KPI"
      ws.append(["Période", f"{start_date} → {end_date}"])
      ws.append(["Période S-1", f"{prev_start} → {prev_end}"])
      ws.append([])
      ws.append([L("col_metric"), "Courant", L("col_prev_period")])
      ws.append([L("row_optim_runs"), cur_agg["n"], prev_agg["n"]])
      ws.append([L("row_avg_dist"), f"{cur_agg['dist']:.2f}", f"{prev_agg['dist']:.2f}"])
      ws.append([L("row_avg_respect"), f"{cur_agg['resp']:.1f}", f"{prev_agg['resp']:.1f}"])
      ws.append([])
      ws.append(["Date", "Runs", "Dist moy", "Respect %"])
      for row in cur:
        rd = _row_dict(row)
        ws.append([rd.get("d"), rd.get("n"), rd.get("dist"), rd.get("resp")])
      wb.save(output_path)
      _history_insert("kpi_report", params, output_path)
      return output_path

    _require_reportlab()
    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = []
    story.append(
      Paragraph(
        f"<b>{L('kpi_title')}</b><br/>{start_date} → {end_date} "
        f"<i>(vs période précédente {prev_start} → {prev_end})</i>",
        ParagraphStyle("t", fontSize=14, textColor=CLR_PRIMARY),
      )
    )
    story.append(Spacer(1, 0.4 * cm))

    cmp_data = [
      [L("col_metric"), L("col_period"), L("col_prev_period")],
      [L("row_optim_runs"), str(cur_agg["n"]), str(prev_agg["n"])],
      [L("row_avg_dist"), f"{cur_agg['dist']:.2f}", f"{prev_agg['dist']:.2f}"],
      [L("row_avg_respect"), f"{cur_agg['resp']:.1f}", f"{prev_agg['resp']:.1f}"],
    ]
    ct = Table(cmp_data)
    ct.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_SECONDARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.3, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
      )
    )
    story.append(ct)
    story.append(Spacer(1, 0.5 * cm))

    if HAS_MPL and days:
      fig = Figure(figsize=(8, 3.2))
      ax = fig.add_subplot(111)
      ax.plot(days, counts, marker="o", color="#3B82F6")
      ax.set_title(L("chart_optim_per_day"))
      ax.tick_params(axis="x", rotation=45)
      fig.tight_layout()
      story.append(_mpl_figure_to_image(fig, 17, 6.5))

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("kpi_report", params, output_path)
    return output_path

  # ── 4. Comparaison algorithmes ────────────────────────────────────────────

  def generate_algo_comparison_report(self, result_ids: Sequence[int], output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    if not result_ids:
      raise ValueError("result_ids vide")
    conn = get_connection()
    placeholders = ",".join("?" * len(result_ids))
    rows = conn.execute(
      f"SELECT * FROM algo_results WHERE id IN ({placeholders}) ORDER BY total_distance ASC",
      tuple(result_ids),
    ).fetchall()
    conn.close()
    if not rows:
      raise ValueError("Aucun algo_results pour ces id")

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
      Paragraph(
        f"<b>{L('algo_cmp_title')}</b>",
        ParagraphStyle("x", fontSize=14, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.4 * cm),
    ]
    hdr = [
      L("col_algo"), L("col_dist_km"), L("col_duration"), L("col_cost"),
      L("col_respect_pct"), L("col_avg_delay"), L("col_cpu_ms"), L("col_clients"),
    ]
    data = [hdr]
    best_dist = None
    for row in rows:
      r = _row_dict(row)
      dist = float(r.get("total_distance") or 0)
      if best_dist is None:
        best_dist = dist
      data.append(
        [
          str(r.get("algorithm", "")),
          f"{dist:.2f}",
          f"{r.get('total_duration') or 0:.0f}",
          f"{r.get('total_cost') or 0:.2f}",
          f"{r.get('respect_rate') or 0:.1f}",
          f"{r.get('avg_delay') or 0:.1f}",
          f"{r.get('cpu_time_ms') or 0:.0f}",
          str(r.get("client_count") or "—"),
        ]
      )

    tbl = Table(data, repeatRows=1)
    style = [
      ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
      ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
      ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
      ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]
    for i, row in enumerate(rows, 1):
      r = _row_dict(row)
      if float(r.get("total_distance") or 0) == best_dist:
        style.append(("BACKGROUND", (0, i), (-1, i), colors.Color(0.85, 0.95, 0.85)))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    if HAS_MPL and len(rows) >= 2:
      fig = Figure(figsize=(8, 3.5))
      ax = fig.add_subplot(111)
      names = [(_row_dict(x).get("algorithm") or "")[:14] for x in rows]
      dists = [float(_row_dict(x).get("total_distance") or 0) for x in rows]
      ax.barh(names, dists, color="#3B82F6")
      ax.set_xlabel("km")
      fig.tight_layout()
      story.append(Spacer(1, 0.3 * cm))
      story.append(_mpl_figure_to_image(fig, 16, 7))

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("algo_comparison", {"result_ids": list(result_ids)}, output_path)
    return output_path

  # ── 5. Fiche client ─────────────────────────────────────────────────────

  def generate_client_report(self, client_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    conn = get_connection()
    c = conn.execute("SELECT * FROM clients WHERE id= ?", (client_id,)).fetchone()
    if not c:
      conn.close()
      raise ValueError(f"Client {client_id} introuvable")
    cd = _row_dict(c)
    orders = conn.execute(
      """SELECT reference, status, scheduled_date, quantity_kg, operation_type
        FROM orders WHERE client_id= ? AND archived=0
        ORDER BY created_at DESC LIMIT 30""",
      (client_id,),
    ).fetchall()
    conn.close()

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
      Paragraph(
        f"<b>{L('client_report_title')}</b> — {cd.get('name', '')}",
        ParagraphStyle("t", fontSize=14, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.3 * cm),
    ]
    info = [
      [L("col_field"), L("col_value")],
      [L("col_id"), str(cd.get("id"))],
      [L("col_name"), str(cd.get("name"))],
      [L("col_address"), str(cd.get("address") or "—")],
      [L("col_coords"), f"{cd.get('latitude')}, {cd.get('longitude')}"],
      [L("col_demand_kg"), str(cd.get("demand_kg"))],
      [L("col_type"), str(cd.get("client_type"))],
      [L("col_priority"), str(cd.get("priority"))],
      [L("col_contact"), str(cd.get("contact") or "—")],
      [L("col_phone"), str(cd.get("phone") or "—")],
    ]
    t = Table(info, colWidths=[4 * cm, 12 * cm])
    t.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_SECONDARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
      )
    )
    story.append(t)
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(f"<b>{L('section_recent_orders')}</b>", getSampleStyleSheet()["Heading2"]))
    od = [[L("col_ref"), L("col_status"), L("col_date"), L("col_kg"), L("col_type")]]
    for o in orders:
      od.append(
        [
          str(_row_dict(o).get("reference") or ""),
          str(_row_dict(o).get("status") or ""),
          str(_row_dict(o).get("scheduled_date") or "—"),
          str(_row_dict(o).get("quantity_kg") or ""),
          str(_row_dict(o).get("operation_type") or ""),
        ]
      )
    if len(od) == 1:
      od.append(["—", L("msg_no_order"), "", "", ""])
    ot = Table(od)
    ot.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]
      )
    )
    story.append(ot)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("client_report", {"client_id": client_id}, output_path)
    return output_path

  # ── 6. Performance chauffeurs ───────────────────────────────────────────

  def generate_driver_performance_report(
    self, period_days: int, output_path: str, fmt: str = "pdf", lang: str = "fr"
  ) -> str:
    fmt = (fmt or "pdf").lower()
    L = lambda k: _RL(lang, k)
    since = (date.today() - timedelta(days=max(1, period_days))).isoformat()

    conn = get_connection()
    rows = conn.execute(
      """SELECT d.id, d.first_name, d.last_name,
           COUNT(r.id) as n_routes,
           SUM(r.total_km) as km,
           SUM(r.on_time_count) as on_time,
           SUM(r.stops_count) as stops
        FROM drivers d
        LEFT JOIN routes r ON r.driver_id = d.id AND r.planned_date >= ?
        WHERE d.archived = 0
        GROUP BY d.id
        ORDER BY COALESCE(km, 0) DESC""",
      (since,),
    ).fetchall()
    conn.close()

    if fmt == "xlsx":
      _require_openpyxl()
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "Performance"
      ws.append(["Période (jours)", period_days, "Depuis", since])
      ws.append([])
      ws.append(["ID", "Nom", "Tournées", "Km", "Arrêts à l'heure", "Arrêts tot."])
      for x in rows:
        r = _row_dict(x)
        ws.append(
          [
            r.get("id"),
            f"{r.get('first_name', '')} {r.get('last_name', '')}",
            r.get("n_routes"),
            r.get("km") or 0,
            r.get("on_time"),
            r.get("stops"),
          ]
        )
      wb.save(output_path)
      _history_insert("driver_performance", {"days": period_days, "fmt": fmt}, output_path)
      return output_path

    _require_reportlab()
    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
      Paragraph(
        f"<b>{L('drv_perf_title')}</b> — {period_days} j (depuis {since})",
        ParagraphStyle("t", fontSize=13, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.35 * cm),
    ]
    data = [[L("col_driver"), L("col_n_routes"), L("col_km"), L("col_on_time"), L("col_stops")]]
    for x in rows:
      r = _row_dict(x)
      data.append(
        [
          f"{r.get('first_name', '')} {r.get('last_name', '')}"[:28],
          str(r.get("n_routes") or 0),
          f"{float(r.get('km') or 0):.1f}",
          str(r.get("on_time") or 0),
          str(r.get("stops") or 0),
        ]
      )
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
      )
    )
    story.append(tbl)

    if HAS_MPL and rows:
      fig = Figure(figsize=(8, 3.5))
      ax = fig.add_subplot(111)
      labels = [f"{_row_dict(x).get('first_name', '')[:1]}. {_row_dict(x).get('last_name', '')[:10]}" for x in rows[:12]]
      vals = [float(_row_dict(x).get("km") or 0) for x in rows[:12]]
      ax.bar(labels, vals, color="#22C55E")
      ax.tick_params(axis="x", rotation=35)
      ax.set_ylabel("km")
      fig.tight_layout()
      story.append(Spacer(1, 0.4 * cm))
      story.append(_mpl_figure_to_image(fig, 17, 7))

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("driver_performance", {"days": period_days, "fmt": "pdf"}, output_path)
    return output_path

  # ── 7. Conformité RSE (paramètres chauffeurs + activité) ────────────────

  def generate_rse_compliance_report(self, start_date: str, end_date: str, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    conn = get_connection()
    drivers = conn.execute(
      "SELECT id, first_name, last_name, max_drive_before_break_min, "
      "min_break_minutes, max_daily_hours, min_daily_rest_minutes "
      "FROM drivers WHERE archived=0"
    ).fetchall()
    route_load = conn.execute(
      """SELECT driver_id, SUM(total_duration_min) as mins
        FROM routes
        WHERE planned_date BETWEEN ? AND ? AND driver_id IS NOT NULL
        GROUP BY driver_id""",
      (start_date, end_date),
    ).fetchall()
    conn.close()
    load_map = {int(_row_dict(x)["driver_id"]): float(_row_dict(x).get("mins") or 0) for x in route_load}

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
      Paragraph(
        f"<b>{L('rse_title')}</b><br/>{start_date} → {end_date}",
        ParagraphStyle("t", fontSize=13, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.3 * cm),
      Paragraph(
        L("rse_subtitle"),
        ParagraphStyle("s", fontSize=9, textColor=CLR_MUTED),
      ),
      Spacer(1, 0.25 * cm),
    ]
    data = [[L("col_driver"), L("col_max_drive"), L("col_pause_min"), L("col_max_h_day"), L("col_rest_min_day"), L("col_route_duration_min"), L("col_alert")]]
    for d in drivers:
      r = _row_dict(d)
      did = int(r["id"])
      mins = load_map.get(did, 0.0)
      max_drive = int(r.get("max_drive_before_break_min") or 270)
      max_daily_h = float(r.get("max_daily_hours") or 10) * 60
      alert = "—"
      if mins > max_daily_h * 0.95:
        alert = L("alert_near_limit")
      elif mins > max_drive * 3:
        alert = L("alert_check_breaks")
      data.append(
        [
          f"{r.get('first_name', '')} {r.get('last_name', '')}"[:24],
          str(max_drive),
          str(r.get("min_break_minutes") or 45),
          str(r.get("max_daily_hours") or 10),
          str(r.get("min_daily_rest_minutes") or 660),
          f"{mins:.0f}",
          alert,
        ]
      )
    tbl = Table(data, repeatRows=1, colWidths=[3.5 * cm, 2.2 * cm, 1.5 * cm, 1.5 * cm, 2 * cm, 2.2 * cm, 3.6 * cm])
    tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("FONTSIZE", (0, 0), (-1, -1), 7),
          ("GRID", (0, 0), (-1, -1), 0.2, CLR_BORDER),
        ]
      )
    )
    story.append(tbl)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("rse_compliance", {"start": start_date, "end": end_date}, output_path)
    return output_path

  # ── 8. Transporteurs ────────────────────────────────────────────────────

  def generate_carrier_report(self, carrier_id: int | None = None, output_path: str | None = None, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    if not output_path:
      raise ValueError("output_path requis")
    conn = get_connection()
    if carrier_id is not None:
      carriers = conn.execute(
        "SELECT * FROM carriers WHERE id= ? AND archived=0", (carrier_id,)
      ).fetchall()
    else:
      carriers = conn.execute("SELECT * FROM carriers WHERE archived=0 ORDER BY name").fetchall()
    story_blocks = []
    for c in carriers:
      cd = _row_dict(c)
      cid = cd["id"]
      ship = conn.execute(
        """SELECT COUNT(*) as n,
             SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as deliv,
             SUM(cost) as cost
          FROM carrier_shipments WHERE carrier_id= ?""",
        (cid,),
      ).fetchone()
      sd = _row_dict(ship) if ship else {}
      story_blocks.append(
        (
          cd,
          sd,
        )
      )
    conn.close()

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    story = [
      Paragraph(
        f"<b>{L('carrier_title')}</b>",
        ParagraphStyle("t", fontSize=14, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.35 * cm),
    ]
    data = [[L("label_carrier_short"), L("col_zones"), L("col_cost_per_km"), L("col_rating"), L("col_shipments"), L("col_delivered"), L("col_total_cost")]]
    for cd, sd in story_blocks:
      data.append(
        [
          str(cd.get("name", ""))[:22],
          str((cd.get("zones_covered") or "—"))[:18],
          f"{cd.get('cost_per_km') or 0:.2f}",
          f"{cd.get('rating') or 0:.1f}",
          str(sd.get("n") or 0),
          str(sd.get("deliv") or 0),
          f"{float(sd.get('cost') or 0):.2f}",
        ]
      )
    if len(data) == 1:
      data.append(["—", L("msg_no_carrier"), "", "", "", "", ""])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.25, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]
      )
    )
    story.append(tbl)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("carrier_report", {"carrier_id": carrier_id}, output_path)
    return output_path

  # ── 9. Export Excel multi-feuilles ─────────────────────────────────────

  def export_to_excel(self, output_path: str) -> str:
    _require_openpyxl()
    conn = get_connection()
    wb = openpyxl.Workbook()
    # remove default
    wb.remove(wb.active)

    def add_sheet(name: str, sql: str):
      ws = wb.create_sheet(title=name[:31])
      rows = conn.execute(sql).fetchall()
      if not rows:
        ws.append(["(vide)"])
        return
      keys = rows[0].keys()
      ws.append(list(keys))
      for row in rows:
        ws.append([row[k] for k in keys])

    add_sheet("Clients", "SELECT * FROM clients WHERE archived=0")
    add_sheet("Vehicules", "SELECT * FROM vehicles")
    add_sheet("Chauffeurs", "SELECT * FROM drivers WHERE archived=0")
    add_sheet("Commandes", "SELECT * FROM orders WHERE archived=0")
    add_sheet("Tournees", "SELECT * FROM routes")
    add_sheet("Journal", "SELECT * FROM logs ORDER BY created_at DESC LIMIT 5000")

    conn.close()
    wb.save(output_path)
    _history_insert("export_excel", {}, output_path)
    log_action("EXPORT_EXCEL_FULL", output_path)
    return output_path

  # ── 10. Snapshot JSON complet ───────────────────────────────────────────

  def generate_full_snapshot(self, output_path: str) -> str:
    tables = [
      "clients",
      "vehicles",
      "depots",
      "drivers",
      "orders",
      "routes",
      "route_stops",
      "algo_results",
      "carriers",
      "carrier_shipments",
      "notifications",
      "logs",
      "zones",
      "scenarios",
    ]
    conn = get_connection()
    snap = {"exported_at": datetime.now().isoformat(), "tables": {}}
    for t in tables:
      try:
        snap["tables"][t] = [_row_dict(r) for r in conn.execute(f"SELECT * FROM {t}").fetchall()]
      except sqlite3.Error:
        snap["tables"][t] = []
    conn.close()
    with open(output_path, "w", encoding="utf-8") as f:
      json.dump(snap, f, ensure_ascii=False, indent=2, default=str)
    _history_insert("full_snapshot", {}, output_path)
    log_action("EXPORT_JSON_SNAPSHOT", output_path)
    return output_path

  # ── Bon de livraison (BL) ───────────────────────────────────────────────

  def generate_delivery_note(self, order_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    conn = get_connection()
    o = conn.execute(
      """SELECT o.*, c.name AS client_name, c.address AS client_address,
           c.phone AS client_phone, c.contact AS client_contact
        FROM orders o
        LEFT JOIN clients c ON c.id = o.client_id
        WHERE o.id= ? AND o.archived=0""",
      (order_id,),
    ).fetchone()
    if not o:
      conn.close()
      raise ValueError(f"Commande {order_id} introuvable")
    od = _row_dict(o)

    depot = None
    did = od.get("depot_id")
    if did:
      depot = conn.execute("SELECT * FROM depots WHERE id= ?", (did,)).fetchone()
    if not depot:
      depot = conn.execute("SELECT * FROM depots ORDER BY id LIMIT 1").fetchone()
    dd = _row_dict(depot) if depot else {}
    conn.close()

    bl_no = f"BL-{order_id}-{datetime.now().strftime('%Y%m%d%H%M')}"
    styles = getSampleStyleSheet()
    story = [
      Paragraph(
        f"<b><font color='#1E3A5F' size='14'>CityPulse Logistics</font></b><br/>"
        f"<font size='11'><b>{L('bl_title')}</b></font>",
        styles["Normal"],
      ),
      Spacer(1, 0.2 * cm),
      Paragraph(f"<b>{L('label_doc_no')} :</b> {bl_no}", ParagraphStyle("n", fontSize=10)),
      Paragraph(f"<b>{L('label_order')} :</b> {od.get('reference') or '—'}", ParagraphStyle("n2", fontSize=9)),
      Spacer(1, 0.35 * cm),
    ]

    exp_lines = (
      f"<b>{dd.get('name') or 'Dépôt'}</b><br/>"
      f"{dd.get('address') or '—'}<br/>"
      f"{L('label_phone')} {dd.get('phone') or '—'}"
    )
    dest_lines = (
      f"<b>{od.get('client_name') or 'Client'}</b><br/>"
      f"{od.get('client_address') or '—'}<br/>"
      f"{L('col_contact')} : {od.get('client_contact') or '—'} — {od.get('client_phone') or '—'}"
    )
    hdr_tbl = Table(
      [
        [
          Paragraph(f"<b>{L('label_shipper')}</b><br/>" + exp_lines, ParagraphStyle("e", fontSize=8)),
          Paragraph(f"<b>{L('label_recipient')}</b><br/>" + dest_lines, ParagraphStyle("d", fontSize=8)),
        ]
      ],
      colWidths=[8.5 * cm, 8.5 * cm],
    )
    hdr_tbl.setStyle(
      TableStyle(
        [
          ("BOX", (0, 0), (-1, -1), 0.8, CLR_PRIMARY),
          ("BACKGROUND", (0, 0), (-1, -1), CLR_LIGHT),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("LEFTPADDING", (0, 0), (-1, -1), 6),
          ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]
      )
    )
    story.append(hdr_tbl)
    story.append(Spacer(1, 0.4 * cm))

    goods = str(od.get("goods_category") or "Marchandises diverses")
    q_cmd = float(od.get("quantity_kg") or 0)
    u_cnt = int(od.get("units_count") or 1)
    art_data = [
      [L("col_article_desc"), L("col_qty_ordered"), L("col_packages"), L("col_qty_delivered"), L("col_observations")],
      [goods[:55], f"{q_cmd:.2f}", str(u_cnt), "___________", ""],
    ]
    art = Table(art_data, colWidths=[6.5 * cm, 2.5 * cm, 2 * cm, 2.5 * cm, 3.5 * cm], repeatRows=1)
    art.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_SECONDARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.4, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 8),
          ("MINROWHEIGHT", (0, 1), (-1, -1), 28),
        ]
      )
    )
    story.append(art)
    story.append(Spacer(1, 0.5 * cm))
    story.append(
      Paragraph(
        f"<b>{L('label_reserves')} :</b><br/><br/>"
        "_________________________________________________________________<br/><br/>"
        f"<b>{L('bl_sig_text')}</b><br/><br/><br/>"
        f"{L('label_name')} : _________________________&nbsp;&nbsp;&nbsp;{L('label_date')} : ___________&nbsp;&nbsp;&nbsp;{L('label_signature')} : _________________________",
        ParagraphStyle("sig", fontSize=9, textColor=CLR_TEXT, leading=12),
      )
    )
    story.append(Spacer(1, 0.3 * cm))
    story.append(
      Paragraph(
        f"<font size='7' color='#6B7280'>{L('footer_generated')} {datetime.now().strftime('%d/%m/%Y %H:%M')} — {L('footer_keep')}</font>",
        ParagraphStyle("ft", alignment=TA_CENTER, fontSize=7),
      )
    )

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("delivery_note", {"order_id": order_id, "bl_no": bl_no}, output_path)
    return output_path

  # ── CMR (lettre de voiture) — formulaire cases numérotées ───────────────

  def _cmr_story_blocks(
    self,
    case1_exp: str,
    case2_dest: str,
    case3_place: str,
    case6_carrier: str,
    goods_nature: str,
    goods_weight: str,
    goods_packaging: str,
    case18_reserves: str,
    lang: str = "fr",
  ) -> list:
    L = lambda k: _RL(lang, k)
    def box(num: str, title: str, body: str):
      p_title = Paragraph(f"<b>{L('cmr_case')} {num}</b> — <i>{title}</i>", ParagraphStyle("ct", fontSize=8, textColor=CLR_PRIMARY))
      p_body = Paragraph(body.replace("\n", "<br/>"), ParagraphStyle("cb", fontSize=8, textColor=CLR_TEXT))
      t = Table([[p_title], [p_body]], colWidths=[17 * cm])
      t.setStyle(
        TableStyle(
          [
            ("BOX", (0, 0), (-1, -1), 0.6, CLR_BORDER),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
          ]
        )
      )
      return t

    story = [
      Paragraph(
        f"<b>{L('cmr_title')}</b> — "
        "<i>Document généré par logiciel — à valider juridiquement.</i>",
        ParagraphStyle("cmrh", fontSize=8, textColor=CLR_MUTED, leading=11),
      ),
      Spacer(1, 0.25 * cm),
      box("1", L("cmr_c1_title"), case1_exp),
      Spacer(1, 0.15 * cm),
      box("2", L("cmr_c2_title"), case2_dest),
      Spacer(1, 0.15 * cm),
      box("3", L("cmr_c3_title"), case3_place),
      Spacer(1, 0.15 * cm),
      box("6", L("cmr_c6_title"), case6_carrier),
      Spacer(1, 0.15 * cm),
    ]
    # Cases 11–13 marchandises (regroupées sur une ligne de tableau)
    g_tbl = Table(
      [
        [
          Paragraph(f"<b>{L('cmr_case')} 11</b><br/>{L('cmr_c11_title')}", ParagraphStyle("g", fontSize=7)),
          Paragraph(f"<b>{L('cmr_case')} 12</b><br/>{L('cmr_c12_title')}", ParagraphStyle("g", fontSize=7)),
          Paragraph(f"<b>{L('cmr_case')} 13</b><br/>{L('cmr_c13_title')}", ParagraphStyle("g", fontSize=7)),
        ],
        [
          Paragraph(goods_nature, ParagraphStyle("gv", fontSize=8)),
          Paragraph(goods_packaging, ParagraphStyle("gv", fontSize=8)),
          Paragraph(goods_weight, ParagraphStyle("gv", fontSize=8)),
        ],
      ],
      colWidths=[5.5 * cm, 5.5 * cm, 6 * cm],
    )
    g_tbl.setStyle(
      TableStyle(
        [
          ("BOX", (0, 0), (-1, -1), 0.6, CLR_BORDER),
          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
          ("GRID", (0, 0), (-1, -1), 0.4, CLR_BORDER),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]
      )
    )
    story.append(g_tbl)
    story.append(Spacer(1, 0.15 * cm))
    story.append(box("18", L("cmr_c18_title"), case18_reserves))
    story.append(Spacer(1, 0.25 * cm))
    sig_tbl = Table(
      [
        [
          Paragraph(
            f"<b>{L('cmr_case')} 23</b><br/>{L('cmr_c23_title')}<br/><br/><br/>_________________________",
            ParagraphStyle("s", fontSize=8),
          ),
          Paragraph(
            f"<b>{L('cmr_case')} 24</b><br/>{L('cmr_c24_title')}<br/><br/><br/>_________________________",
            ParagraphStyle("s", fontSize=8),
          ),
        ]
      ],
      colWidths=[8.5 * cm, 8.5 * cm],
    )
    sig_tbl.setStyle(
      TableStyle(
        [
          ("BOX", (0, 0), (-1, -1), 0.6, CLR_BORDER),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]
      )
    )
    story.append(sig_tbl)
    return story

  def generate_cmr(self, order_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    conn = get_connection()
    o = conn.execute(
      """SELECT o.*, c.name AS client_name, c.address AS client_address,
           c.latitude AS clat, c.longitude AS clon
        FROM orders o
        LEFT JOIN clients c ON c.id = o.client_id
        WHERE o.id= ? AND o.archived=0""",
      (order_id,),
    ).fetchone()
    if not o:
      conn.close()
      raise ValueError(f"Commande {order_id} introuvable")
    od = _row_dict(o)

    depot = None
    if od.get("depot_id"):
      depot = conn.execute("SELECT * FROM depots WHERE id= ?", (od["depot_id"],)).fetchone()
    if not depot:
      depot = conn.execute("SELECT * FROM depots ORDER BY id LIMIT 1").fetchone()
    dd = _row_dict(depot) if depot else {}

    veh = None
    if od.get("vehicle_id"):
      veh = conn.execute("SELECT * FROM vehicles WHERE id= ?", (od["vehicle_id"],)).fetchone()
    conn.close()
    vd = _row_dict(veh) if veh else {}

    exp = f"{dd.get('name') or '—'}\n{dd.get('address') or '—'}\nMaroc"
    dest = f"{od.get('client_name') or '—'}\n{od.get('client_address') or '—'}"
    place = f"{od.get('client_address') or '—'}\n(GPS approx. {od.get('clat') or '—'}, {od.get('clon') or '—'})"
    carrier = (
      f"CityPulse Logistics / sous-traitant agréé\n"
      f"Véhicule : {vd.get('registration') or 'à affecter'}\n"
      f"Type : {vd.get('type') or '—'}"
    )
    nature = str(od.get("goods_category") or "Marchandises")
    weight = f"{float(od.get('quantity_kg') or 0):.2f} kg — Réf. {od.get('reference') or '—'}"
    pack = f"Colis : {int(od.get('units_count') or 1)} — Temp. {od.get('temperature_required') or 'ambient'}"
    reserves = "Réserves : ________________________________"

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.4 * cm, rightMargin=1.4 * cm)
    story = self._cmr_story_blocks(exp, dest, place, carrier, nature, weight, pack, reserves, lang=lang)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("cmr", {"order_id": order_id}, output_path)
    return output_path

  def generate_cmr_from_optimization_route(self, route_info: dict, output_path: str) -> str:
    """CMR à partir d'une route VRP en mémoire (sans order_id obligatoire)."""
    _require_reportlab()
    veh = route_info.get("vehicle") or {}
    stops = [s for s in route_info.get("route", []) if s.get("type") == "delivery"]
    first = stops[0] if stops else {}
    c = first.get("client") or {}
    conn = get_connection()
    depot = conn.execute("SELECT * FROM depots ORDER BY id LIMIT 1").fetchone()
    dd = _row_dict(depot) if depot else {}
    conn.close()
    exp = f"{dd.get('name') or 'Dépôt'}\n{dd.get('address') or '—'}"
    dest = f"{c.get('name') or 'Destinataire'}\n{c.get('address') or '—'}"
    place = dest
    carrier = f"CityPulse Logistics\nVéhicule : {veh.get('registration') or '—'}\nType : {veh.get('type') or '—'}"
    nature = "Livraisons groupées (optimisation)"
    weight = f"{float(route_info.get('load_kg') or 0):.2f} kg (charge totale tournée)"
    pack = f"{len(stops)} arrêt(s) livraison"
    reserves = "—"

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.4 * cm, rightMargin=1.4 * cm)
    story = self._cmr_story_blocks(exp, dest, place, carrier, nature, weight, pack, reserves)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("cmr", {"source": "optimization_route"}, output_path)
    return output_path

  # ── Document ADR ─────────────────────────────────────────────────────────

  _ADR_UN_FALLBACK = {
    "1": ("UN 0494", "Matières et objets présentant un risque de explosion massive", "II"),
    "2": ("UN 1002", "Gaz comprimé N.S.A.", "II"),
    "3": ("UN 1202", "Gazole ou fioul domestique", "III"),
    "4.1": ("UN 2557", "Nitrocellulose modifiée", "II"),
    "4.2": ("UN 1383", "Matières pyrophoriques ou substances auto-réactives", "II"),
    "4.3": ("UN 1398", "Aluminium phosphure", "II"),
    "5.1": ("UN 1477", "Nitrates inorganiques N.S.A.", "II"),
    "5.2": ("UN 3109", "Peroxyde organique type F", "II"),
    "6.1": ("UN 2810", "Matières infectieuses", "II"),
    "6.2": ("UN 3373", "Échantillons biologiques catégorie B", "II"),
    "8": ("UN 2796", "Acide sulfurique", "II"),
    "9": ("UN 3077", "Matières dangereuses pour l'environnement", "III"),
  }

  def generate_adr_document(self, order_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    L = lambda k: _RL(lang, k)
    conn = get_connection()
    o = conn.execute(
      """SELECT o.*, c.name AS client_name, c.address AS client_address
        FROM orders o
        LEFT JOIN clients c ON c.id = o.client_id
        WHERE o.id= ? AND o.archived=0""",
      (order_id,),
    ).fetchone()
    if not o:
      conn.close()
      raise ValueError(f"Commande {order_id} introuvable")
    od = _row_dict(o)
    adr = (od.get("adr_class") or "").strip()
    if not adr:
      conn.close()
      raise ValueError("Cette commande n'a pas de classe ADR (adr_class vide).")
    conn.close()

    key = adr.split(".")[0] if adr else ""
    un_no, desig, pg = self._ADR_UN_FALLBACK.get(
      adr, self._ADR_UN_FALLBACK.get(key, ("UN —", "Matière dangereuse — voir fiche de sécurité", "—"))
    )

    decl = L("adr_declaration")

    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    story = [
      Paragraph(
        f"<b>{L('adr_title')}</b>",
        ParagraphStyle("at", fontSize=14, textColor=CLR_PRIMARY),
      ),
      Spacer(1, 0.3 * cm),
      Paragraph(f"<b>{L('label_order')} :</b> {od.get('reference') or order_id}", ParagraphStyle("a", fontSize=10)),
      Paragraph(f"<b>{L('label_recipient_colon')} :</b> {od.get('client_name') or '—'}", ParagraphStyle("a", fontSize=10)),
      Spacer(1, 0.4 * cm),
    ]
    data = [
      [L("col_field"), L("col_value")],
      [L("col_un_designation"), desig],
      [L("col_un_no"), un_no],
      [L("col_adr_class"), adr],
      [L("col_packaging_group"), pg],
      [L("col_qty_packages"), f"{od.get('quantity_kg') or 0} kg — {od.get('units_count') or 1}"],
      [L("col_temperature"), str(od.get("temperature_required") or "ambient")],
    ]
    t = Table(data, colWidths=[5 * cm, 11 * cm])
    t.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_DANGER),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.4, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]
      )
    )
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(decl, ParagraphStyle("decl", fontSize=9, textColor=CLR_TEXT, leading=13)))
    story.append(Spacer(1, 0.4 * cm))
    story.append(
      Paragraph(
        f"<b>{L('label_signatures')}</b><br/><br/>{L('label_shipper_short')} : _______________________ &nbsp;&nbsp; "
        f"{L('label_carrier_short')} : _______________________ &nbsp;&nbsp; {L('label_date')} : ___________",
        ParagraphStyle("sg", fontSize=9),
      )
    )
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("adr_document", {"order_id": order_id, "adr_class": adr}, output_path)
    return output_path

  # ── Manifeste de chargement ─────────────────────────────────────────────

  def generate_load_manifest(self, route_id: int, output_path: str, lang: str = "fr") -> str:
    _require_reportlab()
    conn = get_connection()
    r = conn.execute(
      """SELECT r.*, v.registration, v.capacity_kg, v.capacity_m3
        FROM routes r
        JOIN vehicles v ON v.id = r.vehicle_id
        WHERE r.id= ?""",
      (route_id,),
    ).fetchone()
    if not r:
      conn.close()
      raise ValueError(f"Tournée {route_id} introuvable")
    rd = _row_dict(r)

    stops = conn.execute(
      """SELECT rs.*, o.reference, o.quantity_kg, o.volume_m3, o.units_count,
           o.delivery_notes, o.access_instructions, c.name AS client_name
        FROM route_stops rs
        LEFT JOIN orders o ON o.id = rs.order_id
        LEFT JOIN clients c ON c.id = o.client_id
        WHERE rs.route_id= ? ORDER BY rs.stop_order""",
      (route_id,),
    ).fetchall()
    conn.close()

    cap_kg = float(rd.get("capacity_kg") or 1)
    tot_kg = sum(float(_row_dict(s).get("quantity_kg") or 0) for s in stops)
    tot_vol = sum(float(_row_dict(s).get("volume_m3") or 0) for s in stops)
    n_colis = sum(int(_row_dict(s).get("units_count") or 1) for s in stops)
    fill = min(100.0, (tot_kg / cap_kg) * 100) if cap_kg else 0

    story = self._load_manifest_story(rd, stops, tot_kg, tot_vol, n_colis, fill, cap_kg, lang=lang)
    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("load_manifest", {"route_id": route_id}, output_path)
    return output_path

  def generate_load_manifest_from_optimization_route(self, route_info: dict, output_path: str, lang: str = "fr") -> str:
    """Manifeste à partir du résultat VRP en mémoire."""
    _require_reportlab()
    veh = route_info.get("vehicle") or {}
    rd = {
      "registration": veh.get("registration", "—"),
      "planned_date": datetime.now().strftime("%Y-%m-%d"),
      "capacity_kg": float(veh.get("capacity_kg") or 1000),
      "capacity_m3": float(veh.get("capacity_m3") or 10),
    }
    stops = []
    for i, s in enumerate(route_info.get("route", []) or []):
      if s.get("type") != "delivery":
        continue
      c = s.get("client") or {}
      stops.append(
        _FakeRow(
          {
            "stop_order": i + 1,
            "reference": s.get("order_ref") or f"STOP-{i+1}",
            "quantity_kg": c.get("demand_kg", 0),
            "volume_m3": c.get("demand_m3", 0),
            "units_count": 1,
            "delivery_notes": c.get("notes") or "",
            "access_instructions": c.get("instructions") or "",
            "client_name": c.get("name") or "",
          }
        )
      )

    cap_kg = rd["capacity_kg"]
    tot_kg = sum(float(_row_dict(s).get("quantity_kg") or 0) for s in stops)
    tot_vol = sum(float(_row_dict(s).get("volume_m3") or 0) for s in stops)
    n_colis = sum(int(_row_dict(s).get("units_count") or 1) for s in stops)
    fill = min(100.0, (tot_kg / cap_kg) * 100) if cap_kg else 0

    story = self._load_manifest_story(rd, stops, tot_kg, tot_vol, n_colis, fill, cap_kg, lang=lang)
    doc = SimpleDocTemplate(output_path, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("load_manifest", {"source": "optimization_route"}, output_path)
    return output_path

  def _load_manifest_story(self, rd: dict, stops: list, tot_kg: float, tot_vol: float, n_colis: int, fill: float, cap_kg: float, lang: str = "fr") -> list:
    L = lambda k: _RL(lang, k)
    story = [
      Paragraph(
        f"<b><font color='#1E3A5F' size='13'>{L('manifest_title')}</font></b><br/>"
        f"<font size='9'>{L('label_vehicle')} : <b>{rd.get('registration')}</b> — {L('label_date')} : {rd.get('planned_date') or '—'}</font>",
        getSampleStyleSheet()["Normal"],
      ),
      Spacer(1, 0.35 * cm),
    ]
    data = [[L("col_num"), L("col_client"), L("col_reference"), L("col_qty"), L("col_weight_kg"), L("col_vol_m3"), L("col_instructions")]]
    for s in stops:
      d = _row_dict(s)
      data.append(
        [
          str(d.get("stop_order") or ""),
          str(d.get("client_name") or "")[:22],
          str(d.get("reference") or "—")[:14],
          str(d.get("units_count") or 1),
          f"{float(d.get('quantity_kg') or 0):.1f}",
          f"{float(d.get('volume_m3') or 0):.2f}",
          str((d.get("access_instructions") or d.get("delivery_notes") or "—"))[:40],
        ]
      )
    if len(data) == 1:
      data.append(["—", "—", "—", "—", "—", "—", L("msg_no_stop")])

    tbl = Table(data, repeatRows=1, colWidths=[0.9 * cm, 3.2 * cm, 2.2 * cm, 1.1 * cm, 1.8 * cm, 1.5 * cm, 4.3 * cm])
    tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("GRID", (0, 0), (-1, -1), 0.35, CLR_BORDER),
          ("FONTSIZE", (0, 0), (-1, -1), 7),
        ]
      )
    )
    story.append(tbl)
    story.append(Spacer(1, 0.35 * cm))

    cap_m3 = float(rd.get("capacity_m3") or 0)
    sum_tbl = Table(
      [
        [L("label_totals"), f"{L('label_weight')} : {tot_kg:.1f} kg", f"{L('label_volume')} : {tot_vol:.2f} m³", f"{L('label_packages')} : {n_colis}"],
        [
          L("label_fill_rate"),
          f"{fill:.1f} % (cap. {cap_kg:.0f} kg)",
          f"Cap. volume {cap_m3:.1f} m³" if cap_m3 else "—",
          "",
        ],
      ],
      colWidths=[4 * cm, 4.5 * cm, 4.5 * cm, 4 * cm],
    )
    sum_tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, -1), CLR_LIGHT),
          ("BOX", (0, 0), (-1, -1), 0.6, CLR_SECONDARY),
          ("FONTSIZE", (0, 0), (-1, -1), 9),
          ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ]
      )
    )
    story.append(sum_tbl)
    story.append(Spacer(1, 0.5 * cm))
    story.append(
      Paragraph(
        f"<b>{L('section_dock_check')}</b><br/><br/>"
        f"{L('manifest_certif')}<br/><br/>"
        f"{L('label_name')} : ___________________________ {L('label_signature')} : ___________________________ {L('label_date')} / heure : ___________",
        ParagraphStyle("mq", fontSize=9, leading=12),
      )
    )
    return story

  def generate_legal_notice_pdf(self, output_path: str, doc_type: str = "terms", lang: str = "fr") -> str:
    _require_reportlab()

    # ── Infos entreprise depuis settings.json ────────────────────────────────
    co: dict = {}
    try:
      _sp = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "settings.json"))
      with open(_sp, encoding="utf-8") as _f:
        co = json.load(_f).get("company", {})
    except Exception:
      pass
    co_name  = co.get("name")    or "CityPulse Logistics"
    co_addr  = co.get("address") or "—"
    co_phone = co.get("phone")   or "—"
    co_email = co.get("email")   or "—"
    now_str  = datetime.now().strftime("%d/%m/%Y")

    # ── Styles ───────────────────────────────────────────────────────────────
    s_title = ParagraphStyle("lt",  fontSize=16, textColor=CLR_PRIMARY,   spaceAfter=4,  fontName="Helvetica-Bold")
    s_sub   = ParagraphStyle("ls",  fontSize=10, textColor=CLR_SECONDARY, spaceAfter=14, fontName="Helvetica")
    s_art   = ParagraphStyle("la",  fontSize=11, textColor=CLR_PRIMARY,   spaceAfter=4,  spaceBefore=14, fontName="Helvetica-Bold")
    s_body  = ParagraphStyle("lb",  fontSize=9,  textColor=CLR_TEXT,      leading=14,    spaceAfter=4,   fontName="Helvetica")

    # ── Bloc en-tête entreprise ───────────────────────────────────────────────
    hdr = Table([[
      Paragraph(f"<b>{co_name}</b><br/>{co_addr}",
                ParagraphStyle("h1", fontSize=9, textColor=CLR_TEXT)),
      Paragraph(f"<b>Tél :</b> {co_phone}<br/><b>Email :</b> {co_email}<br/><b>Date :</b> {now_str}",
                ParagraphStyle("h2", fontSize=9, textColor=CLR_TEXT, alignment=TA_RIGHT)),
    ]], colWidths=[9 * cm, 8.5 * cm])
    hdr.setStyle(TableStyle([
      ("BOX",           (0, 0), (-1, -1), 0.5, CLR_PRIMARY),
      ("BACKGROUND",    (0, 0), (-1, -1), CLR_LIGHT),
      ("VALIGN",        (0, 0), (-1, -1), "TOP"),
      ("LEFTPADDING",   (0, 0), (-1, -1), 8),
      ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
      ("TOPPADDING",    (0, 0), (-1, -1), 8),
      ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    def art(num, titre): return Paragraph(f"Article {num} — {titre}", s_art)
    def body(txt):       return Paragraph(txt, s_body)
    hr = HRFlowable(width="100%", thickness=0.8, color=CLR_SECONDARY)

    story = []

    if doc_type == "terms":
      story += [
        Paragraph("Conditions Générales d'Utilisation", s_title),
        Paragraph(f"{co_name} — Version du {now_str}", s_sub),
        hdr, Spacer(1, 0.4 * cm), hr, Spacer(1, 0.3 * cm),
        art(1, "Objet"),
        body(f"Le présent document définit les conditions d'utilisation du logiciel de gestion et "
             f"d'optimisation de tournées exploité par <b>{co_name}</b>. Il s'applique à tout "
             f"utilisateur ayant accès au système (opérateur, planificateur, administrateur)."),
        art(2, "Accès au système"),
        body("L'accès est réservé aux personnes habilitées. Chaque utilisateur dispose d'identifiants "
             "personnels (login + mot de passe hashé). Le partage de compte est interdit. "
             "Toute compromission doit être signalée immédiatement à l'administrateur."),
        art(3, "Utilisation acceptable"),
        body("Le logiciel est utilisé exclusivement dans le cadre professionnel de l'entreprise. "
             "Toute utilisation à des fins personnelles, concurrentielles ou frauduleuses est interdite. "
             "L'utilisateur s'engage à ne pas tenter de contourner les mécanismes de sécurité."),
        art(4, "Données saisies et responsabilité"),
        body(f"Les données saisies (clients, commandes, coordonnées GPS, véhicules, tournées) restent "
             f"la propriété de <b>{co_name}</b>. L'utilisateur est responsable de l'exactitude "
             f"des informations qu'il saisit. L'éditeur décline toute responsabilité pour les "
             f"décisions prises sur la base de données erronées."),
        art(5, "Conservation des données"),
        body("Les données opérationnelles sont conservées en base SQLite locale selon la politique "
             "de rétention interne. Des sauvegardes peuvent être générées depuis "
             "Paramètres → Sauvegarde. Il est recommandé d'effectuer une sauvegarde hebdomadaire."),
        art(6, "Droits des utilisateurs"),
        body("Tout utilisateur peut demander la consultation, rectification ou suppression de ses "
             "données personnelles (nom, téléphone, email) auprès de l'administrateur système. "
             "Les logs d'audit ne peuvent être supprimés qu'en accord avec les obligations légales."),
        art(7, "Modifications"),
        body(f"Ces conditions peuvent être mises à jour par l'administrateur de <b>{co_name}</b>. "
             f"Les utilisateurs seront informés de toute modification substantielle."),
        art(8, "Contact"),
        body(f"Pour toute question relative à ce document :<br/>"
             f"<b>{co_name}</b> — {co_addr}<br/>"
             f"Tél : <b>{co_phone}</b> — Email : <b>{co_email}</b>"),
      ]
    else:  # privacy
      story += [
        Paragraph("Politique de Protection des Données Personnelles", s_title),
        Paragraph(f"{co_name} — Version du {now_str}", s_sub),
        hdr, Spacer(1, 0.4 * cm), hr, Spacer(1, 0.3 * cm),
        art(1, "Responsable du traitement"),
        body(f"Le responsable du traitement est <b>{co_name}</b>, dont le siège est situé à "
             f"<b>{co_addr}</b>. Contact : <b>{co_email}</b> — Tél : <b>{co_phone}</b>."),
        art(2, "Données collectées"),
        body("Dans le cadre de l'utilisation du logiciel, les données suivantes sont traitées :<br/>"
             "• <b>Clients :</b> nom, adresse de livraison, coordonnées GPS, téléphone, email ;<br/>"
             "• <b>Chauffeurs :</b> nom, prénom, numéro de permis, qualifications, photo ;<br/>"
             "• <b>Opérationnel :</b> commandes, tournées, horaires, distances, coûts, CO₂ ;<br/>"
             "• <b>Système :</b> logs d'audit, sessions utilisateurs, horodatages."),
        art(3, "Finalités du traitement"),
        body("Ces données sont traitées exclusivement pour :<br/>"
             "• Planification et optimisation des tournées de livraison ;<br/>"
             "• Gestion de la flotte de véhicules et des ressources humaines ;<br/>"
             "• Suivi opérationnel, reporting et facturation ;<br/>"
             "• Conformité réglementaire (CE 561/2006, ADR, ZFE)."),
        art(4, "Base légale"),
        body("Le traitement est fondé sur l'intérêt légitime de l'entreprise pour la gestion de son "
             "activité de transport, ainsi que sur les obligations légales applicables au secteur "
             "(réglementation CE 561/2006, ADR, droit du travail)."),
        art(5, "Durée de conservation"),
        body("• Données clients et commandes : durée de la relation commerciale + 5 ans ;<br/>"
             "• Données chauffeurs : durée du contrat + 5 ans ;<br/>"
             "• Logs d'audit : 3 ans ;<br/>"
             "• Données de géolocalisation (tournées) : 2 ans."),
        art(6, "Droits des personnes concernées"),
        body(f"Conformément à la réglementation applicable, toute personne dispose des droits "
             f"d'accès, rectification, effacement, opposition et portabilité. "
             f"Pour exercer ces droits : <b>{co_email}</b>."),
        art(7, "Sécurité des données"),
        body("Les données sont stockées localement dans une base SQLite. "
             "Les mots de passe sont hachés SHA-256 + sel. "
             "Les clés API tierces sont stockées dans le trousseau système (OS keyring). "
             "Des sauvegardes régulières sont recommandées (Paramètres → Sauvegarde)."),
        art(8, "Contact DPO"),
        body(f"Pour toute question relative à la protection des données :<br/>"
             f"<b>{co_name}</b> — {co_addr}<br/>"
             f"Email : <b>{co_email}</b> — Tél : <b>{co_phone}</b>"),
      ]

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2.5 * cm)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("legal_notice", {"doc_type": doc_type, "company": co_name}, output_path)
    log_action("LEGAL_DOC", f"Document légal '{doc_type}' généré pour {co_name}")
    return output_path


  def generate_legal_from_text(self, text: str, output_path: str) -> str:
    """Génère un PDF à partir du texte brut édité par l'utilisateur."""
    _require_reportlab()
    styles = getSampleStyleSheet()
    s_h1   = ParagraphStyle("lh1", fontSize=16, textColor=CLR_PRIMARY,   spaceAfter=4,  fontName="Helvetica-Bold")
    s_sub  = ParagraphStyle("lsb", fontSize=10, textColor=CLR_SECONDARY, spaceAfter=12, fontName="Helvetica")
    s_art  = ParagraphStyle("lar", fontSize=11, textColor=CLR_PRIMARY,   spaceAfter=4,  spaceBefore=14, fontName="Helvetica-Bold")
    s_bul  = ParagraphStyle("lbu", fontSize=9,  textColor=CLR_TEXT,      leading=14,    leftIndent=14, spaceAfter=2)
    s_body = ParagraphStyle("lbd", fontSize=9,  textColor=CLR_TEXT,      leading=14,    spaceAfter=4)

    story = []
    paragraphs = text.split("\n\n")
    for i, block in enumerate(paragraphs):
      lines = block.strip().splitlines()
      if not lines:
        story.append(Spacer(1, 0.2 * cm))
        continue
      first = lines[0].strip()
      if i == 0:
        story.append(Paragraph(first, s_h1))
        for sub in lines[1:]:
          story.append(Paragraph(sub.strip(), s_sub))
      elif first.startswith("Article "):
        story.append(Paragraph(first, s_art))
        for body_line in lines[1:]:
          bl = body_line.strip()
          if bl.startswith("•"):
            story.append(Paragraph(bl, s_bul))
          elif bl:
            story.append(Paragraph(bl, s_body))
      else:
        for bl in lines:
          bl = bl.strip()
          if bl.startswith("•"):
            story.append(Paragraph(bl, s_bul))
          elif bl:
            story.append(Paragraph(bl, s_body))
      story.append(Spacer(1, 0.1 * cm))

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2.5 * cm)
    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    _history_insert("legal_custom", {}, output_path)
    log_action("LEGAL_DOC", f"Document légal personnalisé généré : {output_path}")
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# Compatibilité ascendante (optimization_widget / ancien code)
# ══════════════════════════════════════════════════════════════════════════════


def _header_table(reg, v_type, cap, algo, now_str, lang: str = "fr"):
  L = lambda k: _RL(lang, k)
  styles = getSampleStyleSheet()
  header_data = [
    [
      Paragraph(
        f"<b><font color='#1E3A5F' size='16'>CityPulse Logistics</font></b><br/>"
        f"<font color='#3B82F6' size='11'>{L('route_sheet_title')}</font>",
        styles["Normal"],
      ),
      Paragraph(
        f"<b>{L('label_vehicle')} :</b> {reg}<br/>"
        f"<b>{L('label_type')} :</b> {v_type} | <b>{L('label_capacity')} :</b> {cap:.0f} kg<br/>"
        f"<b>{L('label_algo')} :</b> {algo}<br/>"
        f"<b>{L('label_generated_at')} :</b> {now_str}",
        ParagraphStyle("meta", fontSize=9, textColor=colors.HexColor("#374151")),
      ),
    ]
  ]
  tbl = Table(header_data, colWidths=[8 * cm, 9.5 * cm])
  tbl.setStyle(
    TableStyle(
      [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
      ]
    )
  )
  return tbl


def _add_page_number(canvas, doc):
  canvas.saveState()
  canvas.setFont("Helvetica", 8)
  canvas.setFillColor(CLR_MUTED)
  canvas.drawRightString(A4[0] - 1.8 * cm, 1.2 * cm, f"Page {doc.page}")
  canvas.restoreState()


def generate_route_pdf(route_info: dict, result: dict, output_path: str, lang: str = "fr") -> bool:
  """Ancienne API — retourne bool. Délègue à la logique historique."""
  if not REPORTLAB_OK:
    logger.error("reportlab non disponible")
    return False
  try:
    vehicle = route_info.get("vehicle", {})
    stops = route_info.get("route", [])
    reg = vehicle.get("registration", "Véhicule")
    v_type = vehicle.get("type", "fourgon")
    cap = vehicle.get("capacity_kg", 1000)
    algo = result.get("algorithm", "—")
    now_str = datetime.now().strftime("%d/%m/%Y à %H:%M")

    doc = SimpleDocTemplate(
      output_path,
      pagesize=A4,
      leftMargin=1.8 * cm,
      rightMargin=1.8 * cm,
      topMargin=2 * cm,
      bottomMargin=2 * cm,
      title=f"{_RL(lang, 'route_sheet_title')} — {reg}",
      author="CityPulse Logistics",
    )
    L = lambda k: _RL(lang, k)
    styles = getSampleStyleSheet()
    story = []
    story.append(_header_table(reg, v_type, cap, algo, now_str, lang=lang))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=CLR_SECONDARY, spaceAfter=0.3 * cm))

    dist = route_info.get("distance_km", 0)
    load = route_info.get("load_kg", 0)
    dur = route_info.get("duration_min", 0)
    cost = dist * float(vehicle.get("cost_per_km", 0.5) or 0.5)
    on_time = sum(1 for s in stops if s.get("delay", 0) == 0)
    resp = (on_time / len(stops) * 100) if stops else 0
    src = result.get("distance_source", "haversine")

    summary_data = [
      [L("col_metric"), L("col_value")],
      [L("row_total_dist"), f"{dist:.2f} km"],
      [L("row_est_duration"), f"{dur:.0f} min ({dur/60:.1f}h)"],
      [L("row_total_load"), f"{load:.0f} kg / {cap:.0f} kg ({(load/cap*100) if cap else 0:.0f}%)"],
      [L("row_est_cost"), f"{cost:.2f} €"],
      [L("section_planned_stops"), str(len(stops))],
      [L("row_time_respect"), f"{resp:.0f}% ({on_time}/{len(stops)})"],
      [L("row_dist_source"), L("val_osrm_source") if src == "osrm" else L("val_haversine_source")],
    ]
    summary_tbl = Table(summary_data, colWidths=[7 * cm, 9 * cm])
    summary_tbl.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), CLR_PRIMARY),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
          ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
          ("FONTSIZE", (0, 0), (-1, 0), 10),
          ("BACKGROUND", (0, 1), (0, -1), CLR_LIGHT),
          ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
          ("FONTSIZE", (0, 1), (-1, -1), 9),
          ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CLR_LIGHT]),
          ("GRID", (0, 0), (-1, -1), 0.5, CLR_BORDER),
          ("TOPPADDING", (0, 0), (-1, -1), 5),
          ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
          ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]
      )
    )
    story.append(summary_tbl)
    story.append(Spacer(1, 0.5 * cm))
    story.append(
      Paragraph(
        f"<b>{L('section_stop_detail')}</b>",
        ParagraphStyle(
          "h2", fontSize=12, textColor=CLR_PRIMARY, spaceAfter=6, fontName="Helvetica-Bold"
        ),
      )
    )

    stops_header = [L("col_num"), L("col_client"), L("col_time_window"), L("col_arrival"), L("col_departure"), L("col_load_kg_paren"), L("col_status")]
    stops_data = [stops_header]

    def fmt_min(m):
      h, mn = divmod(int(m), 60)
      return f"{h:02d}:{mn:02d}"

    for i, stop in enumerate(stops, 1):
      c = stop.get("client", {})
      name = c.get("name", f"Client {i}")[:28]
      ready = c.get("ready_time", 0)
      due = c.get("due_time", 1440)
      arrival = stop.get("arrival_time", 0)
      depart = stop.get("departure_time", 0)
      demand = c.get("demand_kg", 0)
      delay = stop.get("delay", 0)
      status = f"+{delay:.0f}min" if delay > 0 else "OK"
      stops_data.append(
        [
          str(i),
          name,
          f"{fmt_min(ready)} – {fmt_min(due)}",
          fmt_min(arrival),
          fmt_min(depart),
          f"{demand:.0f}",
          status,
        ]
      )

    col_w = [0.8 * cm, 5.5 * cm, 3.2 * cm, 1.9 * cm, 1.9 * cm, 2.0 * cm, 2.2 * cm]
    stops_tbl = Table(stops_data, colWidths=col_w, repeatRows=1)
    tbl_style = [
      ("BACKGROUND", (0, 0), (-1, 0), CLR_SECONDARY),
      ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
      ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
      ("FONTSIZE", (0, 0), (-1, 0), 8),
      ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CLR_LIGHT]),
      ("FONTSIZE", (0, 1), (-1, -1), 8),
      ("GRID", (0, 0), (-1, -1), 0.4, CLR_BORDER),
      ("ALIGN", (0, 0), (0, -1), "CENTER"),
      ("ALIGN", (2, 0), (-1, -1), "CENTER"),
      ("TOPPADDING", (0, 0), (-1, -1), 4),
      ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
      ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]
    for row_i, stop in enumerate(stops, 1):
      if stop.get("delay", 0) > 0:
        tbl_style.append(("TEXTCOLOR", (6, row_i), (6, row_i), CLR_DANGER))
        tbl_style.append(("FONTNAME", (6, row_i), (6, row_i), "Helvetica-Bold"))
      else:
        tbl_style.append(("TEXTCOLOR", (6, row_i), (6, row_i), CLR_SUCCESS))
    stops_tbl.setStyle(TableStyle(tbl_style))
    story.append(stops_tbl)
    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=CLR_BORDER))
    story.append(Spacer(1, 0.2 * cm))
    story.append(
      Paragraph(
        f"<font color='#6B7280' size='8'>{L('footer_generated_by')} — "
        f"{now_str} | {L('label_algo')} : {algo} | {L('footer_confidential')}</font>",
        ParagraphStyle("footer", fontSize=8, textColor=CLR_MUTED, alignment=TA_CENTER),
      )
    )
    doc.build(story, onFirstPage=_add_page_number, onLaterPages=_add_page_number)
    return True
  except Exception:
    logger.exception("Erreur génération PDF")
    return False


def generate_all_vehicles_pdf(result: dict, output_dir: str) -> list:
  paths = []
  for i, route_info in enumerate(result.get("routes", [])):
    if not route_info.get("route"):
      continue
    reg = (route_info.get("vehicle") or {}).get("registration", f"V{i+1}")
    safe = reg.replace("/", "_").replace("\\", "_").replace(" ", "_")
    path = os.path.join(output_dir, f"feuille_route_{safe}.pdf")
    if generate_route_pdf(route_info, result, path):
      paths.append(path)
  return paths
